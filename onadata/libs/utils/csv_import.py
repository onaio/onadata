# -*- coding: utf-8 -*-
"""
CSV data import module.
"""
import functools
import sys
import uuid
from builtins import str as text
from collections import defaultdict
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.storage import default_storage
from django.utils import timezone

import unicodecsv as ucsv
from celery import current_task
from celery.backends.rpc import BacklogLimitExceeded
from celery.result import AsyncResult
from dateutil.parser import parse
from multidb.pinning import use_master
from openpyxl import load_workbook
from six import iteritems

from onadata.apps.logger.models import Instance, XForm
from onadata.apps.messaging.constants import SUBMISSION_DELETED, XFORM
from onadata.apps.messaging.serializers import send_message
from onadata.celeryapp import app
from onadata.libs.serializers.metadata_serializer import MetaDataSerializer
from onadata.libs.utils import analytics
from onadata.libs.utils.async_status import FAILED, async_status, celery_state_to_status
from onadata.libs.utils.common_tags import (
    EXCEL_TRUE,
    IMPORTED_VIA_CSV_BY,
    INSTANCE_CREATE_EVENT,
    INSTANCE_UPDATE_EVENT,
    MULTIPLE_SELECT_TYPE,
    NA_REP,
    UUID,
    XLS_DATE_FIELDS,
    XLS_DATETIME_FIELDS,
)
from onadata.libs.utils.common_tools import report_exception
from onadata.libs.utils.dict_tools import csv_dict_to_nested_dict
from onadata.libs.utils.logger_tools import (
    OpenRosaResponse,
    dict2xml,
    safe_create_instance,
)

DEFAULT_UPDATE_BATCH = 100
PROGRESS_BATCH_UPDATE = getattr(
    settings, "EXPORT_TASK_PROGRESS_UPDATE_BATCH", DEFAULT_UPDATE_BATCH
)
IGNORED_COLUMNS = ["formhub/uuid", "meta/instanceID"]

# pylint: disable=invalid-name
User = get_user_model()


def get_submission_meta_dict(xform, instance_id):
    """Generates metadata for our submission

    Checks if `instance_id` belongs to an existing submission.
    If it does, it's considered an edit and its uuid gets deprecated.
    In either case, a new one is generated and assigned.

    :param onadata.apps.logger.models.XForm xform: The submission's XForm.
    :param string instance_id: The submission/instance `uuid`.

    :return: The metadata dict
    :rtype:  dict
    """
    uuid_arg = instance_id or f"uuid:{uuid.uuid4()}"
    meta = {"instanceID": uuid_arg}

    update = 0

    if (
        instance_id
        and xform.instances.filter(uuid=instance_id.replace("uuid:", "")).count() > 0
    ):
        uuid_arg = f"uuid:{uuid.uuid4()}"
        meta.update({"instanceID": uuid_arg, "deprecatedID": instance_id})
        update += 1
    return [meta, update]


def dict2xmlsubmission(submission_dict, xform, instance_id, submission_date):
    """Creates and xml submission from an appropriate dict (& other data)

    :param dict submission_dict: A dict containing form submission data.
    :param onadata.apps.logger.models.XForm xfrom: The submission's XForm.
    :param string instance_id: The submission/instance `uuid`.
    :param string submission_date: An isoformatted datetime string.

    :return: An xml submission string
    :rtype: string
    """

    xform_dict = xform.json_dict()
    return (
        '<?xml version="1.0" ?>'
        '<{0} id="{1}" instanceID="uuid:{2}" submissionDate="{3}">{4}'
        "</{0}>".format(
            xform_dict.get("name", xform.id_string),
            xform.id_string,
            instance_id,
            submission_date,
            dict2xml(submission_dict).replace("\n", ""),
        )
    ).encode("utf-8")


def dict_merge(part_a, part_b):
    """Returns a merger of two dicts a and b

    credits: https://www.xormedia.com/recursively-merge-dictionaries-in-python

    :param dict a: The "Part A" dict
    :param dict b: The "Part B" dict
    :return: The merger
    :rtype: dict
    """
    if not isinstance(part_b, dict):
        return part_b
    result = deepcopy(part_a)
    for k, v in iteritems(part_b):
        if k in result and isinstance(result[k], dict):
            result[k] = dict_merge(result[k], v)
        else:
            result[k] = deepcopy(v)
    return result


def dict_pathkeys_to_nested_dicts(dictionary):
    """Turns a flat dict to a nested dict

    Takes a dict with pathkeys or "slash-namespaced" keys and inflates
    them into nested dictionaries i.e:-
    `d['/path/key/123']` -> `d['path']['key']['123']`

    :param dict dictionary: A dict with one or more "slash-namespaced" keys
    :return: A nested dict
    :rtype: dict
    """
    data = dictionary.copy()
    for key in list(data):
        if r"/" in key:
            data = dict_merge(
                functools.reduce(
                    lambda v, k: {k: v}, (key.split("/") + [data.pop(key)])[::-1]
                ),
                data,
            )

    return data


@app.task()
def submit_csv_async(username, xform_id, file_path, overwrite=False):
    """Imports CSV data to an existing xform asynchrounously."""
    xform = XForm.objects.get(pk=xform_id)

    with default_storage.open(file_path) as csv_file:
        return submit_csv(username, xform, csv_file, overwrite)


def failed_import(rollback_uuids, xform, exception, status_message):
    """Report a failed import.
    :param rollback_uuids: The rollback UUIDs
    :param xform: The XForm that failed to import to
    :param exception: The exception object
    :return: The async_status result
    """
    Instance.objects.filter(uuid__in=rollback_uuids, xform=xform).delete()
    xform.submission_count(True)
    report_exception(
        f"CSV Import Failed : {xform.pk} - {xform.id_string} - {xform.title}",
        exception,
        sys.exc_info(),
    )
    return async_status(FAILED, status_message)


def validate_csv_file(csv_file, xform):
    """Validates a CSV File

    Takes a CSV Formatted file or sring containing rows of submission
    data and validates that the file is valid enough to be processed.

    :param (str or file) csv_file: A CSV formatted file or string with
    submission rows
    :param onadata.apps.logger.models.XForm xform: The submission's XForm
    :rtype: Dict
    :returns: A dict containing the validity of the CSV file as well as
    additional columns(additional_col) if any when successful else
    it returns an error message(error_msg)
    """
    # Validate csv_file is utf-8 encoded or unicode
    if isinstance(csv_file, str):
        csv_file = BytesIO(csv_file)
    elif csv_file is None or not hasattr(csv_file, "read"):
        return {
            "error_msg": (
                "Invalid param type for csv_file`."
                "Expected utf-8 encoded file or unicode"
                f" string got {type(csv_file).__name__} instead"
            ),
            "valid": False,
        }

    # Ensure stream position is at the start of the file
    csv_file.seek(0)

    # Retrieve CSV Headers from the CSV File
    csv_headers = ucsv.DictReader(csv_file, encoding="utf-8-sig").fieldnames

    # Make sure CSV headers have no spaces
    # because these are converted to XLSForm names
    # which cannot have spaces
    if any(" " in header for header in csv_headers):
        return {
            "error_msg": "CSV file fieldnames should not contain spaces",
            "valid": False,
        }

    # Get headers from stored data dictionary
    xform_headers = xform.get_headers()

    # Identify any missing columns between XForm and
    # imported CSV ignoring repeat and metadata columns
    missing_col = sorted(
        [
            col
            for col in set(xform_headers).difference(csv_headers)
            if col.find("[") == -1
            and not col.startswith("_")
            and col not in IGNORED_COLUMNS
            and "/_" not in col
        ]
    )

    mutliple_select_col = []

    # Find all multiple_select type columns and remove them from
    # the missing_col list
    for col in csv_headers:
        survey_element = xform.get_survey_element(col)
        if (
            survey_element
            and hasattr(survey_element, "type")
            and survey_element.get("type") == MULTIPLE_SELECT_TYPE
        ):
            # remove from the missing list
            missing_col = [x for x in missing_col if not x.startswith(col)]
            mutliple_select_col.append(col)

    if missing_col:
        return {
            "error_msg": (
                "Sorry uploaded file does not match the form. "
                "The file is missing the column(s): "
                f"{', '.join(missing_col)}."
            ),
            "valid": False,
        }

    # Identify any additional columns between XForm and
    # imported CSV ignoring repeat and multiple_select columns
    additional_col = [
        col
        for col in set(csv_headers).difference(xform_headers)
        if col.find("[") == -1 and col not in mutliple_select_col
    ]

    return {"valid": True, "additional_col": additional_col}


def flatten_split_select_multiples(
    row: Dict[str, Any], select_multiples: List[str]
) -> dict:
    """
    Flattens a select_multiple question that was previously split into
    different choice columns into one column
    """
    for key, value in row.items():
        if key in select_multiples and isinstance(value, dict):
            picked_choices = [
                k for k, v in value.items() if v.upper() in ["1", "TRUE"] or v == k
            ]
            new_value = " ".join(picked_choices)
            row.update({key: new_value})
        elif isinstance(value, dict):
            # Handle cases where select_multiples are within a group
            new_value = flatten_split_select_multiples(value, select_multiples)
            row.update({key: new_value})
        elif isinstance(value, list):
            # Handle case where we have repeat questions
            new_value = []

            for repeat_question in value:
                flattened_question = flatten_split_select_multiples(
                    repeat_question, select_multiples
                )
                new_value.append(flattened_question)

            row.update({key: new_value})

    return row


# pylint: disable=too-many-locals,too-many-branches,too-many-statements
@use_master
def submit_csv(username, xform, csv_file, overwrite=False):  # noqa
    """Imports CSV data to an existing form

    Takes a csv formatted file or string containing rows of submission/instance
    and converts those to xml submissions and finally submits them by calling
    :py:func:`onadata.libs.utils.logger_tools.safe_create_instance`

    :param str username: the submission user
    :param onadata.apps.logger.models.XForm xform: The submission's XForm.
    :param (str or file) csv_file: A CSV formatted file with submission rows.
    :return: If sucessful, a dict with import summary else dict with error str.
    :rtype: Dict
    """
    csv_file_validation_summary = validate_csv_file(csv_file, xform)

    if csv_file_validation_summary.get("valid"):
        additional_col = csv_file_validation_summary.get("additional_col")
    else:
        return async_status(FAILED, csv_file_validation_summary.get("error_msg"))

    num_rows = sum(1 for row in csv_file) - 1

    # Change stream position to start of file
    csv_file.seek(0)

    csv_reader = ucsv.DictReader(csv_file, encoding="utf-8-sig")
    xform_json = xform.json_dict()
    select_multiples = [
        qstn.name for qstn in xform.get_survey_elements_of_type(MULTIPLE_SELECT_TYPE)
    ]
    ona_uuid = {"formhub": {"uuid": xform.uuid}}
    additions = duplicates = inserts = 0
    rollback_uuids = []
    errors = {}

    # Retrieve the columns we should validate values for
    # Currently validating date, datetime, integer and decimal columns
    col_to_validate = {
        "date": (get_columns_by_type(XLS_DATE_FIELDS, xform_json), parse),
        "datetime": (get_columns_by_type(XLS_DATETIME_FIELDS, xform_json), parse),
        "integer": (get_columns_by_type(["integer"], xform_json), int),
        "decimal": (get_columns_by_type(["decimal"], xform_json), float),
    }

    if overwrite:
        instance_ids = [i["id"] for i in xform.instances.values("id")]
        xform.instances.filter(deleted_at__isnull=True).update(
            deleted_at=timezone.now(), deleted_by=User.objects.get(username=username)
        )
        # updates the form count
        xform.submission_count(True)
        # send message
        send_message(
            instance_id=instance_ids,
            target_id=xform.id,
            target_type=XFORM,
            user=User.objects.get(username=username),
            message_verb=SUBMISSION_DELETED,
        )

    try:  # pylint: disable=too-many-nested-blocks
        for row_no, row in enumerate(csv_reader):
            # Remove additional columns
            for index in additional_col:
                del row[index]

            # Remove 'n/a' and '' values from csv
            row = {k: v for (k, v) in row.items() if v not in [NA_REP, ""]}

            row, error = validate_row(row, col_to_validate)

            if error:
                errors[row_no] = error

            # Only continue the process if no errors where encountered while
            # validating the data
            if not errors:
                location_data = {}

                for key in list(row):
                    # Collect row location data into separate location_data
                    # dict
                    if key.endswith(
                        (".latitude", ".longitude", ".altitude", ".precision")
                    ):
                        location_key, location_prop = key.rsplit(".", 1)
                        location_data.setdefault(location_key, {}).update(
                            {location_prop: row.get(key, "0")}
                        )

                # collect all location K-V pairs into single geopoint field(s)
                # in location_data dict
                for location_key in list(location_data):
                    geo_dict = defaultdict(lambda: "", location_data.get(location_key))
                    location_data.update(
                        {
                            location_key: (
                                f"{geo_dict['latitude']} {geo_dict['longitude']} "
                                f"{geo_dict['altitude']} {geo_dict['precision']}"
                            )
                        }
                    )

                nested_dict = csv_dict_to_nested_dict(
                    row, select_multiples=select_multiples
                )
                row = flatten_split_select_multiples(
                    nested_dict, select_multiples=select_multiples
                )
                location_data = csv_dict_to_nested_dict(location_data)
                # Merge location_data into the Row data
                row = dict_merge(row, location_data)

                submission_time = datetime.utcnow().isoformat()
                row_uuid = (
                    row.get("meta/instanceID") or f"uuid:{row.get(UUID)}"
                    if row.get(UUID) and not overwrite
                    else None
                )
                submitted_by = row.get("_submitted_by")
                submission_date = row.get("_submission_time", submission_time)

                for key in list(row):
                    # remove metadata (keys starting with '_')
                    if key.startswith("_"):
                        del row[key]

                # Inject our forms uuid into the submission
                row.update(ona_uuid)

                old_meta = row.get("meta", {})
                new_meta, update = get_submission_meta_dict(xform, row_uuid)
                inserts += update
                old_meta.update(new_meta)
                row.update({"meta": old_meta})

                row_uuid = row.get("meta").get("instanceID")
                rollback_uuids.append(row_uuid.replace("uuid:", ""))

                try:
                    xml_file = BytesIO(
                        dict2xmlsubmission(row, xform, row_uuid, submission_date)
                    )

                    try:
                        error, instance = safe_create_instance(
                            username,
                            xml_file,
                            [],
                            xform.uuid,
                            None,
                            instance_status="imported_via_csv",
                        )
                    except ValueError as e:
                        error = e

                    if error:
                        if not (
                            isinstance(error, OpenRosaResponse)
                            and error.status_code == 202
                        ):
                            Instance.objects.filter(
                                uuid__in=rollback_uuids, xform=xform
                            ).delete()
                            return async_status(FAILED, text(error))
                        duplicates += 1
                    else:
                        additions += 1

                        if additions % PROGRESS_BATCH_UPDATE == 0:
                            try:
                                current_task.update_state(
                                    state="PROGRESS",
                                    meta={
                                        "progress": additions,
                                        "total": num_rows,
                                        "info": additional_col,
                                    },
                                )
                            except AttributeError:
                                pass
                            finally:
                                xform.submission_count(True)

                        users = (
                            User.objects.filter(username=submitted_by)
                            if submitted_by
                            else []
                        )

                        # Store user who imported data in metadata
                        serializer = MetaDataSerializer(
                            data={
                                "instance": instance.id,
                                "data_type": IMPORTED_VIA_CSV_BY,
                                "data_value": username,
                            }
                        )

                        if serializer.is_valid():
                            serializer.save()

                        if users:
                            instance.user = users[0]
                            instance.save()
                except Exception as e:  # pylint: disable=broad-except
                    return failed_import(rollback_uuids, xform, e, text(e))
    except UnicodeDecodeError as e:
        return failed_import(rollback_uuids, xform, e, "CSV file must be utf-8 encoded")

    if errors:
        # Rollback all created instances if an error occurred during
        # validation
        Instance.objects.filter(uuid__in=rollback_uuids, xform=xform).delete()
        xform.submission_count(True)
        return async_status(
            FAILED,
            f"Invalid CSV data imported in row(s): {errors}",
        )

    xform.submission_count(True)
    added_submissions = additions - inserts
    event_by = User.objects.get(username=username)
    event_name = None
    tracking_properties = {
        "xform_id": xform.pk,
        "project_id": xform.project.pk,
        "submitted_by": event_by,
        "label": f"csv-import-for-form-{xform.pk}",
        "from": "CSV Import",
    }
    if added_submissions > 0:
        tracking_properties["value"] = added_submissions
        event_name = INSTANCE_CREATE_EVENT
        analytics.track(event_by, event_name, properties=tracking_properties)

    if inserts > 0:
        tracking_properties["value"] = inserts
        event_name = INSTANCE_UPDATE_EVENT
        analytics.track(event_by, event_name, properties=tracking_properties)

    additional_columns = ", ".join(list(additional_col))
    return {
        "additions": added_submissions,
        "duplicates": duplicates,
        "updates": inserts,
        "info": (
            "Additional column(s) excluded from the upload: " f"'{additional_columns}'."
        ),
    }


def get_async_csv_submission_status(job_uuid):
    """Gets CSV Submision progress or result
    Can be used to pol long running submissions
    :param str job_uuid: The submission job uuid returned by _submit_csv.delay
    :return: Dict with import progress info (insertions & total)
    :rtype: Dict
    """
    if not job_uuid:
        return async_status(FAILED, "Empty job uuid")

    job = AsyncResult(job_uuid)
    try:
        # result = (job.result or job.state)
        if job.state not in ["SUCCESS", "FAILURE"]:
            response = async_status(celery_state_to_status(job.state))
            if isinstance(job.info, dict):
                response.update(job.info)

            return response

        if job.state == "FAILURE":
            return async_status(celery_state_to_status(job.state), text(job.result))

    except BacklogLimitExceeded:
        return async_status(celery_state_to_status("PENDING"))

    return job.get()


def submission_xls_to_csv(xls_file):  # noqa
    """Convert a submission xls file to submissions csv file

    :param xls_file: submissions xls file
    :return: csv_file
    """
    xls_file.seek(0)
    xl_workbook = load_workbook(filename=xls_file)
    sheet_name = xl_workbook.get_sheet_names()[0]
    first_sheet = xl_workbook.get_sheet_by_name(sheet_name)

    csv_file = BytesIO()
    csv_writer = ucsv.writer(csv_file)

    date_columns = []
    boolean_columns = []

    # write the header
    csv_writer.writerow(list(first_sheet.values)[0])

    # check for any dates or boolean in the first row of data
    for index in range(1, first_sheet.max_column + 1):
        row = 1

        # If the field is not required the first row may have
        # a null and thus XLS Dates (floats) or XLS Booleans
        # will not be properly converted in the next steps,
        # therefore we find the first non-empty row.
        while (
            first_sheet.cell(row, index).value is not None
            and row < first_sheet.max_row - 1
        ):
            row += 1

        if first_sheet.cell(row, index).is_date:
            date_columns.append(index)
        elif first_sheet.cell(row, index).data_type == "b":
            boolean_columns.append(index)

    for row_values in first_sheet.iter_rows(
        min_row=2, max_row=first_sheet.max_row, values_only=True
    ):
        if not any(row_values):
            continue
        # convert excel dates(floats) to datetime
        for date_column_index in date_columns:
            try:
                row_values[date_column_index - 1] = row_values[
                    date_column_index - 1
                ].isoformat()
            except (ValueError, TypeError):
                pass

        # convert excel boolean to true/false
        for boolean_column_index in boolean_columns:
            row_values[boolean_column_index - 1] = bool(
                row_values[boolean_column_index - 1] == EXCEL_TRUE
            )

        csv_writer.writerow(row_values)

    return csv_file


def get_columns_by_type(type_list, form_json):
    """Returns a column that match types passed within
    the field_list

    :param list type_list: A list containing strings that represent
    XLS field types
    :param dict form_json: A dict representing the contents of a
    form
    :return: A list containing the column names that store values passed
    within the type_list
    :rtype: list
    """

    def _column_by_type(item_list, prefix=""):
        found = []
        for item in item_list:
            if item["type"] in ["group", "repeat"]:
                prefix = "/".join([prefix, item["name"]]) if prefix else item["name"]
                found.extend(_column_by_type(item["children"], prefix))
                prefix = ""  # Reset prefix to blank
            else:
                if item["type"] in type_list:
                    name = f"{prefix}/{item['name']}" if prefix else item["name"]
                    found.append(name)

        return found

    return _column_by_type(form_json["children"])


def validate_row(row, columns):
    """Validate row of data making sure column constraints are enforced

    Takes a csv row containing data from a submission and validates the
    data making sure data types are enforced.

    :param Dict row: The csv row
    :param Dict columns: A dict containing the column headers to be validated
    and the function that should check that the columns constraint is not
    broken
    :return: Returns a tuple containing the validated row and errors found
    within the row if any
    :rtype: Tuple
    """
    # Check data doesn't infringe on XForm data constraints
    errors = []
    for datatype in columns:
        column, constraint_check = columns.get(datatype)
        valid, data = validate_column(row, column, constraint_check)

        if valid:
            if datatype in ["date", "datetime"]:
                for key in data:
                    # ODK XForms accept date and datetime values formatted in
                    # accordance to the XML 1.0 spec only deviating when it
                    # comes to the inclusion of the timezone offset in
                    # datetime values. This specification matches ISO 8601
                    value = data.get(key).isoformat()

                    if datatype == "date" and value.endswith("T00:00:00"):
                        # Remove the Time string section from dates
                        # to follow along with the date datetype specification
                        # in the ODK XForms Spec
                        value = value.replace("T00:00:00", "")

                    data.update({key: value})

            row.update(data)
        else:
            errors.append(f"Unknown {datatype} format(s): {', '.join(data)}")

    return (row, errors)


def validate_column(row, columns, constraint_check):
    """Validates columns within a row making sure data constraints are
    not broken

    Takes a list of column headers to validate and a function of which
    is used to validate the data is valid.

    :param Dict row: The row of which the columns values should be validated
    :param List columns: A list of headers to be validated within the row.
    :param func constraint_check: A function used to validate column data
    :return: Returns a tuple containing the validity status of the rows
    data and the validated_data if successful else it contains
    the invalid_data
    """
    invalid_data = []
    validated_data = {}

    for key in columns:
        value = row.get(key, "")

        if value:
            try:
                value = constraint_check(value)
            except ValueError:
                invalid_data.append(value)
            else:
                validated_data[key] = value

    return (False, invalid_data) if invalid_data else (True, validated_data)
