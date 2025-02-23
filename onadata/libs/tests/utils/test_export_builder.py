# -*- coding: utf-8 -*-
"""
Tests Export Builder Functionality
"""

from __future__ import unicode_literals

import csv
import datetime
import os
import shutil
import tempfile
import zipfile
from collections import OrderedDict
from ctypes import ArgumentError
from io import BytesIO
from unittest.mock import patch

from django.conf import settings
from django.core.files.temp import NamedTemporaryFile

from openpyxl import load_workbook
from pyxform.builder import create_survey_from_xls
from savReaderWriter import SavHeaderReader, SavReader

from onadata.apps.logger.import_tools import django_file
from onadata.apps.logger.models import DataView
from onadata.apps.main.tests.test_base import TestBase
from onadata.apps.viewer.models.data_dictionary import DataDictionary
from onadata.apps.viewer.models.parsed_instance import _encode_for_mongo, query_data
from onadata.apps.viewer.tests.export_helpers import viewer_fixture_path
from onadata.libs.utils.common_tags import (
    MULTIPLE_SELECT_TYPE,
    REVIEW_COMMENT,
    REVIEW_DATE,
    REVIEW_STATUS,
    SELECT_BIND_TYPE,
)
from onadata.libs.utils.csv_builder import CSVDataFrameBuilder, get_labels_from_columns
from onadata.libs.utils.export_builder import (
    ExportBuilder,
    decode_mongo_encoded_section_names,
    dict_to_joined_export,
    string_to_date_with_xls_validation,
)
from onadata.libs.utils.export_tools import get_columns_with_hxl
from onadata.libs.utils.logger_tools import create_instance


def _logger_fixture_path(*args):
    return os.path.join(
        settings.PROJECT_ROOT, "apps", "logger", "tests", "fixtures", *args
    )


def _str_if_bytes(val):
    """Returns val as string if it is of type bytes otherwise returns bytes"""
    return str(val, "utf-8") if isinstance(val, bytes) else val


class TestExportBuilder(TestBase):
    """Test onadata.libs.utils.export_builder functions."""

    data = [
        {
            "name": "Abe",
            "age": 35,
            "tel/telLg==office": "020123456",
            "_review_status": "Rejected",
            "_review_comment": "Wrong Location",
            REVIEW_DATE: "2021-05-25T02:27:19",
            "children": [
                {
                    "children/name": "Mike",
                    "children/age": 5,
                    "children/fav_colors": "red blue",
                    "children/iceLg==creams": "vanilla chocolate",
                    "children/cartoons": [
                        {
                            "children/cartoons/name": "Tom & Jerry",
                            "children/cartoons/why": "Tom is silly",
                        },
                        {
                            "children/cartoons/name": "Flinstones",
                            "children/cartoons/why": "I like bam bam\u0107",
                            # throw in a unicode character
                        },
                    ],
                },
                {"children/name": "John", "children/age": 2, "children/cartoons": []},
                {
                    "children/name": "Imora",
                    "children/age": 3,
                    "children/cartoons": [
                        {
                            "children/cartoons/name": "Shrek",
                            "children/cartoons/why": "He's so funny",
                        },
                        {
                            "children/cartoons/name": "Dexter's Lab",
                            "children/cartoons/why": "He thinks hes smart",
                            "children/cartoons/characters": [
                                {
                                    "children/cartoons/characters/name": "Dee Dee",
                                    "children/cartoons/characters/good_or_evi"
                                    "l": "good",
                                },
                                {
                                    "children/cartoons/characters/name": "Dexter",
                                    "children/cartoons/characters/good_or_evi"
                                    "l": "evil",
                                },
                            ],
                        },
                    ],
                },
            ],
        },
        {
            # blank data just to be sure
            "children": []
        },
    ]
    long_survey_data = [
        {
            "name": "Abe",
            "age": 35,
            "childrens_survey_with_a_very_lo": [
                {
                    "childrens_survey_with_a_very_lo/name": "Mike",
                    "childrens_survey_with_a_very_lo/age": 5,
                    "childrens_survey_with_a_very_lo/fav_colors": "red blue",
                    "childrens_survey_with_a_very_lo/cartoons": [
                        {
                            "childrens_survey_with_a_very_lo/cartoons/name": "Tom & Jerry",
                            "childrens_survey_with_a_very_lo/cartoons/why": "Tom is silly",
                        },
                        {
                            "childrens_survey_with_a_very_lo/cartoons/name": "Flinstones",
                            "childrens_survey_with_a_very_lo/cartoons/why": "I like bam bam\u0107",
                            # throw in a unicode character
                        },
                    ],
                },
                {
                    "childrens_survey_with_a_very_lo/name": "John",
                    "childrens_survey_with_a_very_lo/age": 2,
                    "childrens_survey_with_a_very_lo/cartoons": [],
                },
                {
                    "childrens_survey_with_a_very_lo/name": "Imora",
                    "childrens_survey_with_a_very_lo/age": 3,
                    "childrens_survey_with_a_very_lo/cartoons": [
                        {
                            "childrens_survey_with_a_very_lo/cartoons/name": "Shrek",
                            "childrens_survey_with_a_very_lo/cartoons/why": "He's so funny",
                        },
                        {
                            "childrens_survey_with_a_very_lo/cartoons/name": "Dexter's Lab",
                            "childrens_survey_with_a_very_lo/cartoons/why": "He thinks hes smart",
                            "childrens_survey_with_a_very_lo/cartoons/characters": [
                                {
                                    "childrens_survey_with_a_very_lo/cartoons/"
                                    "characters/name": "Dee Dee",
                                    "children/cartoons/characters/good_or_evi"
                                    "l": "good",
                                },
                                {
                                    "childrens_survey_with_a_very_lo/cartoons/"
                                    "characters/name": "Dexter",
                                    "children/cartoons/characters/good_or_evi"
                                    "l": "evil",
                                },
                            ],
                        },
                    ],
                },
            ],
        }
    ]
    data_utf8 = [
        {
            "name": "Abe",
            "age": 35,
            "tel/telLg==office": "020123456",
            "childrenLg==info": [
                {
                    "childrenLg==info/nameLg==first": "Mike",
                    "childrenLg==info/age": 5,
                    "childrenLg==info/fav_colors": "red's blue's",
                    "childrenLg==info/ice_creams": "vanilla chocolate",
                    "childrenLg==info/cartoons": [
                        {
                            "childrenLg==info/cartoons/name": "Tom & Jerry",
                            "childrenLg==info/cartoons/why": "Tom is silly",
                        },
                        {
                            "childrenLg==info/cartoons/name": "Flinstones",
                            "childrenLg==info/cartoons/why":
                            # throw in a unicode character
                            "I like bam bam\u0107",
                        },
                    ],
                }
            ],
        }
    ]
    osm_data = [
        {
            "photo": "1424308569120.jpg",
            "osm_road": "OSMWay234134797.osm",
            "osm_building": "OSMWay34298972.osm",
            "fav_color": "red",
            "osm_road:ctr:lat": "23.708174238006087",
            "osm_road:ctr:lon": "90.40946505581161",
            "osm_road:highway": "tertiary",
            "osm_road:lanes": 2,
            "osm_road:name": "Patuatuli Road",
            "osm_building:building": "yes",
            "osm_building:building:levels": 4,
            "osm_building:ctr:lat": "23.707316084046038",
            "osm_building:ctr:lon": "90.40849938337506",
            "osm_building:name": "kol",
            "_review_status": "Rejected",
            "_review_comment": "Wrong Location",
            REVIEW_DATE: "2021-05-25T02:27:19",
        }
    ]

    def _create_childrens_survey(self, filename="childrens_survey.xlsx"):
        survey = create_survey_from_xls(
            _logger_fixture_path(filename), default_name=filename.split(".")[0]
        )
        self.data_dictionary = DataDictionary()
        setattr(self.data_dictionary, "_survey", survey)
        return survey

    # pylint: disable=invalid-name
    def test_build_sections_for_multilanguage_form(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("multi_lingual_form.xlsx"),
            default_name="multi_lingual_form",
        )

        # check the default langauge
        self.assertEqual(survey.to_json_dict().get("default_language"), "English")
        export_builder = ExportBuilder()
        export_builder.INCLUDE_LABELS_ONLY = True
        export_builder.set_survey(survey)
        expected_sections = [survey.name]
        self.assertEqual(
            expected_sections, [s["name"] for s in export_builder.sections]
        )
        expected_element_names = [
            "Name of respondent",
            "Age",
            "Sex of respondent",
            "Fruits",
            "Fruits/Apple",
            "Fruits/Banana",
            "Fruits/Pear",
            "Fruits/Mango",
            "Fruits/Other",
            "Fruits/None of the above",
            "Cities",
            "meta/instanceID",
        ]
        section = export_builder.section_by_name(survey.name)
        element_names = [element["label"] for element in section["elements"]]
        self.assertEqual(sorted(expected_element_names), sorted(element_names))

        export_builder.language = "French"
        export_builder.set_survey(survey)
        section = export_builder.section_by_name(survey.name)
        element_names = [element["label"] for element in section["elements"]]
        expected_element_names = [
            "Des fruits",
            "Fruits/Aucune de ces réponses",
            "Fruits/Autre",
            "Fruits/Banane",
            "Fruits/Mangue",
            "Fruits/Poire",
            "Fruits/Pomme",
            "L'age",
            "Le genre",
            "Nom de personne interrogée",
            "Villes",
            "meta/instanceID",
        ]
        self.assertEqual(sorted(expected_element_names), sorted(element_names))

        # use default language when the language passed does not exist
        export_builder.language = "Kiswahili"
        export_builder.set_survey(survey)
        expected_element_names = [
            "Name of respondent",
            "Age",
            "Sex of respondent",
            "Fruits",
            "Fruits/Apple",
            "Fruits/Banana",
            "Fruits/Pear",
            "Fruits/Mango",
            "Fruits/Other",
            "Fruits/None of the above",
            "Cities",
            "meta/instanceID",
        ]
        section = export_builder.section_by_name(survey.name)
        element_names = [element["label"] for element in section["elements"]]
        self.assertEqual(sorted(expected_element_names), sorted(element_names))

    def test_build_sections_from_survey(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # test that we generate the proper sections
        expected_sections = [
            survey.name,
            "children",
            "children/cartoons",
            "children/cartoons/characters",
        ]
        self.assertEqual(
            expected_sections, [s["name"] for s in export_builder.sections]
        )
        # main section should have split geolocations
        expected_element_names = [
            "name",
            "age",
            "geo/geolocation",
            "geo/_geolocation_longitude",
            "geo/_geolocation_latitude",
            "geo/_geolocation_altitude",
            "geo/_geolocation_precision",
            "tel/tel.office",
            "tel/tel.mobile",
            "meta/instanceID",
        ]
        section = export_builder.section_by_name(survey.name)
        element_names = [element["xpath"] for element in section["elements"]]
        # fav_colors should have its choices split
        self.assertEqual(sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            "children/name",
            "children/age",
            "children/fav_colors",
            "children/fav_colors/red",
            "children/fav_colors/blue",
            "children/fav_colors/pink",
            "children/ice.creams",
            "children/ice.creams/vanilla",
            "children/ice.creams/strawberry",
            "children/ice.creams/chocolate",
        ]
        section = export_builder.section_by_name("children")
        element_names = [element["xpath"] for element in section["elements"]]
        self.assertEqual(sorted(expected_element_names), sorted(element_names))

        expected_element_names = ["children/cartoons/name", "children/cartoons/why"]
        section = export_builder.section_by_name("children/cartoons")
        element_names = [element["xpath"] for element in section["elements"]]

        self.assertEqual(sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            "children/cartoons/characters/name",
            "children/cartoons/characters/good_or_evil",
        ]
        section = export_builder.section_by_name("children/cartoons/characters")
        element_names = [element["xpath"] for element in section["elements"]]
        self.assertEqual(sorted(expected_element_names), sorted(element_names))

    # pylint: disable=too-many-locals
    def test_zipped_csv_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(d, index, indices, survey_name, survey, d, None)
            )
            index += 1

        # check that each file exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, f"{survey.name}.csv")))
        with open(
            os.path.join(temp_dir, f"{survey.name}.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)

            # open comparison file
            with open(
                _logger_fixture_path("csvs", "childrens_survey.csv"), encoding="utf-8"
            ) as fixture_csv:
                expected_rows = list(csv.reader(fixture_csv))
                self.assertEqual(rows, expected_rows)

        self.assertTrue(os.path.exists(os.path.join(temp_dir, "children.csv")))
        with open(os.path.join(temp_dir, "children.csv"), encoding="utf-8") as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)

            # open comparison file
            with open(
                _logger_fixture_path("csvs", "children.csv"), encoding="utf-8"
            ) as fixture_csv:
                expected_rows = list(csv.reader(fixture_csv))
                self.assertEqual(rows, expected_rows)

        self.assertTrue(os.path.exists(os.path.join(temp_dir, "children_cartoons.csv")))
        with open(
            os.path.join(temp_dir, "children_cartoons.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)

            # open comparison file
            with open(
                _logger_fixture_path("csvs", "children_cartoons.csv"), encoding="utf-8"
            ) as fixture_csv:
                expected_rows = list(csv.reader(fixture_csv))
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(os.path.join(temp_dir, "children_cartoons_characters.csv"))
        )
        with open(
            os.path.join(temp_dir, "children_cartoons_characters.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)

            # open comparison file
            with open(
                _logger_fixture_path("csvs", "children_cartoons_characters.csv"),
                encoding="utf-8",
            ) as fixture_csv:
                expected_rows = list(csv.reader(fixture_csv))
                self.assertEqual(rows, expected_rows)

        shutil.rmtree(temp_dir)

    def test_xlsx_export_with_osm_data(self):
        """
        Tests that osm data is included in xls export
        """
        survey = self._create_osm_survey()
        xform = self.xform
        export_builder = ExportBuilder()
        export_builder.set_survey(survey, xform)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.osm_data)
            temp_xls_file.seek(0)
            workbook = load_workbook(temp_xls_file.name)
            osm_data_sheet = workbook["osm"]
            rows = list(osm_data_sheet.rows)
            xls_headers = [a.value for a in rows[0]]

        expected_column_headers = [
            "photo",
            "osm_road",
            "osm_building",
            "fav_color",
            "form_completed",
            "meta/instanceID",
            "_id",
            "_uuid",
            "_submission_time",
            "_index",
            "_parent_table_name",
            "_parent_index",
            "_tags",
            "_notes",
            "_version",
            "_duration",
            "_submitted_by",
            "osm_road:ctr:lat",
            "osm_road:ctr:lon",
            "osm_road:highway",
            "osm_road:lanes",
            "osm_road:name",
            "osm_road:way:id",
            "osm_building:building",
            "osm_building:building:levels",
            "osm_building:ctr:lat",
            "osm_building:ctr:lon",
            "osm_building:name",
            "osm_building:way:id",
        ]
        self.assertEqual(sorted(expected_column_headers), sorted(xls_headers))

        submission = [a.value for a in rows[1]]
        self.assertEqual(submission[0], "1424308569120.jpg")
        self.assertEqual(submission[2], "23.708174238006087")
        self.assertEqual(submission[4], "tertiary")
        self.assertEqual(submission[6], "Patuatuli Road")
        self.assertEqual(submission[11], "23.707316084046038")
        self.assertEqual(submission[13], "kol")

    # pylint: disable=invalid-name
    def test_decode_mongo_encoded_section_names(self):
        data = {
            "main_section": [1, 2, 3, 4],
            "sectionLg==1/info": [1, 2, 3, 4],
            "sectionLg==2/info": [1, 2, 3, 4],
        }
        result = decode_mongo_encoded_section_names(data)
        expected_result = {
            "main_section": [1, 2, 3, 4],
            "section.1/info": [1, 2, 3, 4],
            "section.2/info": [1, 2, 3, 4],
        }
        self.assertEqual(result, expected_result)

    # pylint: disable=invalid-name
    def test_zipped_csv_export_works_with_unicode(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(
            os.path.join(temp_dir, "children.info.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = [
                "children.info/name.first",
                "children.info/age",
                "children.info/fav_colors",
                "children.info/fav_colors/red's",
                "children.info/fav_colors/blue's",
                "children.info/fav_colors/pink's",
                "children.info/ice_creams",
                "children.info/ice_creams/vanilla",
                "children.info/ice_creams/strawberry",
                "children.info/ice_creams/chocolate",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            rows = list(reader)
            actual_headers = list(rows[0])
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(data["children.info/fav_colors/red's"], "True")
            self.assertEqual(data["children.info/fav_colors/blue's"], "True")
            self.assertEqual(data["children.info/fav_colors/pink's"], "False")
            # check that red and blue are set to true

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_date_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | date              | expense_date | Expense Date |
        |        | begin group       | A            | A group      |
        |        | date              | gdate        | Good Day     |
        |        | end group         |              |              |

        | choices |
        |         | list name | name   | label  |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [
            {
                "expense_date": "2013-01-03",
                "A/gdate": "2017-06-13",
                "_submission_time": "2016-11-21T03:43:43.000-08:00",
            }
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            self.assertTrue(len(rows) > 1)
            self.assertEqual(_str_if_bytes(rows[0][0]), "expense_date")
            self.assertEqual(_str_if_bytes(rows[1][0]), "2013-01-03")
            self.assertEqual(_str_if_bytes(rows[0][1]), "A.gdate")
            self.assertEqual(_str_if_bytes(rows[1][1]), "2017-06-13")
            self.assertEqual(_str_if_bytes(rows[0][5]), "@_submission_time")
            self.assertEqual(_str_if_bytes(rows[1][5]), "2016-11-21 03:43:43")

        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_zipped_sav_export_dynamic_select_multiple(self):
        md = """
        | survey |
        |        | type                  | name           | label                         | calculation                       |
        |        | select_one sex        | sex            | Are they male or female?      |                                   |
        |        | calculate             | text           |                               | if((${sex}=’male’), 'his', ‘her’) |
        |        | text                  | favorite_brand | What's their favorite brand ? |                                   |
        |        | text                  | name           | What's ${text} name?          |                                   |
        |        | select_multiple brand | brand_known    | Brands known in category      |                                   |

        | choices |
        |         | list name | name             | label             |
        |         | sex       | male             | Male              |
        |         | sex       | female           | Female            |
        |         | brand     | ${text}          | ${text}           |
        |         | brand     | ${favorite_brand}| ${favorite_brand} |
        |         | brand     | a                | a                 |
        |         | brand     | b                | b                 |
        """  # noqa
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [
            {
                "sex": "male",
                "text": "his",
                "favorite_brand": "Generic",
                "name": "Davis",
                "brand_known": "${text} ${favorite_brand} a",
            }
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            rows[0] = list(map(_str_if_bytes, rows[0]))
            rows[1] = list(map(_str_if_bytes, rows[1]))
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[0][0], "sex")
            self.assertEqual(rows[1][0], "male")
            self.assertEqual(rows[0][1], "text")
            self.assertEqual(rows[1][1], "his")
            self.assertEqual(rows[0][2], "favorite_brand")
            self.assertEqual(rows[1][2], "Generic")
            self.assertEqual(rows[0][3], "name")
            self.assertEqual(rows[1][3], "Davis")
            self.assertEqual(rows[0][4], "brand_known")
            self.assertEqual(rows[1][4], "his Generic a")
            self.assertEqual(rows[0][5], "brand_known.$text")
            self.assertEqual(rows[1][5], 1.0)
            self.assertEqual(rows[0][6], "brand_known.$favorite_brand")
            self.assertEqual(rows[1][6], 1.0)
            self.assertEqual(rows[0][7], "brand_known.a")
            self.assertEqual(rows[1][7], 1.0)
            self.assertEqual(rows[0][8], "brand_known.b")
            self.assertEqual(rows[1][8], 0.0)

        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_zero_padded_select_one_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | select one yes_no | expensed     | Expensed?    |

        | choices |           |      |       |
        |         | list name | name | label |
        |         | yes_no    | 1    | Yes   |
        |         | yes_no    | 09   | No    |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [{"expensed": "09", "_submission_time": "2016-11-21T03:43:43.000-08:00"}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[1][0].decode("utf-8"), "09")
            self.assertEqual(rows[1][4].decode("utf-8"), "2016-11-21 03:43:43")

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_numeric_select_one_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | select one yes_no | expensed     | Expensed?    |
        |        | begin group       | A            | A            |
        |        | select one yes_no | q1           | Q1           |
        |        | end group         |              |              |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 1      | Yes    |
        |         | yes_no    | 0      | No     |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [
            {
                "expensed": "1",
                "A/q1": "1",
                "_submission_time": "2016-11-21T03:43:43.000-08:00",
            }
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            self.assertTrue(len(rows) > 1)

            # expensed 1
            self.assertEqual(_str_if_bytes(rows[0][0]), "expensed")
            self.assertEqual(_str_if_bytes(rows[1][0]), 1)

            # A/q1 1
            self.assertEqual(_str_if_bytes(rows[0][1]), "A.q1")
            self.assertEqual(rows[1][1], 1)

            # _submission_time is a date string
            self.assertEqual(_str_if_bytes(rows[0][5]), "@_submission_time")
            self.assertEqual(_str_if_bytes(rows[1][5]), "2016-11-21 03:43:43")

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_duplicate_field_different_groups(self):
        """
        Test SAV exports duplicate fields, same group - one field in repeat

        This is to test the fixing of a curious bug which happens when:
            1. You have duplicate fields in the same group
            2. One of those fields is in a repeat
            3. export_builder.TRUNCATE_GROUP_TITLE = True
        """
        md = """
        | survey |                     |          |           |           |
        |        | type                | name     | label     | required  |
        |        | text                | name     | Name      |           |
        |        | begin group         | A        | A         |           |
        |        | begin repeat        | rep      | B         |           |
        |        | select_one y_n      | allaite  | One       | yes       |
        |        | end repeat          | rep      |           |           |
        |        | select_multiple x_y | allaite  | Two       | yes       |
        |        | end group           | A        |           |           |

        | choices |           |      |       |
        |         | list name | name | label |
        |         | y_n       | 1    | Yes   |
        |         | y_n       | 2    | No    |
        |         |           |      |       |
        |         | x_y       | 1    | Oui   |
        |         | x_y       | 2    | Non   |

        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)

        # pylint: disable=protected-access
        labels = export_builder._get_sav_value_labels({"A/allaite": "allaite"})
        self.assertEqual(labels, {"allaite": {"1": "Oui", "2": "Non"}})

        # pylint: disable=protected-access
        repeat_group_labels = export_builder._get_sav_value_labels(
            {"A/rep/allaite": "allaite"}
        )
        self.assertEqual(repeat_group_labels, {"allaite": {1: "Yes", 2: "No"}})

        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            try:
                export_builder.to_zipped_sav(temp_zip_file.name, [])
            except ArgumentError as e:
                self.fail(e)

    # pylint: disable=invalid-name
    def test_split_select_multiples_choices_with_randomize_param(self):
        """
        Test that `_get_select_mulitples_choices` function generates
        choices correctly when random param set to true
        """
        md_xform = """
        | survey  |
        |         | type                     | name   | label  | parameters     |
        |         | select_one states        | state  | state  |                |
        |         | select_multiple counties | county | county | randomize=true |
        |         | select_one cities        | city   | city   |                |

        | choices |
        |         | list name | name        | label       |
        |         | states    | texas       | Texas       |
        |         | states    | washington  | Washington  |
        |         | counties  | king        | King        |
        |         | counties  | pierce      | Pierce      |
        |         | counties  | king        | King        |
        |         | counties  | cameron     | Cameron     |
        |         | cities    | dumont      | Dumont      |
        |         | cities    | finney      | Finney      |
        |         | cities    | brownsville | brownsville |
        |         | cities    | harlingen   | harlingen   |
        |         | cities    | seattle     | Seattle     |
        |         | cities    | redmond     | Redmond     |
        |         | cities    | tacoma      | Tacoma      |
        |         | cities    | puyallup    | Puyallup    |

        | settings |                         |
        |          | allow_choice_duplicates |
        |          | Yes                     |
        """  # noqa: E501
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        dd = DataDictionary()
        # pylint: disable=protected-access
        dd._survey = survey
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        child = [
            e
            for e in dd.get_survey_elements_with_choices()
            if e.bind.get("type") == SELECT_BIND_TYPE and e.type == MULTIPLE_SELECT_TYPE
        ][0]
        # pylint: disable=protected-access
        choices = export_builder._get_select_mulitples_choices(
            child, dd, ExportBuilder.GROUP_DELIMITER, ExportBuilder.TRUNCATE_GROUP_TITLE
        )

        expected_choices = [
            {
                "_label": "King",
                "_label_xpath": "county/King",
                "label": "county/King",
                "title": "county/king",
                "type": "string",
                "xpath": "county/king",
            },
            {
                "_label": "Pierce",
                "_label_xpath": "county/Pierce",
                "label": "county/Pierce",
                "title": "county/pierce",
                "type": "string",
                "xpath": "county/pierce",
            },
            {
                "_label": "King",
                "_label_xpath": "county/King",
                "label": "county/King",
                "title": "county/king",
                "type": "string",
                "xpath": "county/king",
            },
            {
                "_label": "Cameron",
                "_label_xpath": "county/Cameron",
                "label": "county/Cameron",
                "title": "county/cameron",
                "type": "string",
                "xpath": "county/cameron",
            },
        ]
        self.assertEqual(choices, expected_choices)

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_numeric_select_multiple_field(self):
        md = """
        | survey |                     |          |           |               |
        |        | type                | name     | label     | choice_filter |
        |        | select_multiple y_n | expensed | Expensed? |               |
        |        | begin group         | A        | A         |               |
        |        | select_multiple y_n | q1       | Q1        |               |
        |        | select_multiple y_n | q2       | Q2        | n=1           |
        |        | end group           |          |           |               |

        | choices |           |      |       |
        |         | list name | name | label |
        |         | y_n       | 1    | Yes   |
        |         | y_n       | 0    | No    |
        |         |           |      |       |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [
            {
                "expensed": "1",
                "A/q1": "1",
                "A/q2": "1",
                "_submission_time": "2016-11-21T03:43:43.000-08:00",
            }
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            self.assertTrue(len(rows) > 1)

            self.assertEqual(_str_if_bytes(rows[0][0]), "expensed")
            self.assertEqual(_str_if_bytes(rows[1][0]), "1")

            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(_str_if_bytes(rows[0][1]), "expensed.1")
            self.assertEqual(rows[1][1], 1)

            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(_str_if_bytes(rows[0][2]), "expensed.0")
            self.assertEqual(rows[1][2], 0)

            self.assertEqual(_str_if_bytes(rows[0][3]), "A.q1")
            self.assertEqual(_str_if_bytes(rows[1][3]), "1")

            # ensure you get a numeric value for multiple select with choice
            # filters
            self.assertEqual(_str_if_bytes(rows[0][6]), "A.q2")
            self.assertEqual(_str_if_bytes(rows[1][6]), "1")

            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(_str_if_bytes(rows[0][4]), "A.q1.1")
            self.assertEqual(rows[1][4], 1)

            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(_str_if_bytes(rows[0][5]), "A.q1.0")
            self.assertEqual(rows[1][5], 0)

            self.assertEqual(_str_if_bytes(rows[0][12]), "@_submission_time")
            self.assertEqual(_str_if_bytes(rows[1][12]), "2016-11-21 03:43:43")

        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_zero_padded_select_multiple_field(self):
        md = """
        | survey |                        |          |           |
        |        | type                   | name     | label     |
        |        | select_multiple yes_no | expensed | Expensed? |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 1      | Yes    |
        |         | yes_no    | 09     | No     |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [{"expensed": "1", "_submission_time": "2016-11-21T03:43:43.000-08:00"}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
            pass
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[1][0], b"1")
            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(rows[1][1], 1)
            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(rows[1][2], 0)
            self.assertEqual(rows[1][6], b"2016-11-21 03:43:43")

        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_values_split_select_multiple(self):
        md = """
        | survey |                        |          |           |
        |        | type                   | name     | label     |
        |        | select_multiple yes_no | expensed | Expensed? |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 2      | Yes    |
        |         | yes_no    | 09     | No     |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [
            {"expensed": "2 09", "_submission_time": "2016-11-21T03:43:43.000-08:00"}
        ]
        export_builder = ExportBuilder()
        export_builder.VALUE_SELECT_MULTIPLES = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"), returnHeader=True) as reader:
            rows = list(reader)
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[1][0], b"2 09")
            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(rows[1][1], 2)
            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(rows[1][2], b"09")
            self.assertEqual(rows[1][6], b"2016-11-21 03:43:43")

        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_duplicate_name_in_choice_list(self):
        md = """
        | survey |                         |      |                  |
        |        | type                    | name | label            |
        |        | select_multiple animals | q1   | Favorite animal? |

        | choices |           |      |            |
        |         | list_name | name | label      |
        |         | animals   | 1    | Goat       |
        |         | animals   | 2    | Camel      |
        |         | animals   | 3    | Cow        |
        |         | animals   | 4    | Sheep      |
        |         | animals   | 5    | Donkey     |
        |         | animals   | 6    | Hen        |
        |         | animals   | 6    | Other      | # <--- duplicate name
        |         | animals   | 8    | None       |
        |         | animals   | 9    | Don't Know |

        | settings |                         |
        |          | allow_choice_duplicates |
        |          | Yes                     |
        """
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        data = [
            {"q1": "1", "_submission_time": "2016-11-21T03:43:43.000-08:00"},
            {"q1": "6", "_submission_time": "2016-11-21T03:43:43.000-08:00"},
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))

    # pylint: disable=invalid-name
    def test_zipped_sav_export_external_choices(self):  # pylint: disable=invalid-name
        """
        Test that an SPSS export does not fail when it has choices from a file.
        """
        xform_markdown = """
        | survey |                                  |      |                  |
        |        | type                             | name | label            |
        |        | select_one_from_file animals.csv | q1   | Favorite animal? |
        """
        survey = self.md_to_pyxform_survey(xform_markdown, {"name": "exp"})
        data = [
            {"q1": "1", "_submission_time": "2016-11-21T03:43:43.000-08:00"},
            {"q1": "6", "_submission_time": "2016-11-21T03:43:43.000-08:00"},
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "exp.sav")))

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_duplicate_column_name(self):
        """
        Test that SAV  exports with duplicate column names
        """
        md = """
        | survey |              |          |                 |
        |        | type         | name     | label           |
        |        | text         | Sport    | Which sport     |
        |        | text         | sport    | Which fun sports|
        """
        survey = self.md_to_pyxform_survey(md, {"name": "sports"})
        data = [
            {
                "Sport": "Basketball",
                "sport": "Soccer",
                "_submission_time": "2016-11-21T03:43:43.000-08:00",
            }
        ]

        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "sports.sav")))
        # check file's contents

        with SavReader(
            os.path.join(temp_dir, "sports.sav"), returnHeader=True
        ) as reader:
            rows = list(reader)

            # Check that columns are present
            self.assertIn("Sport", [_str_if_bytes(item) for item in rows[0]])
            # Check for sport in first 5 characters
            # because rows contains 'sport@d4b6'
            self.assertIn("sport", list(map(_str_if_bytes, [x[0:5] for x in rows[0]])))

    # pylint: disable=invalid-name
    def test_xlsx_export_works_with_unicode(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrenss_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.data_utf8)
            temp_xls_file.seek(0)
            # check that values for red\'s and blue\'s are set to true
            workbook = load_workbook(temp_xls_file.name)
            children_sheet = workbook["children.info"]
            data = {x.value: y.value for x, y in children_sheet.columns}
            self.assertTrue(data["children.info/fav_colors/red's"])
            self.assertTrue(data["children.info/fav_colors/blue's"])
            self.assertFalse(data["children.info/fav_colors/pink's"])

    # pylint: disable=invalid-name
    def test_xlsx_export_with_hxl_adds_extra_row(self):
        # hxl_example.xlsx contains `instance::hxl` column whose value is #age
        xlsform_path = os.path.join(
            settings.PROJECT_ROOT,
            "apps",
            "main",
            "tests",
            "fixtures",
            "hxl_test",
            "hxl_example.xlsx",
        )
        survey = create_survey_from_xls(
            xlsform_path, default_name=xlsform_path.split("/")[-1].split(".")[0]
        )
        export_builder = ExportBuilder()
        export_builder.INCLUDE_HXL = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")

        survey_elements = [
            survey_item[1]
            for survey_item in survey.items()
            if survey_item[0] == "children"
        ][0]

        columns_with_hxl = export_builder.INCLUDE_HXL and get_columns_with_hxl(
            survey_elements
        )

        export_builder.to_xlsx_export(
            temp_xls_file.name, self.data_utf8, columns_with_hxl=columns_with_hxl
        )
        temp_xls_file.seek(0)
        workbook = load_workbook(temp_xls_file.name)
        children_sheet = workbook["hxl_example"]
        self.assertTrue(children_sheet)

        # we pick the second row because the first row has xform fieldnames
        rows = list(children_sheet.rows)
        hxl_row = [a.value for a in rows[1]]
        self.assertIn("#age", hxl_row)

    # pylint: disable=invalid-name,too-many-locals
    def test_export_with_image_attachments(self):
        """
        Test that the url for images is displayed correctly in exports
        """
        md = """
        | survey |       |        |       |
        |        | type  | name   | label |
        |        | image | image1 | Photo |
        """
        self._create_user_and_login()
        self.xform = self._publish_markdown(md, self.user)

        xml_string = f"""
        <data id="{self.xform.id_string}">
            <meta>
                <instanceID>uuid:UJ6jSMAJ1Jz4EszdgHy8n852AsKaqBPO5</instanceID>
            </meta>
            <image1>1300221157303.jpg</image1>
        </data>
        """

        file_path = (
            f"{settings.PROJECT_ROOT}/apps/logger/tests/Health_2011_03_13."
            "xml_2011-03-15_20-30-28/1300221157303.jpg"
        )
        media_file = django_file(
            path=file_path, field_name="image1", content_type="image/jpeg"
        )
        create_instance(
            self.user.username,
            BytesIO(xml_string.strip().encode("utf-8")),
            media_files=[media_file],
        )

        xdata = [datum for datum in query_data(self.xform)]
        survey = self.md_to_pyxform_survey(md, {"name": "exp"})
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(
                temp_xls_file, xdata, options={"host": "example.com"}
            )
            temp_xls_file.seek(0)
            workbook = load_workbook(temp_xls_file)
            children_sheet = workbook["exp"]
            self.assertTrue(children_sheet)
            rows = list(children_sheet.rows)
            row = [a.value for a in rows[1]]
            attachment_id = xdata[0]["_attachments"][0]["id"]
            attachment_filename = xdata[0]["_attachments"][0]["filename"]
            attachment_url = f"http://example.com/api/v1/files/{attachment_id}?filename={attachment_filename}"  # noqa
            self.assertIn(attachment_url, row)

    # pylint: disable=invalid-name
    def test_generation_of_multi_selects_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_select_multiples = {
            "children": {
                "children/fav_colors": [
                    "children/fav_colors/red",
                    "children/fav_colors/blue",
                    "children/fav_colors/pink",
                ],
                "children/ice.creams": [
                    "children/ice.creams/vanilla",
                    "children/ice.creams/strawberry",
                    "children/ice.creams/chocolate",
                ],
            }
        }
        select_multiples = export_builder.select_multiples
        self.assertTrue("children" in select_multiples)
        self.assertTrue("children/fav_colors" in select_multiples["children"])
        self.assertTrue("children/ice.creams" in select_multiples["children"])
        self.assertEqual(
            sorted(
                [
                    choice["xpath"]
                    for choice in select_multiples["children"]["children/fav_colors"]
                ]
            ),
            sorted(expected_select_multiples["children"]["children/fav_colors"]),
        )
        self.assertEqual(
            sorted(
                [
                    choice["xpath"]
                    for choice in select_multiples["children"]["children/ice.creams"]
                ]
            ),
            sorted(expected_select_multiples["children"]["children/ice.creams"]),
        )

    # pylint: disable=invalid-name
    def test_split_select_multiples_works(self):
        """
        Test split_select_multiples works as expected.
        """
        select_multiples = {
            "children/fav_colors": [
                {
                    "xpath": "children/fav_colors/red",
                    "label": "fav_colors/Red",
                },
                {
                    "xpath": "children/fav_colors/blue",
                    "label": "fav_colors/Blue",
                },
                {
                    "xpath": "children/fav_colors/pink",
                    "label": "fav_colors/Pink",
                },
            ]
        }
        row = {
            "children/name": "Mike",
            "children/age": 5,
            "children/fav_colors": "red blue",
        }
        new_row = ExportBuilder.split_select_multiples(row, select_multiples)
        expected_row = {
            "children/name": "Mike",
            "children/age": 5,
            "children/fav_colors": "red blue",
            "children/fav_colors/red": True,
            "children/fav_colors/blue": True,
            "children/fav_colors/pink": False,
        }
        self.assertEqual(new_row, expected_row)
        row = {
            "children/name": "Mike",
            "children/age": 5,
        }
        new_row = ExportBuilder.split_select_multiples(row, select_multiples)
        expected_row = {
            "children/name": "Mike",
            "children/age": 5,
            "children/fav_colors/red": None,
            "children/fav_colors/blue": None,
            "children/fav_colors/pink": None,
        }
        self.assertEqual(new_row, expected_row)

    # pylint: disable=invalid-name
    def test_split_select_mutliples_works_with_int_value_in_row(self):
        select_multiples = {
            "children/fav_number": [
                {
                    "xpath": "children/fav_number/1",
                },
                {
                    "xpath": "children/fav_number/2",
                },
                {
                    "xpath": "children/fav_number/3",
                },
            ]
        }
        row = {"children/fav_number": 1}

        expected_row = {
            "children/fav_number/1": True,
            "children/fav_number": 1,
            "children/fav_number/3": False,
            "children/fav_number/2": False,
        }

        new_row = ExportBuilder.split_select_multiples(row, select_multiples)
        self.assertTrue(new_row)
        self.assertEqual(new_row, expected_row)

    # pylint: disable=invalid-name
    def test_split_select_multiples_works_when_data_is_blank(self):
        select_multiples = {
            "children/fav_colors": [
                {
                    "xpath": "children/fav_colors/red",
                    "label": "fav_colors/Red",
                },
                {
                    "xpath": "children/fav_colors/blue",
                    "label": "fav_colors/Blue",
                },
                {
                    "xpath": "children/fav_colors/pink",
                    "label": "fav_colors/Pink",
                },
            ]
        }
        row = {"children/name": "Mike", "children/age": 5, "children/fav_colors": ""}
        new_row = ExportBuilder.split_select_multiples(row, select_multiples)
        expected_row = {
            "children/name": "Mike",
            "children/age": 5,
            "children/fav_colors": "",
            "children/fav_colors/red": None,
            "children/fav_colors/blue": None,
            "children/fav_colors/pink": None,
        }
        self.assertEqual(new_row, expected_row)

    # pylint: disable=invalid-name
    def test_generation_of_gps_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_gps_fields = {
            "childrens_survey": {
                "geo/geolocation": [
                    "geo/_geolocation_latitude",
                    "geo/_geolocation_longitude",
                    "geo/_geolocation_altitude",
                    "geo/_geolocation_precision",
                ]
            }
        }
        gps_fields = export_builder.gps_fields
        self.assertTrue("childrens_survey" in gps_fields)
        self.assertEqual(
            sorted(gps_fields["childrens_survey"]),
            sorted(expected_gps_fields["childrens_survey"]),
        )

    def test_split_gps_components_works(self):
        gps_fields = {
            "geo/geolocation": [
                "geo/_geolocation_latitude",
                "geo/_geolocation_longitude",
                "geo/_geolocation_altitude",
                "geo/_geolocation_precision",
            ]
        }
        row = {
            "geo/geolocation": "1.0 36.1 2000 20",
        }
        new_row = ExportBuilder.split_gps_components(row, gps_fields)
        expected_row = {
            "geo/geolocation": "1.0 36.1 2000 20",
            "geo/_geolocation_latitude": "1.0",
            "geo/_geolocation_longitude": "36.1",
            "geo/_geolocation_altitude": "2000",
            "geo/_geolocation_precision": "20",
        }
        self.assertEqual(new_row, expected_row)

    # pylint: disable=invalid-name
    def test_split_gps_components_works_when_gps_data_is_blank(self):
        gps_fields = {
            "geo/geolocation": [
                "geo/_geolocation_latitude",
                "geo/_geolocation_longitude",
                "geo/_geolocation_altitude",
                "geo/_geolocation_precision",
            ]
        }
        row = {
            "geo/geolocation": "",
        }
        new_row = ExportBuilder.split_gps_components(row, gps_fields)
        expected_row = {
            "geo/geolocation": "",
        }
        self.assertEqual(new_row, expected_row)

    # pylint: disable=invalid-name
    def test_generation_of_mongo_encoded_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_encoded_fields = {
            "childrens_survey": {
                "tel/tel.office": f"tel/{_encode_for_mongo('tel.office')}",
                "tel/tel.mobile": f"tel/{_encode_for_mongo('tel.mobile')}",
            }
        }
        encoded_fields = export_builder.encoded_fields
        self.assertTrue("childrens_survey" in encoded_fields)
        self.assertEqual(
            encoded_fields["childrens_survey"],
            expected_encoded_fields["childrens_survey"],
        )

    # pylint: disable=invalid-name
    def test_decode_fields_names_encoded_for_mongo(self):
        encoded_fields = {"tel/tel.office": f"tel/{_encode_for_mongo('tel.office')}"}
        row = {
            "name": "Abe",
            "age": 35,
            f"tel/{_encode_for_mongo('tel.office')}": "123-456-789",
        }
        new_row = ExportBuilder.decode_mongo_encoded_fields(row, encoded_fields)
        expected_row = {"name": "Abe", "age": 35, "tel/tel.office": "123-456-789"}
        self.assertEqual(new_row, expected_row)

    def test_generate_field_title(self):
        self._create_childrens_survey()
        field_name = ExportBuilder.format_field_title(
            "children/age", ".", data_dictionary=self.data_dictionary
        )
        expected_field_name = "children.age"
        self.assertEqual(field_name, expected_field_name)

    # pylint: disable=invalid-name
    def test_delimiter_replacement_works_existing_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_sections = [
            {
                "name": "children",
                "elements": [{"title": "children.name", "xpath": "children/name"}],
            }
        ]
        children_section = export_builder.section_by_name("children")
        self.assertEqual(
            children_section["elements"][0]["title"],
            expected_sections[0]["elements"][0]["title"],
        )

    # pylint: disable=invalid-name
    def test_delimiter_replacement_works_generated_multi_select_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section = {
            "name": "children",
            "elements": [
                {"title": "children.fav_colors.red", "xpath": "children/fav_colors/red"}
            ],
        }
        childrens_section = export_builder.section_by_name("children")
        match = [
            x
            for x in childrens_section["elements"]
            if expected_section["elements"][0]["xpath"] == x["xpath"]
        ][0]
        self.assertEqual(expected_section["elements"][0]["title"], match["title"])

    # pylint: disable=invalid-name
    def test_delimiter_replacement_works_for_generated_gps_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section = {
            "name": "childrens_survey",
            "elements": [
                {
                    "title": "geo._geolocation_latitude",
                    "xpath": "geo/_geolocation_latitude",
                }
            ],
        }
        main_section = export_builder.section_by_name("childrens_survey")
        match = [
            x
            for x in main_section["elements"]
            if expected_section["elements"][0]["xpath"] == x["xpath"]
        ][0]
        self.assertEqual(expected_section["elements"][0]["title"], match["title"])

    def test_to_xlsx_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as xls_file:
            filename = xls_file.name
            export_builder.to_xlsx_export(filename, self.data)
            xls_file.seek(0)
            workbook = load_workbook(filename)
            # check that we have childrens_survey, children, children_cartoons
            # and children_cartoons_characters sheets
            expected_sheet_names = [
                "childrens_survey",
                "children",
                "children_cartoons",
                "children_cartoons_characters",
            ]
            self.assertEqual(list(workbook.sheetnames), expected_sheet_names)

            # check header columns
            main_sheet = workbook["childrens_survey"]
            expected_column_headers = [
                "name",
                "age",
                "geo/geolocation",
                "geo/_geolocation_latitude",
                "geo/_geolocation_longitude",
                "geo/_geolocation_altitude",
                "geo/_geolocation_precision",
                "tel/tel.office",
                "tel/tel.mobile",
                "_id",
                "meta/instanceID",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_index",
                "_parent_table_name",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            column_headers = list(main_sheet.values)[0]
            self.assertEqual(
                sorted(list(column_headers)), sorted(expected_column_headers)
            )

            childrens_sheet = workbook["children"]
            expected_column_headers = [
                "children/name",
                "children/age",
                "children/fav_colors",
                "children/fav_colors/red",
                "children/fav_colors/blue",
                "children/fav_colors/pink",
                "children/ice.creams",
                "children/ice.creams/vanilla",
                "children/ice.creams/strawberry",
                "children/ice.creams/chocolate",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_index",
                "_parent_table_name",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            column_headers = list(childrens_sheet.values)[0]
            self.assertEqual(
                sorted(list(column_headers)), sorted(expected_column_headers)
            )

            cartoons_sheet = workbook["children_cartoons"]
            expected_column_headers = [
                "children/cartoons/name",
                "children/cartoons/why",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_index",
                "_parent_table_name",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            column_headers = list(cartoons_sheet.values)[0]
            self.assertEqual(
                sorted(list(column_headers)), sorted(expected_column_headers)
            )

            characters_sheet = workbook["children_cartoons_characters"]
            expected_column_headers = [
                "children/cartoons/characters/name",
                "children/cartoons/characters/good_or_evil",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_index",
                "_parent_table_name",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            column_headers = list(characters_sheet.values)[0]
            self.assertEqual(
                sorted(list(column_headers)), sorted(expected_column_headers)
            )

    # pylint: disable=invalid-name
    def test_to_xlsx_export_respects_custom_field_delimiter(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = ExportBuilder.GROUP_DELIMITER_DOT
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as xls_file:
            filename = xls_file.name
            export_builder.to_xlsx_export(filename, self.data)
            xls_file.seek(0)
            workbook = load_workbook(filename)

            # check header columns
            main_sheet = workbook["childrens_survey"]
            expected_column_headers = [
                "name",
                "age",
                "geo.geolocation",
                "geo._geolocation_latitude",
                "geo._geolocation_longitude",
                "geo._geolocation_altitude",
                "geo._geolocation_precision",
                "tel.tel.office",
                "tel.tel.mobile",
                "_id",
                "meta.instanceID",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_index",
                "_parent_table_name",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            column_headers = list(main_sheet.values)[0]
            self.assertEqual(
                sorted(list(column_headers)), sorted(expected_column_headers)
            )

    # pylint: disable=invalid-name
    def test_get_valid_sheet_name_catches_duplicates(self):
        work_sheets = {"childrens_survey": "Worksheet"}
        desired_sheet_name = "childrens_survey"
        expected_sheet_name = "childrens_survey1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheets
        )
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    # pylint: disable=invalid-name
    def test_get_valid_sheet_name_catches_long_names(self):
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_lo"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, []
        )
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    # pylint: disable=invalid-name
    def test_get_valid_sheet_name_catches_long_duplicate_names(self):
        work_sheet_titles = ["childrens_survey_with_a_very_lo"]
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_l1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheet_titles
        )
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    # pylint: disable=invalid-name
    def test_to_xlsx_export_generates_valid_sheet_names(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_with_a_very_long_name.xlsx"),
            default_name="childrens_survey_with_a_very_long_name",
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as xls_file:
            filename = xls_file.name
            export_builder.to_xlsx_export(filename, self.data)
            xls_file.seek(0)
            workbook = load_workbook(filename)
            # check that we have childrens_survey, children, children_cartoons
            # and children_cartoons_characters sheets
            expected_sheet_names = [
                "childrens_survey_with_a_very_lo",
                "childrens_survey_with_a_very_l1",
                "childrens_survey_with_a_very_l2",
                "childrens_survey_with_a_very_l3",
            ]
            self.assertEqual(list(workbook.sheetnames), expected_sheet_names)

    # pylint: disable=invalid-name
    def test_child_record_parent_table_is_updated_when_sheet_is_renamed(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_with_a_very_long_name.xlsx"),
            default_name="childrens_survey_with_a_very_long_name",
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as xls_file:
            filename = xls_file.name
            export_builder.to_xlsx_export(filename, self.long_survey_data)
            xls_file.seek(0)
            workbook = load_workbook(filename)

            # get the children's sheet
            ws1 = workbook["childrens_survey_with_a_very_l1"]

            # parent_table is in cell K2
            parent_table_name = ws1["K2"].value
            expected_parent_table_name = "childrens_survey_with_a_very_lo"
            self.assertEqual(parent_table_name, expected_parent_table_name)

            # get cartoons sheet
            ws2 = workbook["childrens_survey_with_a_very_l2"]
            parent_table_name = ws2["G2"].value
            expected_parent_table_name = "childrens_survey_with_a_very_l1"
            self.assertEqual(parent_table_name, expected_parent_table_name)

    def test_type_conversion(self):
        submission_1 = {
            "_id": 579827,
            "geolocation": "-1.2625482 36.7924794 0.0 21.0",
            "_bamboo_dataset_id": "",
            "meta/instanceID": "uuid:2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "name": "Smith",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "_submission_time": "2013-07-03T08:25:30",
            "age": "107",
            "_uuid": "2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "when": "2013-07-03",
            "amount": "250.0",
            "_geolocation": ["-1.2625482", "36.7924794"],
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "precisely": "2013-07-03T15:24:00.000+03",
            "really": "15:24:00.000+03",
        }

        submission_2 = {
            "_id": 579828,
            "_submission_time": "2013-07-03T08:26:10",
            "_uuid": "5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "_bamboo_dataset_id": "",
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "meta/instanceID": "uuid:5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "amount": "",
        }

        survey = create_survey_from_xls(
            viewer_fixture_path("test_data_types/test_data_types.xlsx"),
            default_name="test_data_types",
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # format submission 1 for export
        survey_name = survey.name
        indices = {survey_name: 0}
        data = dict_to_joined_export(
            submission_1, 1, indices, survey_name, survey, submission_1, None
        )
        new_row = export_builder.pre_process_row(
            data[survey_name], export_builder.sections[0]
        )
        self.assertIsInstance(new_row["age"], int)
        self.assertIsInstance(new_row["when"], datetime.date)
        self.assertIsInstance(new_row["amount"], float)

        # check missing values dont break and empty values return blank strings
        indices = {survey_name: 0}
        data = dict_to_joined_export(
            submission_2, 1, indices, survey_name, survey, submission_2, None
        )
        new_row = export_builder.pre_process_row(
            data[survey_name], export_builder.sections[0]
        )
        self.assertIsInstance(new_row["amount"], str)
        self.assertEqual(new_row["amount"], "")

    # pylint: disable=invalid-name
    def test_xls_convert_dates_before_1900(self):
        survey = create_survey_from_xls(
            viewer_fixture_path("test_data_types/test_data_types.xlsx"),
            default_name="test_data_types",
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        data = [
            {
                "name": "Abe",
                "when": "1899-07-03",
            }
        ]
        # create export file
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, data)
        # this should error if there is a problem, not sure what to assert

    def test_convert_types(self):
        val = "1"
        expected_val = 1
        converted_val = ExportBuilder.convert_type(val, "int")
        self.assertIsInstance(converted_val, int)
        self.assertEqual(converted_val, expected_val)

        val = "1.2"
        expected_val = 1.2
        converted_val = ExportBuilder.convert_type(val, "decimal")
        self.assertIsInstance(converted_val, float)
        self.assertEqual(converted_val, expected_val)

        val = "2012-06-23"
        expected_val = datetime.date(2012, 6, 23)
        converted_val = ExportBuilder.convert_type(val, "date")
        self.assertIsInstance(converted_val, datetime.date)
        self.assertEqual(converted_val, expected_val)

    def test_to_sav_export(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            filename = temp_zip_file.name
            export_builder.to_zipped_sav(filename, self.data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(d, index, indices, survey_name, survey, d, None)
            )
            index += 1

        # check that each file exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, f"{survey.name}.sav")))

        def _test_sav_file(section):
            with SavReader(
                os.path.join(temp_dir, f"{section}.sav"), returnHeader=True
            ) as reader:
                header = next(reader)
                rows = list(reader)

                # open comparison file
                with SavReader(
                    _logger_fixture_path("spss", f"{section}.sav"),
                    returnHeader=True,
                ) as fixture_reader:
                    fixture_header = next(fixture_reader)
                    self.assertEqual(header, fixture_header)
                    expected_rows = list(fixture_reader)
                    self.assertEqual(rows, expected_rows)

                if section == "children_cartoons_charactors":
                    self.assertEqual(
                        reader.valueLabels, {"good_or_evil": {"good": "Good"}}
                    )

        for section in export_builder.sections:
            section_name = section["name"].replace("/", "_")
            _test_sav_file(section_name)

    def test_to_sav_export_language(self):
        survey = self._create_childrens_survey("childrens_survey_sw.xlsx")
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            filename = temp_zip_file.name
            export_builder.to_zipped_sav(filename, self.data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(d, index, indices, survey_name, survey, d, None)
            )
            index += 1

        # check that each file exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, f"{survey.name}.sav")))

        def _test_sav_file(section):
            with SavReader(
                os.path.join(temp_dir, f"{section}.sav"), returnHeader=True
            ) as reader:
                header = next(reader)
                rows = list(reader)
                if section != "childrens_survey_sw":
                    section += "_sw"

                # open comparison file
                with SavReader(
                    _logger_fixture_path("spss", f"{section}.sav"),
                    returnHeader=True,
                ) as fixture_reader:
                    fixture_header = next(fixture_reader)
                    self.assertEqual(header, fixture_header)
                    expected_rows = list(fixture_reader)
                    self.assertEqual(rows, expected_rows)

                if section == "children_cartoons_charactors":
                    self.assertEqual(
                        reader.valueLabels, {"good_or_evil": {"good": "Good"}}
                    )

        for section in export_builder.sections:
            section_name = section["name"].replace("/", "_")
            _test_sav_file(section_name)

    def test_export_zipped_zap_missing_en_label(self):
        """Blank language label defaults to label for default language"""
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_sw_missing_en_label.xlsx"),
            default_name="childrens_survey_sw",
        )
        # default_language is set to swahili
        self.assertEqual(survey.to_json_dict().get("default_language"), "swahili")
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        # export to be in english
        export_builder.language = "english"
        export_builder.set_survey(survey)

        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            filename = temp_zip_file.name
            export_builder.to_zipped_sav(filename, self.data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        # check that each file exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, f"{survey.name}.sav")))
        checks = 0

        for section in export_builder.sections:
            section_name = section["name"]

            if section_name == "childrens_survey_sw":
                # Default swahili label is used incase english label is missing for question
                result = filter(
                    lambda question: question["label"] == "1. Jina lako ni?",
                    section["elements"],
                )
                self.assertEqual(len(list(result)), 1)
                checks += 1

            if section_name == "children":
                # Default swahili label is used incase english label is missing for choice
                result = filter(
                    lambda choice: choice["label"] == "fav_colors/Nyekundu",
                    section["elements"],
                )
                self.assertEqual(len(list(result)), 1)
                checks += 1

        self.assertEqual(checks, 2)

    # pylint: disable=invalid-name
    def test_generate_field_title_truncated_titles(self):
        self._create_childrens_survey()
        field_name = ExportBuilder.format_field_title(
            "children/age",
            "/",
            data_dictionary=self.data_dictionary,
            remove_group_name=True,
        )
        expected_field_name = "age"
        self.assertEqual(field_name, expected_field_name)

    # pylint: disable=invalid-name
    def test_generate_field_title_truncated_titles_select_multiple(self):
        self._create_childrens_survey()
        field_name = ExportBuilder.format_field_title(
            "children/fav_colors/red",
            "/",
            data_dictionary=self.data_dictionary,
            remove_group_name=True,
        )
        expected_field_name = "fav_colors/red"
        self.assertEqual(field_name, expected_field_name)

    # pylint: disable=invalid-name
    def test_xlsx_export_remove_group_name(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.data_utf8)
            temp_xls_file.seek(0)
            # check that values for red\'s and blue\'s are set to true
            workbook = load_workbook(temp_xls_file.name)
            children_sheet = workbook["children.info"]
            data = {r[0].value: r[1].value for r in children_sheet.columns}
            self.assertTrue(data["fav_colors/red's"])
            self.assertTrue(data["fav_colors/blue's"])
            self.assertFalse(data["fav_colors/pink's"])

    # pylint: disable=invalid-name
    def test_gps_xlsx_export_remove_group_name(self):
        self._publish_xls_file_and_set_xform(_logger_fixture_path("gps_data.xlsx"))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(self.xform.survey, self.xform)
        self.assertEqual(self.xform.instances.count(), 0)
        self._make_submission(_logger_fixture_path("gps_data.xml"))
        self.assertEqual(self.xform.instances.count(), 1)
        records = self.xform.instances.all().order_by("id")
        inst_json = records.first().json
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, [inst_json])
            temp_xls_file.seek(0)
            workbook = load_workbook(temp_xls_file.name)
            gps_sheet = workbook["data"]
            expected_headers = [
                "start",
                "end",
                "test",
                "Age",
                "Image",
                "Select_one_gender",
                "respondent_name",
                "_1_2_Where_do_you_live",
                "_1_3_Another_test",
                "GPS",
                "_GPS_latitude",
                "_GPS_longitude",
                "_GPS_altitude",
                "_GPS_precision",
                "Another_GPS",
                "_Another_GPS_latitude",
                "_Another_GPS_longitude",
                "_Another_GPS_altitude",
                "_Another_GPS_precision",
                "__version__",
                "instanceID",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            self.assertEqual(expected_headers, list(tuple(gps_sheet.values)[0]))
            # test exported data
            expected_data = [
                "2023-01-20T12:11:29.278+03:00",
                "2023-01-20T12:12:00.443+03:00",
                "2",
                32,
                None,
                "male",
                "dsa",
                "zcxsd",
                None,
                "-1.248238 36.685681 0 0",
                -1.248238,
                36.685681,
                0,
                0,
                "-1.244172 36.685359 0 0",
                -1.244172,
                36.685359,
                0,
                0,
                "vTEmiygu2uLZpPHBYX8jKj",
                "uuid:76887abc-7bc4-4f9c-ba64-197968359d36",
                inst_json.get("_id"),
                "76887abc-7bc4-4f9c-ba64-197968359d36",
                1,
                None,
                -1,
                None,
                None,
                "vTEmiygu2uLZpPHBYX8jKj",
                31,
                "bob",
            ]
            actual_data = list(tuple(gps_sheet.values)[1])
            # remove submission time
            del actual_data[23]
            self.assertEqual(expected_data, actual_data)

    def test_zipped_csv_export_remove_group_name(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(
            os.path.join(temp_dir, "children.info.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = [
                "name.first",
                "age",
                "fav_colors",
                "fav_colors/red's",
                "fav_colors/blue's",
                "fav_colors/pink's",
                "ice_creams",
                "ice_creams/vanilla",
                "ice_creams/strawberry",
                "ice_creams/chocolate",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            rows = list(reader)
            actual_headers = list(rows[0])
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(data["fav_colors/red's"], "True")
            self.assertEqual(data["fav_colors/blue's"], "True")
            self.assertEqual(data["fav_colors/pink's"], "False")
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    def test_xlsx_export_with_labels(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.data_utf8)
            temp_xls_file.seek(0)
            # check that values for red\'s and blue\'s are set to true
            workbook = load_workbook(temp_xls_file.name)
            children_sheet = workbook["children.info"]
            labels = {r[0].value: r[1].value for r in children_sheet.columns}
            self.assertEqual(labels["name.first"], "3.1 Childs name")
            self.assertEqual(labels["age"], "3.2 Child age")
            self.assertEqual(labels["fav_colors/red's"], "fav_colors/Red")
            self.assertEqual(labels["fav_colors/blue's"], "fav_colors/Blue")
            self.assertEqual(labels["fav_colors/pink's"], "fav_colors/Pink")

            data = {r[0].value: r[2].value for r in children_sheet.columns}
            self.assertEqual(data["name.first"], "Mike")
            self.assertEqual(data["age"], 5)
            self.assertTrue(data["fav_colors/red's"])
            self.assertTrue(data["fav_colors/blue's"])
            self.assertFalse(data["fav_colors/pink's"])

    # pylint: disable=invalid-name
    def test_xlsx_export_with_labels_only(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS_ONLY = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        export_builder.to_xlsx_export(temp_xls_file.name, self.data_utf8)
        temp_xls_file.seek(0)
        # check that values for red\'s and blue\'s are set to true
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb["children.info"]
        data = {r[0].value: r[1].value for r in children_sheet.columns}
        self.assertEqual(data["3.1 Childs name"], "Mike")
        self.assertEqual(data["3.2 Child age"], 5)
        self.assertTrue(data["fav_colors/Red"])
        self.assertTrue(data["fav_colors/Blue"])
        self.assertFalse(data["fav_colors/Pink"])
        temp_xls_file.close()

    # pylint: disable=invalid-name
    def test_zipped_csv_export_with_labels(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(
            os.path.join(temp_dir, "children.info.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = [
                "name.first",
                "age",
                "fav_colors",
                "fav_colors/red's",
                "fav_colors/blue's",
                "fav_colors/pink's",
                "ice_creams",
                "ice_creams/vanilla",
                "ice_creams/strawberry",
                "ice_creams/chocolate",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            expected_labels = [
                "3.1 Childs name",
                "3.2 Child age",
                "3.3 Favorite Colors",
                "fav_colors/Red",
                "fav_colors/Blue",
                "fav_colors/Pink",
                "3.3 Ice Creams",
                "ice_creams/Vanilla",
                "ice_creams/Strawberry",
                "ice_creams/Chocolate",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            rows = list(reader)
            actual_headers = list(rows[0])
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            actual_labels = list(rows[1])
            self.assertEqual(sorted(actual_labels), sorted(expected_labels))
            data = dict(zip(rows[0], rows[2]))
            self.assertEqual(data["fav_colors/red's"], "True")
            self.assertEqual(data["fav_colors/blue's"], "True")
            self.assertEqual(data["fav_colors/pink's"], "False")
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_zipped_csv_export_with_labels_only(self):
        """
        csv writer does not handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_unicode.xlsx"),
            default_name="childrens_survey_unicode",
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS_ONLY = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)
            # check that the children's file (which has the unicode header) exists
            self.assertTrue(os.path.exists(os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(
            os.path.join(temp_dir, "children.info.csv"), encoding="utf-8"
        ) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = [
                "3.1 Childs name",
                "3.2 Child age",
                "3.3 Favorite Colors",
                "fav_colors/Red",
                "fav_colors/Blue",
                "fav_colors/Pink",
                "3.3 Ice Creams",
                "ice_creams/Vanilla",
                "ice_creams/Strawberry",
                "ice_creams/Chocolate",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
            ]
            rows = list(reader)
            actual_headers = list(rows[0])
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(data["fav_colors/Red"], "True")
            self.assertEqual(data["fav_colors/Blue"], "True")
            self.assertEqual(data["fav_colors/Pink"], "False")
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    def test_to_sav_export_with_labels(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            filename = temp_zip_file.name
            export_builder.to_zipped_sav(filename, self.data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for item in self.data:
            outputs.append(
                dict_to_joined_export(
                    item, index, indices, survey_name, survey, item, None
                )
            )
            index += 1

        # check that each file exists
        self.assertTrue(os.path.exists(os.path.join(temp_dir, f"{survey.name}.sav")))

        def _test_sav_file(section):
            sav_path = os.path.join(temp_dir, f"{section}.sav")
            if section == "children_survey":
                with SavHeaderReader(sav_path) as header:
                    expected_labels = [
                        "1. What is your name?",
                        "2. How old are you?",
                        "4. Geo-location",
                        "5.1 Office telephone",
                        "5.2 Mobile telephone",
                        "_duration",
                        "_id",
                        "_index",
                        "_notes",
                        "_parent_index",
                        "_parent_table_name",
                        "_submission_time",
                        "_submitted_by",
                        "_tags",
                        "_uuid",
                        "_version",
                        "geo/_geolocation_altitude",
                        "geo/_geolocation_latitude",
                        "geo/_geolocation_longitude",
                        "geo/_geolocation_precision",
                        "meta/instanceID",
                    ]
                    labels = header.varLabels.values()
                    self.assertEqual(sorted(expected_labels), sorted(labels))

            with SavReader(sav_path, returnHeader=True) as reader:
                header = next(reader)
                rows = list(reader)

                # open comparison file
                with SavReader(
                    _logger_fixture_path("spss", f"{section}.sav"),
                    returnHeader=True,
                ) as fixture_reader:
                    fixture_header = next(fixture_reader)
                    self.assertEqual(header, fixture_header)
                    expected_rows = list(fixture_reader)
                    self.assertEqual(rows, expected_rows)

        for section in export_builder.sections:
            section_name = section["name"].replace("/", "_")
            _test_sav_file(section_name)

    # pylint: disable=invalid-name
    def test_xlsx_export_with_english_labels(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_en.xlsx"),
            default_name="childrens_survey_en",
        )
        # no default_language is not set
        self.assertEqual(survey.to_json_dict().get("default_language"), "default")
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.data)
            temp_xls_file.seek(0)
            workbook = load_workbook(temp_xls_file.name)
            childrens_survey_sheet = workbook["childrens_survey_en"]
            labels = {r[0].value: r[1].value for r in childrens_survey_sheet.columns}
            self.assertEqual(labels["name"], "1. What is your name?")
            self.assertEqual(labels["age"], "2. How old are you?")

            children_sheet = workbook["children"]
            labels = {r[0].value: r[1].value for r in children_sheet.columns}
            self.assertEqual(labels["fav_colors/red"], "fav_colors/Red")
            self.assertEqual(labels["fav_colors/blue"], "fav_colors/Blue")

    # pylint: disable=invalid-name
    def test_xlsx_export_with_swahili_labels(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_sw.xlsx"),
            default_name="childrens_survey_sw",
        )
        # default_language is set to swahili
        self.assertEqual(survey.to_json_dict().get("default_language"), "swahili")
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.data)
            temp_xls_file.seek(0)
            workbook = load_workbook(temp_xls_file.name)
            childrens_survey_sheet = workbook["childrens_survey_sw"]
            labels = {r[0].value: r[1].value for r in childrens_survey_sheet.columns}
            self.assertEqual(labels["name"], "1. Jina lako ni?")
            self.assertEqual(labels["age"], "2. Umri wako ni?")

            children_sheet = workbook["children"]
            labels = {r[0].value: r[1].value for r in children_sheet.columns}
            self.assertEqual(labels["fav_colors/red"], "fav_colors/Nyekundu")
            self.assertEqual(labels["fav_colors/blue"], "fav_colors/Bluu")

    # pylint: disable=invalid-name
    def test_csv_export_with_swahili_labels(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_sw.xlsx"),
            default_name="childrens_survey_sw",
        )
        # default_language is set to swahili
        self.assertEqual(survey.to_json_dict().get("default_language"), "swahili")
        dd = DataDictionary()
        # pylint: disable=protected-access
        dd._survey = survey
        ordered_columns = OrderedDict()
        # pylint: disable=protected-access
        CSVDataFrameBuilder._build_ordered_columns(survey, ordered_columns)
        ordered_columns["children/fav_colors/red"] = None
        labels = get_labels_from_columns(ordered_columns, dd, "/")
        self.assertIn("1. Jina lako ni?", labels)
        self.assertIn("2. Umri wako ni?", labels)
        self.assertIn("fav_colors/Nyekundu", labels)

        # use language provided in keyword argument
        labels = get_labels_from_columns(ordered_columns, dd, "/", language="english")
        self.assertIn("1. What is your name?", labels)
        self.assertIn("2. How old are you?", labels)
        self.assertIn("fav_colors/Red", labels)

        # use default language when language supplied does not exist
        labels = get_labels_from_columns(ordered_columns, dd, "/", language="Chinese")
        self.assertIn("1. Jina lako ni?", labels)
        self.assertIn("2. Umri wako ni?", labels)
        self.assertIn("fav_colors/Nyekundu", labels)

    def test_select_multiples_choices(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("childrens_survey_sw.xlsx"),
            default_name="childrens_survey_sw",
        )
        dd = DataDictionary()
        # pylint: disable=protected-access
        dd._survey = survey
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        child = [
            e
            for e in dd.get_survey_elements_with_choices()
            if e.bind.get("type") == SELECT_BIND_TYPE and e.type == MULTIPLE_SELECT_TYPE
        ][0]
        # pylint: disable=protected-access
        choices = export_builder._get_select_mulitples_choices(
            child, dd, ExportBuilder.GROUP_DELIMITER, ExportBuilder.TRUNCATE_GROUP_TITLE
        )
        expected_choices = [
            {
                "_label": "Nyekundu",
                "_label_xpath": "fav_colors/Nyekundu",
                "xpath": "children/fav_colors/red",
                "title": "children/fav_colors/red",
                "type": "string",
                "label": "fav_colors/Nyekundu",
            },
            {
                "_label": "Bluu",
                "_label_xpath": "fav_colors/Bluu",
                "xpath": "children/fav_colors/blue",
                "title": "children/fav_colors/blue",
                "type": "string",
                "label": "fav_colors/Bluu",
            },
            {
                "_label": "Pink",
                "_label_xpath": "fav_colors/Pink",
                "xpath": "children/fav_colors/pink",
                "title": "children/fav_colors/pink",
                "type": "string",
                "label": "fav_colors/Pink",
            },
        ]
        self.assertEqual(choices, expected_choices)
        select_multiples = {
            "children/fav_colors": [
                ("children/fav_colors/red", "red", "Nyekundu"),
                ("children/fav_colors/blue", "blue", "Bluu"),
                ("children/fav_colors/pink", "pink", "Pink"),
            ],
            "children/ice.creams": [
                ("children/ice.creams/vanilla", "vanilla", "Vanilla"),
                ("children/ice.creams/strawberry", "strawberry", "Strawberry"),
                ("children/ice.creams/chocolate", "chocolate", "Chocolate"),
            ],
        }
        self.assertEqual(
            CSVDataFrameBuilder._collect_select_multiples(dd), select_multiples
        )

    # pylint: disable=invalid-name
    def test_select_multiples_choices_with_choice_filter(self):
        survey = create_survey_from_xls(
            _logger_fixture_path("choice_filter.xlsx"), default_name="choice_filter"
        )
        dd = DataDictionary()
        # pylint: disable=protected-access
        dd._survey = survey
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        child = [
            e
            for e in dd.get_survey_elements_with_choices()
            if e.bind.get("type") == SELECT_BIND_TYPE and e.type == MULTIPLE_SELECT_TYPE
        ][0]
        # pylint: disable=protected-access
        choices = export_builder._get_select_mulitples_choices(
            child, dd, ExportBuilder.GROUP_DELIMITER, ExportBuilder.TRUNCATE_GROUP_TITLE
        )

        expected_choices = [
            {
                "_label": "King",
                "_label_xpath": "county/King",
                "label": "county/King",
                "title": "county/king",
                "type": "string",
                "xpath": "county/king",
            },
            {
                "_label": "Pierce",
                "_label_xpath": "county/Pierce",
                "label": "county/Pierce",
                "title": "county/pierce",
                "type": "string",
                "xpath": "county/pierce",
            },
            {
                "_label": "King",
                "_label_xpath": "county/King",
                "label": "county/King",
                "title": "county/king",
                "type": "string",
                "xpath": "county/king",
            },
            {
                "_label": "Cameron",
                "_label_xpath": "county/Cameron",
                "label": "county/Cameron",
                "title": "county/cameron",
                "type": "string",
                "xpath": "county/cameron",
            },
        ]
        self.assertEqual(choices, expected_choices)
        select_multiples = {
            "county": [
                ("county/king", "king", "King"),
                ("county/pierce", "pierce", "Pierce"),
                ("county/king", "king", "King"),
                ("county/cameron", "cameron", "Cameron"),
            ]
        }
        # pylint: disable=protected-access
        self.assertEqual(
            CSVDataFrameBuilder._collect_select_multiples(dd), select_multiples
        )

    # pylint: disable=invalid-name
    def test_string_to_date_with_xls_validation(self):
        # test "2016-11-02"
        val = string_to_date_with_xls_validation("2016-11-02")
        self.assertEqual(val, datetime.date(2016, 11, 2))

        # test random string
        val = string_to_date_with_xls_validation("random")
        self.assertEqual(val, "random")

        val = string_to_date_with_xls_validation(0.4)
        self.assertEqual(val, 0.4)

    def _create_osm_survey(self):
        """
        Creates survey for osm tests
        """
        # publish form

        osm_fixtures_dir = os.path.join(
            settings.PROJECT_ROOT, "apps", "api", "tests", "fixtures", "osm"
        )
        xlsform_path = os.path.join(osm_fixtures_dir, "osm.xlsx")
        self._publish_xls_file_and_set_xform(xlsform_path)

        # make submissions
        filenames = [
            "OSMWay234134797.osm",
            "OSMWay34298972.osm",
        ]
        paths = [os.path.join(osm_fixtures_dir, filename) for filename in filenames]
        submission_path = os.path.join(osm_fixtures_dir, "instance_a.xml")
        self._make_submission_w_attachment(submission_path, paths)
        survey = create_survey_from_xls(
            xlsform_path, default_name=xlsform_path.split("/")[-1].split(".")[0]
        )
        return survey

    # pylint: disable=invalid-name
    def test_zip_csv_export_has_submission_review_fields(self):
        """
        Test that review comment, status and date fields are in csv exports
        """
        self._create_user_and_login("dave", "1234")
        survey = self._create_osm_survey()
        xform = self.xform
        export_builder = ExportBuilder()
        export_builder.INCLUDE_REVIEW = True
        export_builder.set_survey(survey, xform, include_reviews=True)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        # check file's contents
        with open(os.path.join(temp_dir, "osm.csv"), encoding="utf-8") as csv_file:
            rows = list(csv.reader(csv_file))
            actual_headers = rows[0]
            self.assertIn(REVIEW_COMMENT, sorted(actual_headers))
            self.assertIn(REVIEW_DATE, sorted(actual_headers))
            self.assertIn(REVIEW_STATUS, sorted(actual_headers))
            submission = rows[1]
            self.assertEqual(submission[29], "Rejected")
            self.assertEqual(submission[30], "Wrong Location")
            self.assertEqual(submission[31], "2021-05-25T02:27:19")
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    # pylint: disable=invalid-name
    def test_xlsx_export_has_submission_review_fields(self):
        """
        Test that review comment, status and date fields are in xls exports
        """
        self._create_user_and_login("dave", "1234")
        survey = self._create_osm_survey()
        xform = self.xform
        export_builder = ExportBuilder()
        export_builder.INCLUDE_REVIEW = True
        export_builder.set_survey(survey, xform, include_reviews=True)
        with NamedTemporaryFile(suffix=".xlsx") as temp_xls_file:
            export_builder.to_xlsx_export(temp_xls_file.name, self.osm_data)
            temp_xls_file.seek(0)
            workbook = load_workbook(temp_xls_file.name)
            osm_data_sheet = workbook["osm"]
            rows = list(osm_data_sheet.rows)
            xls_headers = [a.value for a in rows[0]]
            xls_data = [a.value for a in rows[1]]
        self.assertIn(REVIEW_COMMENT, sorted(xls_headers))
        self.assertIn(REVIEW_DATE, sorted(xls_headers))
        self.assertIn(REVIEW_STATUS, sorted(xls_headers))
        self.assertEqual(xls_data[29], "Rejected")
        self.assertEqual(xls_data[30], "Wrong Location")
        self.assertEqual(xls_data[31], "2021-05-25T02:27:19")

    # pylint: disable=invalid-name
    def test_zipped_sav_has_submission_review_fields(self):
        """
        Test that review comment, status and date fields are in csv exports
        """
        self._create_user_and_login("dave", "1234")
        survey = self._create_osm_survey()
        xform = self.xform
        export_builder = ExportBuilder()
        export_builder.INCLUDE_REVIEW = True
        export_builder.set_survey(survey, xform, include_reviews=True)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, self.osm_data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        with SavReader(os.path.join(temp_dir, "osm.sav"), returnHeader=True) as reader:
            rows = list(reader)
            expected_column_headers = [
                "photo",
                "osm_road",
                "osm_building",
                "fav_color",
                "form_completed",
                "meta.instanceID",
                "@_id",
                "@_uuid",
                "@_submission_time",
                "@_index",
                "@_parent_table_name",
                "@_review_comment",
                f"@{REVIEW_DATE}",
                "@_review_status",
                "@_parent_index",
                "@_tags",
                "@_notes",
                "@_version",
                "@_duration",
                "@_submitted_by",
                "osm_road_ctr_lat",
                "osm_road_ctr_lon",
                "osm_road_highway",
                "osm_road_lanes",
                "osm_road_name",
                "osm_road_way_id",
                "osm_building_building",
                "osm_building_building_levels",
                "osm_building_ctr_lat",
                "osm_building_ctr_lon",
                "osm_building_name",
                "osm_building_way_id",
            ]
            actual_headers = list(map(_str_if_bytes, rows[0]))
            self.assertEqual(sorted(actual_headers), sorted(expected_column_headers))
            self.assertEqual(_str_if_bytes(rows[1][29]), "Rejected")
            self.assertEqual(_str_if_bytes(rows[1][30]), "Wrong Location")
            self.assertEqual(_str_if_bytes(rows[1][31]), "2021-05-25T02:27:19")

    # pylint: disable=invalid-name
    def test_zipped_csv_export_with_osm_data(self):
        """
        Tests that osm data is included in zipped csv export
        """
        survey = self._create_osm_survey()
        xform = self.xform
        export_builder = ExportBuilder()
        export_builder.set_survey(survey, xform)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_csv(temp_zip_file.name, self.osm_data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        with open(os.path.join(temp_dir, "osm.csv"), encoding="utf-8") as csv_file:
            rows = list(csv.reader(csv_file))

            expected_column_headers = [
                "photo",
                "osm_road",
                "osm_building",
                "fav_color",
                "form_completed",
                "meta/instanceID",
                "_id",
                "_uuid",
                "_submission_time",
                "_index",
                "_parent_table_name",
                "_parent_index",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
                "osm_road:ctr:lat",
                "osm_road:ctr:lon",
                "osm_road:highway",
                "osm_road:lanes",
                "osm_road:name",
                "osm_road:way:id",
                "osm_building:building",
                "osm_building:building:levels",
                "osm_building:ctr:lat",
                "osm_building:ctr:lon",
                "osm_building:name",
                "osm_building:way:id",
            ]

            self.assertEqual(sorted(rows[0]), sorted(expected_column_headers))
            self.assertEqual(rows[1][0], "1424308569120.jpg")
            self.assertEqual(rows[1][1], "OSMWay234134797.osm")
            self.assertEqual(rows[1][2], "23.708174238006087")
            self.assertEqual(rows[1][4], "tertiary")
            self.assertEqual(rows[1][6], "Patuatuli Road")
            self.assertEqual(rows[1][13], "kol")

    # pylint: disable=invalid-name
    def test_zipped_sav_export_with_osm_data(self):
        """
        Test that osm data is included in zipped sav export
        """
        survey = self._create_osm_survey()
        xform = self.xform
        osm_data = [
            {
                "photo": "1424308569120.jpg",
                "osm_road": "OSMWay234134797.osm",
                "osm_building": "OSMWay34298972.osm",
                "fav_color": "red",
                "osm_road_ctr_lat": "23.708174238006087",
                "osm_road_ctr_lon": "90.40946505581161",
                "osm_road_highway": "tertiary",
                "osm_road_lanes": "2",
                "osm_road_name": "Patuatuli Road",
                "osm_building_building": "yes",
                "osm_building_building_levels": "4",
                "osm_building_ctr_lat": "23.707316084046038",
                "osm_building_ctr_lon": "90.40849938337506",
                "osm_building_name": "kol",
            }
        ]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey, xform)
        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, osm_data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        with SavReader(os.path.join(temp_dir, "osm.sav"), returnHeader=True) as reader:
            rows = list(reader)
            expected_column_headers = [
                "photo",
                "osm_road",
                "osm_building",
                "fav_color",
                "form_completed",
                "meta.instanceID",
                "@_id",
                "@_uuid",
                "@_submission_time",
                "@_index",
                "@_parent_table_name",
                "@_parent_index",
                "@_tags",
                "@_notes",
                "@_version",
                "@_duration",
                "@_submitted_by",
                "osm_road_ctr_lat",
                "osm_road_ctr_lon",
                "osm_road_highway",
                "osm_road_lanes",
                "osm_road_name",
                "osm_road_way_id",
                "osm_building_building",
                "osm_building_building_levels",
                "osm_building_ctr_lat",
                "osm_building_ctr_lon",
                "osm_building_name",
                "osm_building_way_id",
            ]
            rows[0] = list(map(_str_if_bytes, rows[0]))
            rows[1] = list(map(_str_if_bytes, rows[1]))
            self.assertEqual(sorted(rows[0]), sorted(expected_column_headers))
            self.assertEqual(rows[1][0], "1424308569120.jpg")
            self.assertEqual(rows[1][1], "OSMWay234134797.osm")
            self.assertEqual(rows[1][2], "23.708174238006087")
            self.assertEqual(rows[1][4], "tertiary")
            self.assertEqual(rows[1][6], "Patuatuli Road")
            self.assertEqual(rows[1][13], "kol")

    def test_show_choice_labels(self):
        """
        Test show_choice_labels=true for select one questions.
        """
        md_xform = """
        | survey  |
        |         | type              | name  | label  |
        |         | text              | name  | Name   |
        |         | integer           | age   | Age    |
        |         | select one fruits | fruit | Fruit  |
        |         |                   |       |        |
        | choices | list name         | name  | label  |
        |         | fruits            | 1     | Mango  |
        |         | fruits            | 2     | Orange |
        |         | fruits            | 3     | Apple  |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:3]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [["name", "age", "fruit"], ["Maria", 25, "Mango"]]

        self.assertEqual(expected_result, result)

    def test_show_choice_labels_multi_language(self):  # pylint: disable=invalid-name
        """
        Test show_choice_labels=true for select one questions - multi language
        form.
        """
        md_xform = """
        | survey  |
        |         | type              | name  | label:English | label:French |
        |         | text              | name  | Name          | Prénom       |
        |         | integer           | age   | Age           | Âge          |
        |         | select one fruits | fruit | Fruit         | Fruit        |
        |         |                   |       |               |              |
        | choices | list name         | name  | label:English | label:French |
        |         | fruits            | 1     | Mango         | Mangue       |
        |         | fruits            | 2     | Orange        | Orange       |
        |         | fruits            | 3     | Apple         | Pomme        |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.language = "French"
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:3]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [["name", "age", "fruit"], ["Maria", 25, "Mangue"]]

        self.assertEqual(expected_result, result)

    def test_show_choice_labels_select_multiple(self):  # pylint: disable=invalid-name
        """
        Test show_choice_labels=true for select multiple questions -
        split_select_multiples=true.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label  |
        |         | text                   | name  | Name   |
        |         | integer                | age   | Age    |
        |         | select_multiple fruits | fruit | Fruit  |
        |         |                        |       |        |
        | choices | list name              | name  | label  |
        |         | fruits                 | 1     | Mango  |
        |         | fruits                 | 2     | Orange |
        |         | fruits                 | 3     | Apple  |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 2"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:3]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [["name", "age", "fruit"], ["Maria", 25, "Mango Orange"]]

        self.assertEqual(expected_result, result)

    # pylint: disable=invalid-name
    def test_show_choice_labels_select_multiple_1(self):
        """
        Test show_choice_labels=true for select multiple questions
        split_select_multiples=false.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label  |
        |         | text                   | name  | Name   |
        |         | integer                | age   | Age    |
        |         | select_multiple fruits | fruit | Fruit  |
        |         |                        |       |        |
        | choices | list name              | name  | label  |
        |         | fruits                 | 1     | Mango  |
        |         | fruits                 | 2     | Orange |
        |         | fruits                 | 3     | Apple  |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = False
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 2"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:3]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [["name", "age", "fruit"], ["Maria", 25, "Mango Orange"]]

        self.assertEqual(expected_result, result)

    # pylint: disable=invalid-name
    def test_show_choice_labels_select_multiple_2(self):
        """
        Test show_choice_labels=true for select multiple questions
        split_select_multiples=false.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label  |
        |         | text                   | name  | Name   |
        |         | integer                | age   | Age    |
        |         | select_multiple fruits | fruit | Fruit  |
        |         |                        |       |        |
        | choices | list name              | name  | label  |
        |         | fruits                 | 1     | Mango  |
        |         | fruits                 | 2     | Orange |
        |         | fruits                 | 3     | Apple  |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = True
        export_builder.VALUE_SELECT_MULTIPLES = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 2"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:6]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [
            ["name", "age", "fruit", "fruit/Mango", "fruit/Orange", "fruit/Apple"],
            ["Maria", 25, "Mango Orange", "Mango", "Orange", None],
        ]

        self.maxDiff = None
        self.assertEqual(expected_result, result)

    # pylint: disable=invalid-name
    def test_show_choice_labels_select_multiple_language(self):
        """
        Test show_choice_labels=true for select multiple questions - multi
        language form.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label:Eng  | label:Fr |
        |         | text                   | name  | Name       | Prénom   |
        |         | integer                | age   | Age        | Âge      |
        |         | select_multiple fruits | fruit | Fruit      | Fruit    |
        |         |                        |       |            |          |
        | choices | list name              | name  | label:Eng  | label:Fr |
        |         | fruits                 | 1     | Mango      | Mangue   |
        |         | fruits                 | 2     | Orange     | Orange   |
        |         | fruits                 | 3     | Apple      | Pomme    |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = False
        export_builder.language = "Fr"
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 3"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:3]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [["name", "age", "fruit"], ["Maria", 25, "Mangue Pomme"]]

        self.assertEqual(expected_result, result)

    # pylint: disable=invalid-name
    def test_show_choice_labels_select_multiple_language_1(self):
        """
        Test show_choice_labels=true, split_select_multiples=true, for select
        multiple questions - multi language form.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label:Eng  | label:Fr |
        |         | text                   | name  | Name       | Prénom   |
        |         | integer                | age   | Age        | Âge      |
        |         | select_multiple fruits | fruit | Fruit      | Fruit    |
        |         |                        |       |            |          |
        | choices | list name              | name  | label:Eng  | label:Fr |
        |         | fruits                 | 1     | Mango      | Mangue   |
        |         | fruits                 | 2     | Orange     | Orange   |
        |         | fruits                 | 3     | Apple      | Pomme    |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = True
        export_builder.language = "Fr"
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 3"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:6]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [
            ["name", "age", "fruit", "fruit/Mangue", "fruit/Orange", "fruit/Pomme"],
            ["Maria", 25, "Mangue Pomme", True, False, True],
        ]

        self.assertEqual(expected_result, result)

    # pylint: disable=invalid-name
    def test_show_choice_labels_select_multiple_language_2(self):
        """
        Test show_choice_labels=true, split_select_multiples=true,
        binary_select_multiples=true for select multiple questions - multi
        language form.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label:Eng  | label:Fr |
        |         | text                   | name  | Name       | Prénom   |
        |         | integer                | age   | Age        | Âge      |
        |         | select_multiple fruits | fruit | Fruit      | Fruit    |
        |         |                        |       |            |          |
        | choices | list name              | name  | label:Eng  | label:Fr |
        |         | fruits                 | 1     | Mango      | Mangue   |
        |         | fruits                 | 2     | Orange     | Orange   |
        |         | fruits                 | 3     | Apple      | Pomme    |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = True
        export_builder.BINARY_SELECT_MULTIPLES = True
        export_builder.language = "Fr"
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 3"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:6]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [
            ["name", "age", "fruit", "fruit/Mangue", "fruit/Orange", "fruit/Pomme"],
            ["Maria", 25, "Mangue Pomme", 1, 0, 1],
        ]

        self.assertEqual(expected_result, result)

    # pylint: disable=invalid-name
    def test_show_choice_labels_select_multiple_language_3(self):
        """
        Test show_choice_labels=true, split_select_multiples=true,
        value_select_multiples=true for select multiple questions - multi
        language form.
        """
        md_xform = """
        | survey  |
        |         | type                   | name  | label:Eng  | label:Fr |
        |         | text                   | name  | Name       | Prénom   |
        |         | integer                | age   | Age        | Âge      |
        |         | select_multiple fruits | fruit | Fruit      | Fruit    |
        |         |                        |       |            |          |
        | choices | list name              | name  | label:Eng  | label:Fr |
        |         | fruits                 | 1     | Mango      | Mangue   |
        |         | fruits                 | 2     | Orange     | Orange   |
        |         | fruits                 | 3     | Apple      | Pomme    |
        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = True
        export_builder.VALUE_SELECT_MULTIPLES = True
        export_builder.language = "Fr"
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")
        data = [{"name": "Maria", "age": 25, "fruit": "1 3"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)
        result = [[col.value for col in row[:6]] for row in children_sheet.rows]
        temp_xls_file.close()
        expected_result = [
            ["name", "age", "fruit", "fruit/Mangue", "fruit/Orange", "fruit/Pomme"],
            ["Maria", 25, "Mangue Pomme", "Mangue", None, "Pomme"],
        ]

        self.assertEqual(expected_result, result)

    def test_mulsel_export_with_label_choices(self):
        """test_mulsel_export_with_label_choices"""
        md_xform = """
        | survey |
        |        | type                  | name         | label                                         | choice_filter               | # noqa
        |        | select_multiple banks | banks_deal   | Which banks do you deal with?                 |                             | # noqa
        |        | select_one primary    | primary_bank | Which bank do you consider your primary bank? | selected(${banks_deal}, cf) | # noqa
        | choices |
        |         | list_name | name | label              | cf |
        |         | banks     | 1    | KCB                |    |
        |         | banks     | 2    | Equity             |    |
        |         | banks     | 3    | Co-operative       |    |
        |         | banks     | 4    | CBA                |    |
        |         | banks     | 5    | Stanbic            |    |
        |         | banks     | 6    | City Bank          |    |
        |         | banks     | 7    | Barclays           |    |
        |         | banks     | 8    | Standard Chattered |    |
        |         | banks     | 9    | Other bank         |    |
        |         | primary   | 1    | KCB                | 1  |
        |         | primary   | 2    | Equity             | 2  |
        |         | primary   | 3    | Co-operative       | 3  |
        |         | primary   | 4    | CBA                | 4  |
        |         | primary   | 5    | Stanbic            | 5  |
        |         | primary   | 6    | City Bank          | 6  |
        |         | primary   | 7    | Barclays           | 7  |
        |         | primary   | 8    | Standard Chattered | 8  |
        |         | primary   | 9    | Other bank         | 9  |

        """
        survey = self.md_to_pyxform_survey(md_xform, {"name": "data"})
        export_builder = ExportBuilder()
        export_builder.SHOW_CHOICE_LABELS = True
        export_builder.SPLIT_SELECT_MULTIPLES = False
        export_builder.VALUE_SELECT_MULTIPLES = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix=".xlsx")

        data = [{"banks_deal": "1 2 3 4", "primary_bank": "3"}]  # yapf: disable
        export_builder.to_xlsx_export(temp_xls_file, data)
        temp_xls_file.seek(0)
        children_sheet = load_workbook(temp_xls_file)["data"]
        self.assertTrue(children_sheet)

        result = [[col.value for col in row[:2]] for row in children_sheet.rows]
        expected_result = [
            ["banks_deal", "primary_bank"],
            ["KCB Equity Co-operative CBA", "Co-operative"],
        ]
        temp_xls_file.close()
        self.assertEqual(result, expected_result)

    def test_no_group_name_gps_data(self):
        """
        Test that GPS Data is correctly exported when GPS Fields are within
        groups and the group name is removed
        """
        self._publish_xls_file_and_set_xform(_logger_fixture_path("gps_data.xlsx"))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(self.xform.survey, self.xform)

        # Submit test record
        self._make_submission(_logger_fixture_path("gps_data.xml"))
        self.assertEqual(self.xform.instances.count(), 1)
        csv_export = NamedTemporaryFile(suffix=".csv")
        records = self.xform.instances.all().order_by("id")
        inst_json = records.first().json
        csv_data = records.values_list("json", flat=True).iterator()
        export_builder.to_flat_csv_export(
            csv_export.name,
            csv_data,
            self.xform.user.username,
            self.xform.id_string,
            "",
            xform=self.xform,
            options={},
        )
        expected_data = [
            [
                "start",
                "end",
                "test",
                "Age",
                "Image",
                "Select_one_gender",
                "respondent_name",
                "_1_2_Where_do_you_live",
                "_1_3_Another_test",
                "GPS",
                "_GPS_latitude",
                "_GPS_longitude",
                "_GPS_altitude",
                "_GPS_precision",
                "Another_GPS",
                "_Another_GPS_latitude",
                "_Another_GPS_longitude",
                "_Another_GPS_altitude",
                "_Another_GPS_precision",
                "__version__",
                "instanceID",
                "_id",
                "_uuid",
                "_submission_time",
                "_date_modified",
                "_tags",
                "_notes",
                "_version",
                "_duration",
                "_submitted_by",
                "_total_media",
                "_media_count",
                "_media_all_received",
            ],
            [
                "2023-01-20T12:11:29.278+03:00",
                "2023-01-20T12:12:00.443+03:00",
                "2",
                "32",
                "n/a",
                "male",
                "dsa",
                "zcxsd",
                "n/a",
                "-1.248238 36.685681 0 0",
                "-1.248238",
                "36.685681",
                "0",
                "0",
                "-1.244172 36.685359 0 0",
                "-1.244172",
                "36.685359",
                "0",
                "0",
                inst_json.get("_version"),
                "uuid:" + inst_json.get("_uuid"),
                str(inst_json.get("_id")),
                inst_json.get("_uuid"),
                inst_json.get("_submission_time"),
                inst_json.get("_date_modified"),
                "",
                "",
                "vTEmiygu2uLZpPHBYX8jKj",
                "31.0",
                "bob",
                "0",
                "0",
                "True",
            ],
        ]
        csv_file = open(csv_export.name)
        csvreader = csv.reader(csv_file)
        for idx, row in enumerate(csvreader):
            self.assertEqual(row, expected_data[idx])

        csv_export.close()

    @patch("onadata.libs.utils.export_builder.uuid")
    def test_sav_export_with_duplicate_metadata(self, mock_uuid):
        """
        Test that SAV exports work when a GPS Question is both in a group
        and outside of one
        """
        md = """
        | survey |
        |        | type              | name    | label |
        |        | begin group       | group_a | A     |
        |        | gps               | gps     | GPS   |
        |        | end_group         | group_a | A     |
        |        | gps               | gps     | GPS B |
        """
        name = "data_tsewdm"
        survey = self.md_to_pyxform_survey(md, {"name": name})
        data = [
            {
                "group_a/gps": "4.0 36.1 5000 20",
                "gps": "1.0 36.1 2000 20",
                "_submission_time": "2016-11-21T03:42:43.00-08:00",
            }
        ]
        mock_uuid.uuid4.return_value = "9366af30-52a9-46d0-9cc8-1c6411280c84"
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)

        with NamedTemporaryFile(suffix=".zip") as temp_zip_file:
            export_builder.to_zipped_sav(temp_zip_file.name, data)
            temp_zip_file.seek(0)
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_zip_file.name, "r") as zip_file:
                zip_file.extractall(temp_dir)

        expected_data = [
            [
                "gps",
                "@_gps_latitude",
                "@_gps_longitude",
                "@_gps_altitude",
                "@_gps_precision",
                "gps@52a9",
                "@_gps_latitude_52a9",
                "@_gps_longitude_52a9",
                "@_gps_altitude_52a9",
                "@_gps_precision_52a9",
                "instanceID",
                "@_id",
                "@_uuid",
                "@_submission_time",
                "@_index",
                "@_parent_table_name",
                "@_parent_index",
                "@_tags",
                "@_notes",
                "@_version",
                "@_duration",
                "@_submitted_by",
            ],
            [
                "4.0 36.1 5000 20",
                4.0,
                36.1,
                5000.0,
                20.0,
                "1.0 36.1 2000 20",
                1.0,
                36.1,
                2000.0,
                20.0,
                "",
                None,
                "",
                "2016-11-21 03:42:43",
                1.0,
                "",
                -1.0,
                "",
                "",
                "",
                "",
                "",
            ],
        ]
        with SavReader(
            os.path.join(temp_dir, f"{name}.sav"), returnHeader=True
        ) as reader:
            rows = list(reader)
            self.assertEqual(len(rows), 2)
            rows[0] = list(map(_str_if_bytes, rows[0]))
            rows[1] = list(map(_str_if_bytes, rows[1]))
            self.assertEqual(expected_data, rows)
        shutil.rmtree(temp_dir)

    def test_extra_columns_dataview(self):
        """Extra columns are included in export for dataview

        Extra columns included only if in the dataview
        """
        self._publish_xls_file_and_set_xform(
            _logger_fixture_path("childrens_survey.xlsx")
        )
        export_builder = ExportBuilder()
        export_builder.set_survey(self.xform.survey)
        extra_cols = [
            "_id",
            "_uuid",
            "_submission_time",
            "_index",
            "_parent_table_name",
            "_parent_index",
            "_tags",
            "_notes",
            "_version",
            "_duration",
            "_submitted_by",
        ]

        for extra_col in extra_cols:
            dataview = DataView.objects.create(
                xform=self.xform,
                name="test",
                columns=["name", extra_col],
                project=self.project,
            )
            fields = export_builder.get_fields(
                dataview, export_builder.sections[0], "title"
            )
            self.assertEqual(fields, ["name", extra_col])
