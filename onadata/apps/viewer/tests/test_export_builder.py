import csv
import datetime
import os
import shutil
import tempfile
import xlrd
import zipfile

from collections import OrderedDict
from django.conf import settings
from django.core.files.temp import NamedTemporaryFile
from openpyxl import load_workbook
from pyxform.builder import create_survey_from_xls
from pyxform.tests_v1.pyxform_test_case import PyxformTestCase
from savReaderWriter import SavReader
from savReaderWriter import SavHeaderReader

from onadata.apps.main.tests.test_base import TestBase
from onadata.apps.viewer.models.data_dictionary import DataDictionary
from onadata.apps.viewer.models.parsed_instance import _encode_for_mongo
from onadata.apps.viewer.tests.export_helpers import viewer_fixture_path
from onadata.libs.utils.export_builder import dict_to_joined_export
from onadata.libs.utils.export_tools import ExportBuilder, get_columns_with_hxl
from onadata.libs.utils.csv_builder import CSVDataFrameBuilder
from onadata.libs.utils.csv_builder import get_labels_from_columns


def _logger_fixture_path(*args):
    return os.path.join(settings.PROJECT_ROOT, 'apps', 'logger',
                        'tests', 'fixtures', *args)


class TestExportBuilder(PyxformTestCase, TestBase):
    data = [
        {
            'name': 'Abe',
            'age': 35,
            'tel/telLg==office': '020123456',
            'children':
            [
                {
                    'children/name': 'Mike',
                    'children/age': 5,
                    'children/fav_colors': 'red blue',
                    'children/iceLg==creams': 'vanilla chocolate',
                    'children/cartoons':
                    [
                        {
                            'children/cartoons/name': 'Tom & Jerry',
                            'children/cartoons/why': 'Tom is silly',
                        },
                        {
                            'children/cartoons/name': 'Flinstones',
                            'children/cartoons/why': u"I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                },
                {
                    'children/name': 'John',
                    'children/age': 2,
                    'children/cartoons': []
                },
                {
                    'children/name': 'Imora',
                    'children/age': 3,
                    'children/cartoons':
                    [
                        {
                            'children/cartoons/name': 'Shrek',
                            'children/cartoons/why': 'He\'s so funny'
                        },
                        {
                            'children/cartoons/name': 'Dexter\'s Lab',
                            'children/cartoons/why': 'He thinks hes smart',
                            'children/cartoons/characters':
                            [
                                {
                                    'children/cartoons/characters/name':
                                    'Dee Dee',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'good'
                                },
                                {
                                    'children/cartoons/characters/name':
                                    'Dexter',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'evil'
                                },
                            ]
                        }
                    ]
                }
            ]
        },
        {
            # blank data just to be sure
            'children': []
        }
    ]
    long_survey_data = [
        {
            'name': 'Abe',
            'age': 35,
            'childrens_survey_with_a_very_lo':
            [
                {
                    'childrens_survey_with_a_very_lo/name': 'Mike',
                    'childrens_survey_with_a_very_lo/age': 5,
                    'childrens_survey_with_a_very_lo/fav_colors': 'red blue',
                    'childrens_survey_with_a_very_lo/cartoons':
                    [
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Tom & Jerry',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            'Tom is silly',
                        },
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Flinstones',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            u"I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                },
                {
                    'childrens_survey_with_a_very_lo/name': 'John',
                    'childrens_survey_with_a_very_lo/age': 2,
                    'childrens_survey_with_a_very_lo/cartoons': []
                },
                {
                    'childrens_survey_with_a_very_lo/name': 'Imora',
                    'childrens_survey_with_a_very_lo/age': 3,
                    'childrens_survey_with_a_very_lo/cartoons':
                    [
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Shrek',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            'He\'s so funny'
                        },
                        {
                            'childrens_survey_with_a_very_lo/cartoons/name':
                            'Dexter\'s Lab',
                            'childrens_survey_with_a_very_lo/cartoons/why':
                            'He thinks hes smart',
                            'childrens_survey_with_a_very_lo/cartoons/characte'
                            'rs':
                            [
                                {
                                    'childrens_survey_with_a_very_lo/cartoons/'
                                    'characters/name': 'Dee Dee',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'good'
                                },
                                {
                                    'childrens_survey_with_a_very_lo/cartoons/'
                                    'characters/name': 'Dexter',
                                    'children/cartoons/characters/good_or_evi'
                                    'l': 'evil'
                                },
                            ]
                        }
                    ]
                }
            ]
        }
    ]
    data_utf8 = [
        {
            'name': 'Abe',
            'age': 35,
            'tel/telLg==office': '020123456',
            'childrenLg==info':
            [
                {
                    'childrenLg==info/nameLg==first': 'Mike',
                    'childrenLg==info/age': 5,
                    'childrenLg==info/fav_colors': u'red\'s blue\'s',
                    'childrenLg==info/ice_creams': 'vanilla chocolate',
                    'childrenLg==info/cartoons':
                    [
                        {
                            'childrenLg==info/cartoons/name': 'Tom & Jerry',
                            'childrenLg==info/cartoons/why': 'Tom is silly',
                        },
                        {
                            'childrenLg==info/cartoons/name': 'Flinstones',
                            'childrenLg==info/cartoons/why':
                            u"I like bam bam\u0107"
                            # throw in a unicode character
                        }
                    ]
                }
            ]
        }
    ]

    def _create_childrens_survey(self, filename="childrens_survey.xls"):
        survey = create_survey_from_xls(_logger_fixture_path(
            filename
        ))
        self.dd = DataDictionary()
        self.dd._survey = survey

        return survey

    def test_build_sections_from_survey(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # test that we generate the proper sections
        expected_sections = [
            survey.name, 'children', 'children/cartoons',
            'children/cartoons/characters']
        self.assertEqual(
            expected_sections, [s['name'] for s in export_builder.sections])
        # main section should have split geolocations
        expected_element_names = [
            'name', 'age', 'geo/geolocation', 'geo/_geolocation_longitude',
            'geo/_geolocation_latitude', 'geo/_geolocation_altitude',
            'geo/_geolocation_precision', 'tel/tel.office', 'tel/tel.mobile',
            'meta/instanceID']
        section = export_builder.section_by_name(survey.name)
        element_names = [element['xpath'] for element in section['elements']]
        # fav_colors should have its choices split
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/name', 'children/age', 'children/fav_colors',
            'children/fav_colors/red', 'children/fav_colors/blue',
            'children/fav_colors/pink', 'children/ice.creams',
            'children/ice.creams/vanilla', 'children/ice.creams/strawberry',
            'children/ice.creams/chocolate']
        section = export_builder.section_by_name('children')
        element_names = [element['xpath'] for element in section['elements']]
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/cartoons/name', 'children/cartoons/why']
        section = export_builder.section_by_name('children/cartoons')
        element_names = [element['xpath'] for element in section['elements']]

        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

        expected_element_names = [
            'children/cartoons/characters/name',
            'children/cartoons/characters/good_or_evil']
        section = \
            export_builder.section_by_name('children/cartoons/characters')
        element_names = [element['xpath'] for element in section['elements']]
        self.assertEqual(
            sorted(expected_element_names), sorted(element_names))

    def test_zipped_csv_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(
                    d, index, indices, survey_name, survey, d))
            index += 1

        # check that each file exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "{0}.csv".format(survey.name))))
        with open(
                os.path.join(
                    temp_dir, "{0}.csv".format(survey.name))) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(_logger_fixture_path(
                    'csvs', 'childrens_survey.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.csv")))
        with open(os.path.join(temp_dir, "children.csv")) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(_logger_fixture_path(
                    'csvs', 'children.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children_cartoons.csv")))
        with open(os.path.join(temp_dir, "children_cartoons.csv")) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(_logger_fixture_path(
                    'csvs', 'children_cartoons.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children_cartoons_characters.csv")))
        with open(os.path.join(
                temp_dir, "children_cartoons_characters.csv")) as csv_file:
            reader = csv.reader(csv_file)
            rows = [r for r in reader]

            # open comparison file
            with open(_logger_fixture_path(
                    'csvs',
                    'children_cartoons_characters.csv')) as fixture_csv:
                fixture_reader = csv.reader(fixture_csv)
                expected_rows = [r for r in fixture_reader]
                self.assertEqual(rows, expected_rows)

        shutil.rmtree(temp_dir)

    def test_decode_mongo_encoded_section_names(self):
        data = {
            'main_section': [1, 2, 3, 4],
            'sectionLg==1/info': [1, 2, 3, 4],
            'sectionLg==2/info': [1, 2, 3, 4],
        }
        result = ExportBuilder.decode_mongo_encoded_section_names(data)
        expected_result = {
            'main_section': [1, 2, 3, 4],
            'section.1/info': [1, 2, 3, 4],
            'section.2/info': [1, 2, 3, 4],
        }
        self.assertEqual(result, expected_result)

    def test_zipped_csv_export_works_with_unicode(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(os.path.join(temp_dir, "children.info.csv")) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = ['children.info/name.first',
                                'children.info/age',
                                'children.info/fav_colors',
                                u'children.info/fav_colors/red\'s',
                                u'children.info/fav_colors/blue\'s',
                                u'children.info/fav_colors/pink\'s',
                                'children.info/ice_creams',
                                'children.info/ice_creams/vanilla',
                                'children.info/ice_creams/strawberry',
                                'children.info/ice_creams/chocolate', '_id',
                                '_uuid', '_submission_time', '_index',
                                '_parent_table_name', '_parent_index',
                                u'_tags', '_notes', '_version',
                                '_duration', '_submitted_by']
            rows = [row for row in reader]
            actual_headers = [h.decode('utf-8') for h in rows[0]]
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(
                data[u'children.info/fav_colors/red\'s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'children.info/fav_colors/blue\'s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'children.info/fav_colors/pink\'s'.encode('utf-8')],
                'False')
            # check that red and blue are set to true

    def test_zipped_sav_export_with_date_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | date              | expense_date | Expense Date |
        |        | begin group       | A            | A group      |
        |        | date              | gdate        | Good Day     |
        |        | end group         |              |              |

        | choices |
        |         | list name | name   | label  |
        """
        survey = self.md_to_pyxform_survey(md, {'name': 'exp'})
        data = [{"expense_date": "2013-01-03", "A/gdate": "2017-06-13",
                 '_submission_time': u'2016-11-21T03:43:43.000-08:00'}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_sav(temp_zip_file.name, data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"),
                       returnHeader=True) as reader:
            rows = [r for r in reader]
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[0][0],  'expense_date')
            self.assertEqual(rows[1][0],  '2013-01-03')
            self.assertEqual(rows[0][1],  'A.gdate')
            self.assertEqual(rows[1][1],  '2017-06-13')
            self.assertEqual(rows[0][5], '@_submission_time')
            self.assertEqual(rows[1][5], '2016-11-21 03:43:43')

        shutil.rmtree(temp_dir)

    def test_zipped_sav_export_with_zero_padded_select_one_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | select one yes_no | expensed     | Expensed?    |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 1      | Yes    |
        |         | yes_no    | 09      | No     |
        """
        survey = self.md_to_pyxform_survey(md, {'name': 'exp'})
        data = [{"expensed": "09",
                 '_submission_time': u'2016-11-21T03:43:43.000-08:00'}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_sav(temp_zip_file.name, data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"),
                       returnHeader=True) as reader:
            rows = [r for r in reader]
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[1][0],  "09")
            self.assertEqual(rows[1][4], '2016-11-21 03:43:43')

    def test_zipped_sav_export_with_numeric_select_one_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | select one yes_no | expensed     | Expensed?    |
        |        | begin group       | A            | A            |
        |        | select one yes_no | q1           | Q1           |
        |        | end group         |              |              |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 1      | Yes    |
        |         | yes_no    | 0      | No     |
        """
        survey = self.md_to_pyxform_survey(md, {'name': 'exp'})
        data = [{"expensed": "1", "A/q1": "1",
                 '_submission_time': u'2016-11-21T03:43:43.000-08:00'}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_sav(temp_zip_file.name, data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"),
                       returnHeader=True) as reader:
            rows = [r for r in reader]
            self.assertTrue(len(rows) > 1)

            # expensed 1
            self.assertEqual(rows[0][0],  'expensed')
            self.assertEqual(rows[1][0],  1)

            # A/q1 1
            self.assertEqual(rows[0][1],  'A.q1')
            self.assertEqual(rows[1][1],  1)

            # _submission_time is a date string
            self.assertEqual(rows[0][5], '@_submission_time')
            self.assertEqual(rows[1][5], '2016-11-21 03:43:43')

    def test_zipped_sav_export_with_numeric_select_multiple_field(self):
        md = """
        | survey |
        |        | type                   | name         | label        |
        |        | select_multiple yes_no | expensed     | Expensed?    |
        |        | begin group            | A            | A            |
        |        | select_multiple yes_no | q1           | Q1           |
        |        | end group              |              |              |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 1      | Yes    |
        |         | yes_no    | 0      | No     |
        """
        survey = self.md_to_pyxform_survey(md, {'name': 'exp'})
        data = [{"expensed": "1", "A/q1": "1",
                 '_submission_time': u'2016-11-21T03:43:43.000-08:00'}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_sav(temp_zip_file.name, data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"),
                       returnHeader=True) as reader:
            rows = [r for r in reader]
            self.assertTrue(len(rows) > 1)

            self.assertEqual(rows[0][0],  "expensed")
            self.assertEqual(rows[1][0],  "1")

            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(rows[0][1],  "expensed.1")
            self.assertEqual(rows[1][1], 1)

            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(rows[0][2],  "expensed.0")
            self.assertEqual(rows[1][2], 0)

            self.assertEqual(rows[0][3],  "A.q1")
            self.assertEqual(rows[1][3],  "1")

            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(rows[0][4],  "A.q1.1")
            self.assertEqual(rows[1][4], 1)

            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(rows[0][5],  "A.q1.0")
            self.assertEqual(rows[1][5], 0)

            self.assertEqual(rows[0][9],  "@_submission_time")
            self.assertEqual(rows[1][9], '2016-11-21 03:43:43')

        shutil.rmtree(temp_dir)

    def test_zipped_sav_export_with_zero_padded_select_multiple_field(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | select_multiple yes_no | expensed     | Expensed?    |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 1      | Yes    |
        |         | yes_no    | 09     | No     |
        """
        survey = self.md_to_pyxform_survey(md, {'name': 'exp'})
        data = [{"expensed": "1",
                 '_submission_time': u'2016-11-21T03:43:43.000-08:00'}]
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_sav(temp_zip_file.name, data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"),
                       returnHeader=True) as reader:
            rows = [r for r in reader]
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[1][0],  "1")
            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(rows[1][1], 1)
            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(rows[1][2], 0)
            self.assertEqual(rows[1][6], '2016-11-21 03:43:43')

        shutil.rmtree(temp_dir)

    def test_zipped_sav_export_with_values_split_select_multiple(self):
        md = """
        | survey |
        |        | type              | name         | label        |
        |        | select_multiple yes_no | expensed     | Expensed?    |

        | choices |
        |         | list name | name   | label  |
        |         | yes_no    | 2      | Yes    |
        |         | yes_no    | 09     | No     |
        """
        survey = self.md_to_pyxform_survey(md, {'name': 'exp'})
        data = [{"expensed": "2 09",
                 '_submission_time': u'2016-11-21T03:43:43.000-08:00'}]
        export_builder = ExportBuilder()
        export_builder.VALUE_SELECT_MULTIPLES = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_sav(temp_zip_file.name, data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "exp.sav")))
        # check file's contents

        with SavReader(os.path.join(temp_dir, "exp.sav"),
                       returnHeader=True) as reader:
            rows = [r for r in reader]
            self.assertTrue(len(rows) > 1)
            self.assertEqual(rows[1][0],  "2 09")
            # expensed.1 is selected hence True, 1.00 or 1 in SPSS
            self.assertEqual(rows[1][1], 2)
            # expensed.0 is not selected hence False, .00 or 0 in SPSS
            self.assertEqual(rows[1][2], '09')
            self.assertEqual(rows[1][6], '2016-11-21 03:43:43')

        shutil.rmtree(temp_dir)

    def test_xls_export_works_with_unicode(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data_utf8)
        temp_xls_file.seek(0)
        # check that values for red\'s and blue\'s are set to true
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb.get_sheet_by_name("children.info")
        data = dict([(r[0].value, r[1].value) for r in children_sheet.columns])
        self.assertTrue(data[u'children.info/fav_colors/red\'s'])
        self.assertTrue(data[u'children.info/fav_colors/blue\'s'])
        self.assertFalse(data[u'children.info/fav_colors/pink\'s'])
        temp_xls_file.close()

    def test_xls_export_with_hxl_adds_extra_row(self):
        # hxl_example.xlsx contains `instance::hxl` column whose value is #age
        xlsform_path = os.path.join(
            settings.PROJECT_ROOT, "apps", "main", "tests", "fixtures",
            "hxl_test", "hxl_example.xlsx")
        survey = create_survey_from_xls(xlsform_path)
        export_builder = ExportBuilder()
        export_builder.INCLUDE_HXL = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')

        survey_elements = [
            survey_item[1]
            for survey_item in survey.items()
            if survey_item[0] == u'children'
        ][0]

        columns_with_hxl = export_builder.INCLUDE_HXL and get_columns_with_hxl(
            survey_elements
        )

        export_builder.to_xls_export(
            temp_xls_file.name, self.data_utf8,
            columns_with_hxl=columns_with_hxl)
        temp_xls_file.seek(0)
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb.get_sheet_by_name("hxl_example")
        self.assertTrue(children_sheet)

        # we pick the second row because the first row has xform fieldnames
        rows = [row for row in children_sheet.rows]
        hxl_row = [a.value for a in rows[1]]
        self.assertIn(u'#age', hxl_row)

    def test_generation_of_multi_selects_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_select_multiples =\
            {
                'children':
                {
                    'children/fav_colors':
                    [
                        'children/fav_colors/red', 'children/fav_colors/blue',
                        'children/fav_colors/pink'
                    ],
                    'children/ice.creams':
                    [
                        'children/ice.creams/vanilla',
                        'children/ice.creams/strawberry',
                        'children/ice.creams/chocolate'
                    ]
                }
            }
        select_multiples = export_builder.select_multiples
        self.assertTrue('children' in select_multiples)
        self.assertTrue('children/fav_colors' in select_multiples['children'])
        self.assertTrue('children/ice.creams' in select_multiples['children'])
        self.assertEqual(
            sorted(select_multiples['children']['children/fav_colors']),
            sorted(
                expected_select_multiples['children']['children/fav_colors']))
        self.assertEqual(
            sorted(select_multiples['children']['children/ice.creams']),
            sorted(
                expected_select_multiples['children']['children/ice.creams']))

    def test_split_select_multiples_works(self):
        select_multiples =\
            {
                'children/fav_colors': [
                    'children/fav_colors/red', 'children/fav_colors/blue',
                    'children/fav_colors/pink']
            }
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': 'red blue'
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': 'red blue',
                'children/fav_colors/red': True,
                'children/fav_colors/blue': True,
                'children/fav_colors/pink': False
            }
        self.assertEqual(new_row, expected_row)
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors/red': None,
                'children/fav_colors/blue': None,
                'children/fav_colors/pink': None
            }
        self.assertEqual(new_row, expected_row)

    def test_split_select_mutliples_works_with_int_value_in_row(self):
        select_multiples = {
            'children/fav_number': ['children/fav_number/1',
                                    'children/fav_number/2',
                                    'children/fav_number/3']
        }
        row = {'children/fav_number': 1}

        expected_row = {
            'children/fav_number/1': True,
            'children/fav_number': 1,
            'children/fav_number/3': False,
            'children/fav_number/2': False
        }

        new_row = ExportBuilder.split_select_multiples(row, select_multiples)
        self.assertTrue(new_row)
        self.assertEqual(new_row, expected_row)

    def test_split_select_multiples_works_when_data_is_blank(self):
        select_multiples =\
            {
                'children/fav_colors': [
                    'children/fav_colors/red', 'children/fav_colors/blue',
                    'children/fav_colors/pink']
            }
        row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': ''
            }
        new_row = ExportBuilder.split_select_multiples(
            row, select_multiples)
        expected_row = \
            {
                'children/name': 'Mike',
                'children/age': 5,
                'children/fav_colors': '',
                'children/fav_colors/red': None,
                'children/fav_colors/blue': None,
                'children/fav_colors/pink': None
            }
        self.assertEqual(new_row, expected_row)

    def test_generation_of_gps_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_gps_fields =\
            {
                'childrens_survey':
                {
                    'geo/geolocation':
                    [
                        'geo/_geolocation_latitude',
                        'geo/_geolocation_longitude',
                        'geo/_geolocation_altitude',
                        'geo/_geolocation_precision'
                    ]
                }
            }
        gps_fields = export_builder.gps_fields
        self.assertTrue('childrens_survey' in gps_fields)
        self.assertEqual(
            sorted(gps_fields['childrens_survey']),
            sorted(expected_gps_fields['childrens_survey']))

    def test_split_gps_components_works(self):
        gps_fields =\
            {
                'geo/geolocation':
                [
                    'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                    'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                ]
            }
        row = \
            {
                'geo/geolocation': '1.0 36.1 2000 20',
            }
        new_row = ExportBuilder.split_gps_components(
            row, gps_fields)
        expected_row = \
            {
                'geo/geolocation': '1.0 36.1 2000 20',
                'geo/_geolocation_latitude': '1.0',
                'geo/_geolocation_longitude': '36.1',
                'geo/_geolocation_altitude': '2000',
                'geo/_geolocation_precision': '20'
            }
        self.assertEqual(new_row, expected_row)

    def test_split_gps_components_works_when_gps_data_is_blank(self):
        gps_fields =\
            {
                'geo/geolocation':
                [
                    'geo/_geolocation_latitude', 'geo/_geolocation_longitude',
                    'geo/_geolocation_altitude', 'geo/_geolocation_precision'
                ]
            }
        row = \
            {
                'geo/geolocation': '',
            }
        new_row = ExportBuilder.split_gps_components(
            row, gps_fields)
        expected_row = \
            {
                'geo/geolocation': '',
            }
        self.assertEqual(new_row, expected_row)

    def test_generation_of_mongo_encoded_fields_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        expected_encoded_fields =\
            {
                'childrens_survey':
                {
                    'tel/tel.office': 'tel/{0}'.format(
                        _encode_for_mongo('tel.office')),
                    'tel/tel.mobile': 'tel/{0}'.format(
                        _encode_for_mongo('tel.mobile')),
                }
            }
        encoded_fields = export_builder.encoded_fields
        self.assertTrue('childrens_survey' in encoded_fields)
        self.assertEqual(
            encoded_fields['childrens_survey'],
            expected_encoded_fields['childrens_survey'])

    def test_decode_fields_names_encoded_for_mongo(self):
        encoded_fields = \
            {
                'tel/tel.office': 'tel/{0}'.format(
                    _encode_for_mongo('tel.office'))
            }
        row = \
            {
                'name': 'Abe',
                'age': 35,
                'tel/{0}'.format(
                    _encode_for_mongo('tel.office')): '123-456-789'
            }
        new_row = ExportBuilder.decode_mongo_encoded_fields(
            row, encoded_fields)
        expected_row = \
            {
                'name': 'Abe',
                'age': 35,
                'tel/tel.office': '123-456-789'
            }
        self.assertEqual(new_row, expected_row)

    def test_generate_field_title(self):
        self._create_childrens_survey()
        field_name = ExportBuilder.format_field_title("children/age", ".",
                                                      data_dictionary=self.dd)
        expected_field_name = "children.age"
        self.assertEqual(field_name, expected_field_name)

    def test_delimiter_replacement_works_existing_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_sections =\
            [
                {
                    'name': 'children',
                    'elements': [
                        {
                            'title': 'children.name',
                            'xpath': 'children/name'
                        }
                    ]
                }
            ]
        children_section = export_builder.section_by_name('children')
        self.assertEqual(
            children_section['elements'][0]['title'],
            expected_sections[0]['elements'][0]['title'])

    def test_delimiter_replacement_works_generated_multi_select_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section =\
            {
                'name': 'children',
                'elements': [
                    {
                        'title': 'children.fav_colors.red',
                        'xpath': 'children/fav_colors/red'
                    }
                ]
            }
        childrens_section = export_builder.section_by_name('children')
        match = filter(lambda x: expected_section['elements'][0]['xpath'] ==
                       x['xpath'], childrens_section['elements'])[0]
        self.assertEqual(
            expected_section['elements'][0]['title'], match['title'])

    def test_delimiter_replacement_works_for_generated_gps_fields(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = "."
        export_builder.set_survey(survey)
        expected_section = \
            {
                'name': 'childrens_survey',
                'elements': [
                    {
                        'title': 'geo._geolocation_latitude',
                        'xpath': 'geo/_geolocation_latitude'
                    }
                ]
            }
        main_section = export_builder.section_by_name('childrens_survey')
        match = filter(
            lambda x: (expected_section['elements'][0]['xpath'] ==
                       x['xpath']), main_section['elements'])[0]
        self.assertEqual(
            expected_section['elements'][0]['title'], match['title'])

    def test_to_xls_export_works(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xls')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = xlrd.open_workbook(filename)
        # check that we have childrens_survey, children, children_cartoons
        # and children_cartoons_characters sheets
        expected_sheet_names = ['childrens_survey', 'children',
                                'children_cartoons',
                                'children_cartoons_characters']
        self.assertEqual(wb.sheet_names(), expected_sheet_names)

        # check header columns
        main_sheet = wb.sheet_by_name('childrens_survey')
        expected_column_headers = [
            u'name', u'age', u'geo/geolocation', u'geo/_geolocation_latitude',
            u'geo/_geolocation_longitude', u'geo/_geolocation_altitude',
            u'geo/_geolocation_precision', u'tel/tel.office',
            u'tel/tel.mobile', u'_id', u'meta/instanceID', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name', u'_tags', '_notes', '_version',
            '_duration', '_submitted_by']
        column_headers = main_sheet.row_values(0)
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        childrens_sheet = wb.sheet_by_name('children')
        expected_column_headers = [
            u'children/name', u'children/age', u'children/fav_colors',
            u'children/fav_colors/red', u'children/fav_colors/blue',
            u'children/fav_colors/pink', u'children/ice.creams',
            u'children/ice.creams/vanilla', u'children/ice.creams/strawberry',
            u'children/ice.creams/chocolate', u'_id', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name', u'_tags', '_notes', '_version',
            '_duration', '_submitted_by']
        column_headers = childrens_sheet.row_values(0)
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        cartoons_sheet = wb.sheet_by_name('children_cartoons')
        expected_column_headers = [
            u'children/cartoons/name', u'children/cartoons/why', u'_id',
            u'_uuid', u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name', u'_tags', '_notes', '_version',
            '_duration', '_submitted_by']
        column_headers = cartoons_sheet.row_values(0)
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        characters_sheet = wb.sheet_by_name('children_cartoons_characters')
        expected_column_headers = [
            u'children/cartoons/characters/name',
            u'children/cartoons/characters/good_or_evil', u'_id', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name', u'_tags', '_notes', '_version',
            '_duration', '_submitted_by']
        column_headers = characters_sheet.row_values(0)
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))

        xls_file.close()

    def test_to_xls_export_respects_custom_field_delimiter(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.GROUP_DELIMITER = ExportBuilder.GROUP_DELIMITER_DOT
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xls')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = xlrd.open_workbook(filename)

        # check header columns
        main_sheet = wb.sheet_by_name('childrens_survey')
        expected_column_headers = [
            u'name', u'age', u'geo.geolocation', u'geo._geolocation_latitude',
            u'geo._geolocation_longitude', u'geo._geolocation_altitude',
            u'geo._geolocation_precision', u'tel.tel.office',
            u'tel.tel.mobile', u'_id', u'meta.instanceID', u'_uuid',
            u'_submission_time', u'_index', u'_parent_index',
            u'_parent_table_name', u'_tags', '_notes', '_version',
            '_duration', '_submitted_by']
        column_headers = main_sheet.row_values(0)
        self.assertEqual(sorted(column_headers),
                         sorted(expected_column_headers))
        xls_file.close()

    def test_get_valid_sheet_name_catches_duplicates(self):
        work_sheets = {'childrens_survey': "Worksheet"}
        desired_sheet_name = "childrens_survey"
        expected_sheet_name = "childrens_survey1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheets)
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_get_valid_sheet_name_catches_long_names(self):
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_lo"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, [])
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_get_valid_sheet_name_catches_long_duplicate_names(self):
        work_sheet_titles = ['childrens_survey_with_a_very_lo']
        desired_sheet_name = "childrens_survey_with_a_very_long_name"
        expected_sheet_name = "childrens_survey_with_a_very_l1"
        generated_sheet_name = ExportBuilder.get_valid_sheet_name(
            desired_sheet_name, work_sheet_titles)
        self.assertEqual(generated_sheet_name, expected_sheet_name)

    def test_to_xls_export_generates_valid_sheet_names(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_with_a_very_long_name.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xls')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.data)
        xls_file.seek(0)
        wb = xlrd.open_workbook(filename)
        # check that we have childrens_survey, children, children_cartoons
        # and children_cartoons_characters sheets
        expected_sheet_names = ['childrens_survey_with_a_very_lo',
                                'childrens_survey_with_a_very_l1',
                                'childrens_survey_with_a_very_l2',
                                'childrens_survey_with_a_very_l3']
        self.assertEqual(wb.sheet_names(), expected_sheet_names)
        xls_file.close()

    def test_child_record_parent_table_is_updated_when_sheet_is_renamed(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_with_a_very_long_name.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        xls_file = NamedTemporaryFile(suffix='.xlsx')
        filename = xls_file.name
        export_builder.to_xls_export(filename, self.long_survey_data)
        xls_file.seek(0)
        wb = load_workbook(filename)

        # get the children's sheet
        ws1 = wb.get_sheet_by_name('childrens_survey_with_a_very_l1')

        # parent_table is in cell K2
        parent_table_name = ws1['K2'].value
        expected_parent_table_name = 'childrens_survey_with_a_very_lo'
        self.assertEqual(parent_table_name, expected_parent_table_name)

        # get cartoons sheet
        ws2 = wb.get_sheet_by_name('childrens_survey_with_a_very_l2')
        parent_table_name = ws2['G2'].value
        expected_parent_table_name = 'childrens_survey_with_a_very_l1'
        self.assertEqual(parent_table_name, expected_parent_table_name)
        xls_file.close()

    def test_type_conversion(self):
        submission_1 = {
            "_id": 579827,
            "geolocation": "-1.2625482 36.7924794 0.0 21.0",
            "_bamboo_dataset_id": "",
            "meta/instanceID": "uuid:2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "name": "Smith",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "_submission_time": "2013-07-03T08:25:30",
            "age": "107",
            "_uuid": "2a8129f5-3091-44e1-a579-bed2b07a12cf",
            "when": "2013-07-03",
            "amount": "250.0",
            "_geolocation": [
                "-1.2625482",
                "36.7924794"
            ],
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "precisely": "2013-07-03T15:24:00.000+03",
            "really": "15:24:00.000+03"
        }

        submission_2 = {
            "_id": 579828,
            "_submission_time": "2013-07-03T08:26:10",
            "_uuid": "5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "_bamboo_dataset_id": "",
            "_xform_id_string": "test_data_types",
            "_userform_id": "larryweya_test_data_types",
            "_status": "submitted_via_web",
            "meta/instanceID": "uuid:5b4752eb-e13c-483e-87cb-e67ca6bb61e5",
            "formhub/uuid": "633ec390e024411ba5ce634db7807e62",
            "amount": "",
        }

        survey = create_survey_from_xls(viewer_fixture_path(
            'test_data_types/test_data_types.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        # format submission 1 for export
        survey_name = survey.name
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission_1, 1, indices, survey_name,
                                     survey, submission_1)
        new_row = export_builder.pre_process_row(data[survey_name],
                                                 export_builder.sections[0])
        self.assertIsInstance(new_row['age'], int)
        self.assertIsInstance(new_row['when'], datetime.date)
        self.assertIsInstance(new_row['amount'], float)

        # check missing values dont break and empty values return blank strings
        indices = {survey_name: 0}
        data = dict_to_joined_export(submission_2, 1, indices, survey_name,
                                     survey, submission_2)
        new_row = export_builder.pre_process_row(data[survey_name],
                                                 export_builder.sections[0])
        self.assertIsInstance(new_row['amount'], basestring)
        self.assertEqual(new_row['amount'], '')

    def test_xls_convert_dates_before_1900(self):
        survey = create_survey_from_xls(viewer_fixture_path(
            'test_data_types/test_data_types.xls'))
        export_builder = ExportBuilder()
        export_builder.set_survey(survey)
        data = [
            {
                'name': 'Abe',
                'when': '1899-07-03',
            }
        ]
        # create export file
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, data)
        temp_xls_file.close()
        # this should error if there is a problem, not sure what to assert

    def test_convert_types(self):
        val = '1'
        expected_val = 1
        converted_val = ExportBuilder.convert_type(val, 'int')
        self.assertIsInstance(converted_val, int)
        self.assertEqual(converted_val, expected_val)

        val = '1.2'
        expected_val = 1.2
        converted_val = ExportBuilder.convert_type(val, 'decimal')
        self.assertIsInstance(converted_val, float)
        self.assertEqual(converted_val, expected_val)

        val = '2012-06-23'
        expected_val = datetime.date(2012, 6, 23)
        converted_val = ExportBuilder.convert_type(val, 'date')
        self.assertIsInstance(converted_val, datetime.date)
        self.assertEqual(converted_val, expected_val)

    def test_to_sav_export(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        filename = temp_zip_file.name
        export_builder.to_zipped_sav(filename, self.data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(
                    d, index, indices, survey_name, survey, d))
            index += 1

        # check that each file exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "{0}.sav".format(survey.name))))

        def _test_sav_file(section):
            with SavReader(
                    os.path.join(
                        temp_dir, "{0}.sav".format(section)),
                    returnHeader=True) as reader:
                header = next(reader)
                rows = [r for r in reader]

                # open comparison file
                with SavReader(_logger_fixture_path(
                        'spss', "{0}.sav".format(section)),
                        returnHeader=True) as fixture_reader:
                    fixture_header = next(fixture_reader)
                    self.assertEqual(header, fixture_header)
                    expected_rows = [r for r in fixture_reader]
                    self.assertEqual(rows, expected_rows)

                if section == 'children_cartoons_charactors':
                    self.assertEqual(reader.valueLabels, {
                        'good_or_evil': {'good': 'Good'}
                    })

        for section in export_builder.sections:
            section_name = section['name'].replace('/', '_')
            _test_sav_file(section_name)

    def test_to_sav_export_language(self):
        survey = self._create_childrens_survey('childrens_survey_sw.xls')
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        filename = temp_zip_file.name
        export_builder.to_zipped_sav(filename, self.data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(
                    d, index, indices, survey_name, survey, d))
            index += 1

        # check that each file exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "{0}.sav".format(survey.name))))

        def _test_sav_file(section):
            with SavReader(
                    os.path.join(
                        temp_dir, "{0}.sav".format(section)),
                    returnHeader=True) as reader:
                header = next(reader)
                rows = [r for r in reader]
                if section != 'childrens_survey_sw':
                    section += '_sw'

                # open comparison file
                with SavReader(_logger_fixture_path(
                        'spss', "{0}.sav".format(section)),
                        returnHeader=True) as fixture_reader:
                    fixture_header = next(fixture_reader)
                    self.assertEqual(header, fixture_header)
                    expected_rows = [r for r in fixture_reader]
                    self.assertEqual(rows, expected_rows)

                if section == 'children_cartoons_charactors':
                    self.assertEqual(reader.valueLabels, {
                        'good_or_evil': {'good': 'Good'}
                    })

        for section in export_builder.sections:
            section_name = section['name'].replace('/', '_')
            _test_sav_file(section_name)

    def test_generate_field_title_truncated_titles(self):
        self._create_childrens_survey()
        field_name = ExportBuilder.format_field_title("children/age", "/",
                                                      data_dictionary=self.dd,
                                                      remove_group_name=True)
        expected_field_name = "age"
        self.assertEqual(field_name, expected_field_name)

    def test_generate_field_title_truncated_titles_select_multiple(self):
        self._create_childrens_survey()
        field_name = ExportBuilder.format_field_title(
            "children/fav_colors/red", "/",
            data_dictionary=self.dd,
            remove_group_name=True
        )
        expected_field_name = "fav_colors/red"
        self.assertEqual(field_name, expected_field_name)

    def test_xls_export_remove_group_name(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data_utf8)
        temp_xls_file.seek(0)
        # check that values for red\'s and blue\'s are set to true
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb.get_sheet_by_name("children.info")
        data = dict([(r[0].value, r[1].value) for r in children_sheet.columns])
        self.assertTrue(data[u"fav_colors/red's"])
        self.assertTrue(data[u"fav_colors/blue's"])
        self.assertFalse(data[u"fav_colors/pink's"])
        temp_xls_file.close()

    def test_zipped_csv_export_remove_group_name(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(os.path.join(temp_dir, "children.info.csv")) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = ['name.first',
                                'age',
                                'fav_colors',
                                u'fav_colors/red\'s',
                                u'fav_colors/blue\'s',
                                u'fav_colors/pink\'s',
                                'ice_creams',
                                'ice_creams/vanilla',
                                'ice_creams/strawberry',
                                'ice_creams/chocolate', '_id',
                                '_uuid', '_submission_time', '_index',
                                '_parent_table_name', '_parent_index',
                                u'_tags', '_notes', '_version',
                                '_duration', '_submitted_by']
            rows = [row for row in reader]
            actual_headers = [h.decode('utf-8') for h in rows[0]]
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(
                data[u'fav_colors/red\'s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'fav_colors/blue\'s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'fav_colors/pink\'s'.encode('utf-8')],
                'False')
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    def test_xls_export_with_labels(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data_utf8)
        temp_xls_file.seek(0)
        # check that values for red\'s and blue\'s are set to true
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb.get_sheet_by_name("children.info")
        labels = dict([(r[0].value, r[1].value)
                       for r in children_sheet.columns])
        self.assertEqual(labels[u'name.first'], '3.1 Childs name')
        self.assertEqual(labels[u'age'], '3.2 Child age')
        self.assertEqual(labels[u'fav_colors/red\'s'], 'fav_colors/Red')
        self.assertEqual(labels[u'fav_colors/blue\'s'], 'fav_colors/Blue')
        self.assertEqual(labels[u'fav_colors/pink\'s'], 'fav_colors/Pink')

        data = dict([(r[0].value, r[2].value) for r in children_sheet.columns])
        self.assertEqual(data[u'name.first'], 'Mike')
        self.assertEqual(data[u'age'], 5)
        self.assertTrue(data[u'fav_colors/red\'s'])
        self.assertTrue(data[u'fav_colors/blue\'s'])
        self.assertFalse(data[u'fav_colors/pink\'s'])
        temp_xls_file.close()

    def test_xls_export_with_labels_only(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS_ONLY = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data_utf8)
        temp_xls_file.seek(0)
        # check that values for red\'s and blue\'s are set to true
        wb = load_workbook(temp_xls_file.name)
        children_sheet = wb.get_sheet_by_name("children.info")
        data = dict([(r[0].value, r[1].value) for r in children_sheet.columns])
        self.assertEqual(data['3.1 Childs name'], 'Mike')
        self.assertEqual(data['3.2 Child age'], 5)
        self.assertTrue(data[u'fav_colors/Red'])
        self.assertTrue(data[u'fav_colors/Blue'])
        self.assertFalse(data[u'fav_colors/Pink'])
        temp_xls_file.close()

    def test_zipped_csv_export_with_labels(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(os.path.join(temp_dir, "children.info.csv")) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = ['name.first',
                                'age',
                                'fav_colors',
                                u'fav_colors/red\'s',
                                u'fav_colors/blue\'s',
                                u'fav_colors/pink\'s',
                                'ice_creams',
                                'ice_creams/vanilla',
                                'ice_creams/strawberry',
                                'ice_creams/chocolate', '_id',
                                '_uuid', '_submission_time', '_index',
                                '_parent_table_name', '_parent_index',
                                u'_tags', '_notes', '_version',
                                '_duration', '_submitted_by']
            expected_labels = ['3.1 Childs name',
                               '3.2 Child age',
                               '3.3 Favorite Colors',
                               'fav_colors/Red',
                               'fav_colors/Blue',
                               'fav_colors/Pink',
                               '3.3 Ice Creams',
                               'ice_creams/Vanilla',
                               'ice_creams/Strawberry',
                               'ice_creams/Chocolate', '_id',
                               '_uuid', '_submission_time', '_index',
                               '_parent_table_name', '_parent_index',
                               u'_tags', '_notes', '_version',
                               '_duration', '_submitted_by']
            rows = [row for row in reader]
            actual_headers = [h.decode('utf-8') for h in rows[0]]
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            actual_labels = [h.decode('utf-8') for h in rows[1]]
            self.assertEqual(sorted(actual_labels), sorted(expected_labels))
            data = dict(zip(rows[0], rows[2]))
            self.assertEqual(
                data[u'fav_colors/red\'s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'fav_colors/blue\'s'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'fav_colors/pink\'s'.encode('utf-8')],
                'False')
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    def test_zipped_csv_export_with_labels_only(self):
        """
        cvs writer doesnt handle unicode we we have to encode to ascii
        """
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_unicode.xls'))
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS_ONLY = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        export_builder.to_zipped_csv(temp_zip_file.name, self.data_utf8)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()
        # check that the children's file (which has the unicode header) exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "children.info.csv")))
        # check file's contents
        with open(os.path.join(temp_dir, "children.info.csv")) as csv_file:
            reader = csv.reader(csv_file)
            expected_headers = [
                '3.1 Childs name',
                '3.2 Child age',
                '3.3 Favorite Colors',
                'fav_colors/Red',
                'fav_colors/Blue',
                'fav_colors/Pink',
                '3.3 Ice Creams',
                'ice_creams/Vanilla',
                'ice_creams/Strawberry',
                'ice_creams/Chocolate', '_id',
                '_uuid', '_submission_time', '_index',
                '_parent_table_name', '_parent_index',
                u'_tags', '_notes', '_version',
                '_duration', '_submitted_by'
            ]
            rows = [row for row in reader]
            actual_headers = [h.decode('utf-8') for h in rows[0]]
            self.assertEqual(sorted(actual_headers), sorted(expected_headers))
            data = dict(zip(rows[0], rows[1]))
            self.assertEqual(
                data[u'fav_colors/Red'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'fav_colors/Blue'.encode('utf-8')],
                'True')
            self.assertEqual(
                data[u'fav_colors/Pink'.encode('utf-8')],
                'False')
            # check that red and blue are set to true
        shutil.rmtree(temp_dir)

    def test_to_sav_export_with_labels(self):
        survey = self._create_childrens_survey()
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.set_survey(survey)
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        temp_zip_file = NamedTemporaryFile(suffix='.zip')
        filename = temp_zip_file.name
        export_builder.to_zipped_sav(filename, self.data)
        temp_zip_file.seek(0)
        temp_dir = tempfile.mkdtemp()
        zip_file = zipfile.ZipFile(temp_zip_file.name, "r")
        zip_file.extractall(temp_dir)
        zip_file.close()
        temp_zip_file.close()

        # generate data to compare with
        index = 1
        indices = {}
        survey_name = survey.name
        outputs = []
        for d in self.data:
            outputs.append(
                dict_to_joined_export(
                    d, index, indices, survey_name, survey, d))
            index += 1

        # check that each file exists
        self.assertTrue(
            os.path.exists(
                os.path.join(temp_dir, "{0}.sav".format(survey.name))))

        def _test_sav_file(section):
            sav_path = os.path.join(temp_dir, "{0}.sav".format(section))
            if section == 'children_survey':
                with SavHeaderReader(sav_path) as header:
                    expected_labels = [
                        '1. What is your name?', '2. How old are you?',
                        '4. Geo-location', '5.1 Office telephone',
                        '5.2 Mobile telephone', '_duration', '_id',
                        '_index', '_notes', '_parent_index',
                        '_parent_table_name', '_submission_time',
                        '_submitted_by',
                        '_tags', '_uuid', '_version',
                        'geo/_geolocation_altitude',
                        'geo/_geolocation_latitude',
                        'geo/_geolocation_longitude',
                        'geo/_geolocation_precision',
                        'meta/instanceID'
                    ]
                    labels = header.varLabels.values()
                    self.assertEqual(sorted(expected_labels), sorted(labels))

            with SavReader(sav_path, returnHeader=True) as reader:
                header = next(reader)
                rows = [r for r in reader]

                # open comparison file
                with SavReader(_logger_fixture_path(
                        'spss', "{0}.sav".format(section)),
                        returnHeader=True) as fixture_reader:
                    fixture_header = next(fixture_reader)
                    self.assertEqual(header, fixture_header)
                    expected_rows = [r for r in fixture_reader]
                    self.assertEqual(rows, expected_rows)

        for section in export_builder.sections:
            section_name = section['name'].replace('/', '_')
            _test_sav_file(section_name)

    def test_xls_export_with_english_labels(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_en.xls'))
        # no default_language is not set
        self.assertEqual(
            survey.to_json_dict().get('default_language'), 'default'
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data)
        temp_xls_file.seek(0)
        wb = load_workbook(temp_xls_file.name)
        childrens_survey_sheet = wb.get_sheet_by_name("childrens_survey_en")
        labels = dict([(r[0].value, r[1].value)
                       for r in childrens_survey_sheet.columns])
        self.assertEqual(labels[u'name'], '1. What is your name?')
        self.assertEqual(labels[u'age'], '2. How old are you?')

        children_sheet = wb.get_sheet_by_name("children")
        labels = dict([(r[0].value, r[1].value)
                       for r in children_sheet.columns])
        self.assertEqual(labels['fav_colors/red'], 'fav_colors/Red')
        self.assertEqual(labels['fav_colors/blue'], 'fav_colors/Blue')
        temp_xls_file.close()

    def test_xls_export_with_swahili_labels(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_sw.xls'))
        # default_language is set to swahili
        self.assertEqual(
            survey.to_json_dict().get('default_language'), 'swahili'
        )
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        temp_xls_file = NamedTemporaryFile(suffix='.xlsx')
        export_builder.to_xls_export(temp_xls_file.name, self.data)
        temp_xls_file.seek(0)
        wb = load_workbook(temp_xls_file.name)
        childrens_survey_sheet = wb.get_sheet_by_name("childrens_survey_sw")
        labels = dict([(r[0].value, r[1].value)
                       for r in childrens_survey_sheet.columns])
        self.assertEqual(labels[u'name'], '1. Jina lako ni?')
        self.assertEqual(labels[u'age'], '2. Umri wako ni?')

        children_sheet = wb.get_sheet_by_name("children")
        labels = dict([(r[0].value, r[1].value)
                       for r in children_sheet.columns])
        self.assertEqual(labels['fav_colors/red'], 'fav_colors/Nyekundu')
        self.assertEqual(labels['fav_colors/blue'], 'fav_colors/Bluu')
        temp_xls_file.close()

    def test_csv_export_with_swahili_labels(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_sw.xls'))
        # default_language is set to swahili
        self.assertEqual(
            survey.to_json_dict().get('default_language'), 'swahili'
        )
        dd = DataDictionary()
        dd._survey = survey
        ordered_columns = OrderedDict()
        CSVDataFrameBuilder._build_ordered_columns(survey, ordered_columns)
        ordered_columns['children/fav_colors/red'] = None
        labels = get_labels_from_columns(ordered_columns, dd, '/')
        self.assertIn('1. Jina lako ni?', labels)
        self.assertIn('2. Umri wako ni?', labels)
        self.assertIn('fav_colors/Nyekundu', labels)

    def test_select_multiples_choices(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'childrens_survey_sw.xls'))
        dd = DataDictionary()
        dd._survey = survey
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        child = [e for e in dd.get_survey_elements_with_choices()
                 if e.bind.get('type') == 'select'][0]
        self.assertNotEqual(child.children, [])
        choices = export_builder._get_select_mulitples_choices(
            child, dd, ExportBuilder.GROUP_DELIMITER,
            ExportBuilder.TRUNCATE_GROUP_TITLE
        )
        expected_choices = [
            {
                'xpath': u'children/fav_colors/red',
                'title': u'children/fav_colors/red',
                'type': 'string',
                'label': u'fav_colors/Nyekundu'
            }, {
                'xpath': u'children/fav_colors/blue',
                'title': u'children/fav_colors/blue',
                'type': 'string', 'label': u'fav_colors/Bluu'
            }, {
                'xpath': u'children/fav_colors/pink',
                'title': u'children/fav_colors/pink',
                'type': 'string', 'label': u'fav_colors/Pink'
            }
        ]
        self.assertEqual(choices, expected_choices)
        select_multiples = {
            u'children/fav_colors': [
                u'children/fav_colors/red', u'children/fav_colors/blue',
                u'children/fav_colors/pink'
            ], u'children/ice.creams': [
                u'children/ice.creams/vanilla',
                u'children/ice.creams/strawberry',
                u'children/ice.creams/chocolate'
            ]
        }
        self.assertEqual(CSVDataFrameBuilder._collect_select_multiples(dd),
                         select_multiples)

    def test_select_multiples_choices_with_choice_filter(self):
        survey = create_survey_from_xls(_logger_fixture_path(
            'choice_filter.xls'
        ))
        dd = DataDictionary()
        dd._survey = survey
        export_builder = ExportBuilder()
        export_builder.TRUNCATE_GROUP_TITLE = True
        export_builder.INCLUDE_LABELS = True
        export_builder.set_survey(survey)
        child = [e for e in dd.get_survey_elements_with_choices()
                 if e.bind.get('type') == 'select'][0]
        choices = export_builder._get_select_mulitples_choices(
            child, dd, ExportBuilder.GROUP_DELIMITER,
            ExportBuilder.TRUNCATE_GROUP_TITLE
        )
        self.assertEqual(child.children, [])
        expected_choices = [
            {
                'label': u'county/King',
                'title': u'county/king',
                'type': 'string',
                'xpath': u'county/king'
            },  {
                'label': u'county/Pierce',
                'title': u'county/pierce',
                'type': 'string',
                'xpath': u'county/pierce'
            },  {
                'label': u'county/King',
                'title': u'county/king',
                'type': 'string',
                'xpath': u'county/king'
            },  {
                'label': u'county/Cameron',
                'title': u'county/cameron',
                'type': 'string',
                'xpath': u'county/cameron'
            }
        ]
        self.assertEqual(choices, expected_choices)
        select_multiples = {
            u'county': [
                u'county/king',
                u'county/pierce',
                u'county/king',
                u'county/cameron'
            ]
        }
        self.assertEqual(CSVDataFrameBuilder._collect_select_multiples(dd),
                         select_multiples)

    def test_string_to_date_with_xls_validation(self):
        # test "2016-11-02"
        val = ExportBuilder.string_to_date_with_xls_validation("2016-11-02")
        self.assertEqual(val, datetime.date(2016, 11, 2))

        # test random string
        val = ExportBuilder.string_to_date_with_xls_validation("random")
        self.assertEqual(val, "random")

        val = ExportBuilder.string_to_date_with_xls_validation(0.4)
        self.assertEqual(val, 0.4)
