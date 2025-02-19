# -*- coding: utf-8 -*-
"""
DataDictionary model.
"""

import importlib
import json
import os
from io import BytesIO

from django.contrib.contenttypes.models import ContentType
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction
from django.db.models.signals import post_save, pre_save
from django.utils import timezone
from django.utils.translation import gettext as _

import openpyxl
import unicodecsv as csv
from floip import FloipSurvey
from kombu.exceptions import OperationalError
from pyxform.utils import has_external_choices
from pyxform.xls2json_backends import xlsx_value_to_str

from onadata.apps.logger.models.entity_list import EntityList
from onadata.apps.logger.models.follow_up_form import FollowUpForm
from onadata.apps.logger.models.registration_form import RegistrationForm
from onadata.apps.logger.models.xform import (
    XForm,
    check_version_set,
    check_xform_uuid,
    get_survey_from_file_object,
)
from onadata.apps.logger.xform_instance_parser import XLSFormError
from onadata.apps.main.models.meta_data import MetaData
from onadata.libs.utils.cache_tools import (
    PROJ_BASE_FORMS_CACHE,
    PROJ_FORMS_CACHE,
    safe_delete,
)
from onadata.libs.utils.model_tools import get_columns_with_hxl, set_uuid


def is_newline_error(error):
    """
    Return True is e is a new line error based on the error text.
    Otherwise return False.
    """
    newline_error = (
        "new-line character seen in unquoted field - do you need"
        " to open the file in universal-newline mode?"
    )
    return newline_error == str(error)


def process_xlsform_survey(xls, default_name):
    """
    Process XLSForm file and return the PyXForm Survey for the XLSForm.
    """
    # FLOW Results package is a JSON file.
    if xls.name.endswith("json"):
        return FloipSurvey(xls).survey.to_json_dict()
    return get_survey_from_file_object(xls, name=default_name)


def process_xlsform(xls, default_name):
    """
    Process XLSForm file and return the survey dictionary for the XLSForm.
    """
    return process_xlsform_survey(xls, default_name).to_json_dict()


# adopted from pyxform.utils.sheet_to_csv
# pylint: disable=too-many-branches,too-many-locals
def sheet_to_csv(xls_content, sheet_name):
    """Writes a csv file of a specified sheet from a an excel file

    :param xls_content: Excel file contents
    :param sheet_name: the name of the excel sheet to generate the csv file

    :returns: a (StringIO) csv file object
    """
    workbook = openpyxl.load_workbook(xls_content)
    sheet = workbook[sheet_name]

    if not sheet or sheet.max_column < 2:
        raise ValueError(_(f"Sheet <'{sheet_name}'> has no data."))

    csv_file = BytesIO()

    writer = csv.writer(csv_file, encoding="utf-8", quoting=csv.QUOTE_ALL)
    mask = [v and len(v.strip()) > 0 for v in list(sheet.values)[0]]

    for row in sheet.iter_rows(values_only=True):
        row_values = []
        try:
            for val in row:
                if val is not None:
                    val = xlsx_value_to_str(val)
                    val = val.strip()
                row_values.append(val)
        except TypeError:
            continue

        if not all(v is None for v in row_values):
            writer.writerow([v for v, m in zip(row_values, mask) if m])
    return csv_file


def upload_to(instance, filename, username=None):
    """
    Return XLSForm file upload path.
    """
    if instance:
        username = instance.xform.user.username
    return os.path.join(username, "xls", os.path.split(filename)[1])


class DataDictionary(XForm):  # pylint: disable=too-many-instance-attributes
    """
    DataDictionary model class.
    """

    def __init__(self, *args, **kwargs):
        self.instances_for_export = lambda d: d.instances.all()
        self.has_external_choices = False
        self._id_string_changed = False
        super().__init__(*args, **kwargs)

    def __str__(self):
        return getattr(self, "id_string", "")

    def save(self, *args, **kwargs):
        skip_xls_read = kwargs.get("skip_xls_read")

        if self.xls and not skip_xls_read:
            default_name = None if not self.pk else self.survey.xml_instance().tagName
            survey = process_xlsform_survey(self.xls, default_name)
            survey_dict = survey.to_json_dict()
            if has_external_choices(survey_dict):
                self.has_external_choices = True
            survey = check_version_set(survey)
            if get_columns_with_hxl(survey.get("children")):
                self.has_hxl_support = True
            # if form is being replaced, don't check for id_string uniqueness
            if self.pk is None:
                new_id_string = self.get_unique_id_string(survey.get("id_string"))
                self._id_string_changed = new_id_string != survey.get("id_string")
                survey["id_string"] = new_id_string
                # For flow results packages use the user defined id/uuid
                if self.xls.name.endswith("json"):
                    self.uuid = FloipSurvey(self.xls).descriptor.get("id")
                    if self.uuid:
                        check_xform_uuid(self.uuid)
            elif self.id_string != survey.get("id_string"):
                raise XLSFormError(
                    _(
                        (
                            "Your updated form's id_string '%(new_id)s' must match "
                            "the existing forms' id_string '%(old_id)s'."
                            % {
                                "new_id": survey.get("id_string"),
                                "old_id": self.id_string,
                            }
                        )
                    )
                )
            elif default_name and default_name != survey.get("name"):
                survey["name"] = default_name
            else:
                survey["id_string"] = self.id_string
            self.json = survey.to_json_dict()
            self.xml = survey.to_xml()
            self.version = survey.get("version")
            self.last_updated_at = timezone.now()
            self.title = survey.get("title")
            self.mark_start_time_boolean()
            set_uuid(self)
            self.set_uuid_in_xml()
            self.set_hash()

        if "skip_xls_read" in kwargs:
            del kwargs["skip_xls_read"]

        super().save(*args, **kwargs)

    def file_name(self):
        return (
            os.path.split(self.xls.name)[-1]
            if self.xls.name is not None
            else self.id_string + ".xml"
        )


# pylint: disable=unused-argument
def set_object_permissions(sender, instance=None, created=False, **kwargs):
    """
    Apply the relevant object permissions for the form to all users who should
    have access to it.
    """
    # seems the super is not called, have to get xform from here
    xform = XForm.objects.get(pk=instance.pk)

    if created:
        # pylint: disable=import-outside-toplevel
        from onadata.libs.permissions import OwnerRole

        OwnerRole.add(instance.user, xform)

        if instance.created_by and instance.user != instance.created_by:
            OwnerRole.add(instance.created_by, xform)

        # pylint: disable=import-outside-toplevel
        from onadata.libs.utils.project_utils import (
            set_project_perms_to_xform_async,
        )  # noqa

        try:
            transaction.on_commit(
                lambda: set_project_perms_to_xform_async.delay(
                    xform.pk, instance.project.pk
                )
            )
        except OperationalError:
            # pylint: disable=import-outside-toplevel
            from onadata.libs.utils.project_utils import (
                set_project_perms_to_xform,
            )  # noqa

            set_project_perms_to_xform(xform, instance.project)

    if hasattr(instance, "has_external_choices") and instance.has_external_choices:
        instance.xls.seek(0)
        choices_file = sheet_to_csv(instance.xls, "external_choices")
        choices_file.seek(0, os.SEEK_END)
        size = choices_file.tell()
        choices_file.seek(0)

        data_file = InMemoryUploadedFile(
            file=choices_file,
            field_name="data_file",
            name="itemsets.csv",
            content_type="text/csv",
            size=size,
            charset=None,
        )

        MetaData.media_upload(xform, data_file)


post_save.connect(
    set_object_permissions,
    sender=DataDictionary,
    dispatch_uid="xform_object_permissions",
)


# pylint: disable=unused-argument
def save_project(sender, instance=None, created=False, **kwargs):
    """
    Receive XForm project to update date_modified field of the project and on
    the next XHR request the form will be included in the project data.
    """
    instance.project.save()


pre_save.connect(
    save_project, sender=DataDictionary, dispatch_uid="save_project_datadictionary"
)


def create_registration_form(sender, instance=None, created=False, **kwargs):
    """Create a RegistrationForm for a form that defines entities

    Create an EntityList if it does not exist. If it exists, use the
    the existing EntityList
    """
    instance_json = instance.json

    if isinstance(instance_json, str):
        instance_json = json.loads(instance_json)

    if not instance_json.get("entity_features"):
        return

    children = instance_json.get("children", [])
    meta_list = filter(lambda child: child.get("name") == "meta", children)

    for meta in meta_list:
        for child in meta.get("children", []):
            if child.get("name") == "entity":
                parameters = child.get("parameters", {})
                dataset = parameters.get("dataset")
                entity_list, _ = EntityList.objects.get_or_create(
                    name=dataset, project=instance.project
                )
                (
                    registration_form,
                    registration_form_created,
                ) = RegistrationForm.objects.get_or_create(
                    entity_list=entity_list,
                    xform=instance,
                )

                if registration_form_created:
                    # RegistrationForm contributing to any previous
                    # EntityList should be disabled
                    for form in instance.registration_forms.exclude(
                        entity_list=entity_list, is_active=True
                    ):
                        form.is_active = False
                        form.save()
                elif not registration_form_created and not registration_form.is_active:
                    # If previously disabled, enable it
                    registration_form.is_active = True
                    registration_form.save()

                return


post_save.connect(
    create_registration_form,
    sender=DataDictionary,
    dispatch_uid="create_registration_form_datadictionary",
)


def create_follow_up_form(sender, instance=None, created=False, **kwargs):
    """Create a FollowUpForm for a form that consumes entities

    Check if a form consumes data from a dataset that is an EntityList. If so,
    we create a FollowUpForm
    """
    instance_json = instance.json

    if isinstance(instance_json, str):
        instance_json = json.loads(instance_json)

    children = instance_json.get("children", [])
    active_entity_datasets: list[str] = []
    xform = XForm.objects.get(pk=instance.pk)

    for child in children:
        if child["type"] == "select one" and "itemset" in child:
            dataset_name = child["itemset"].split(".")[0]

            try:
                entity_list = EntityList.objects.get(
                    name=dataset_name, project=instance.project
                )

            except EntityList.DoesNotExist:
                # No EntityList dataset was found with the specified
                # name, we simply do nothing
                continue

            active_entity_datasets.append(entity_list.name)
            follow_up_form, created = FollowUpForm.objects.get_or_create(
                entity_list=entity_list, xform=instance
            )

            if not created and not follow_up_form.is_active:
                # If previously deactivated, re-activate
                follow_up_form.is_active = True
                follow_up_form.save()

            content_type = ContentType.objects.get_for_model(xform)
            MetaData.objects.get_or_create(
                object_id=xform.pk,
                content_type=content_type,
                data_type="media",
                data_value=f"entity_list {entity_list.pk} {entity_list.name}",
            )

    # Deactivate the XForm's FollowUpForms whose EntityList are not
    # referenced by the updated XForm version
    inactive_follow_up_forms = FollowUpForm.objects.filter(xform=xform).exclude(
        entity_list__name__in=active_entity_datasets
    )
    inactive_follow_up_forms.update(is_active=False)


post_save.connect(
    create_follow_up_form,
    sender=DataDictionary,
    dispatch_uid="create_follow_up_datadictionary",
)


def disable_registration_form(sender, instance=None, created=False, **kwargs):
    """Disable registration form if form no longer contains entities definitions"""
    instance_json = instance.json

    if isinstance(instance_json, str):
        instance_json = json.loads(instance_json)

    if not instance_json.get("entity_features"):
        # If form creates entities, disable the registration forms
        for registration_form in instance.registration_forms.filter(is_active=True):
            registration_form.is_active = False
            registration_form.save()


post_save.connect(
    disable_registration_form,
    sender=DataDictionary,
    dispatch_uid="disable_registration_form_datadictionary",
)


def invalidate_caches(sender, instance=None, created=False, **kwargs):
    """Invalidate caches"""
    # Avoid cyclic dependency errors
    api_tools = importlib.import_module("onadata.apps.api.tools")

    xform = XForm.objects.get(pk=instance.pk)

    safe_delete(f"{PROJ_FORMS_CACHE}{instance.project.pk}")
    safe_delete(f"{PROJ_BASE_FORMS_CACHE}{instance.project.pk}")
    api_tools.invalidate_xform_list_cache(xform)


post_save.connect(
    invalidate_caches,
    sender=DataDictionary,
    dispatch_uid="xform_invalidate_caches",
)


def create_or_update_export_register(sender, instance=None, created=False, **kwargs):
    """Create or update export columns register for the form"""
    # Avoid cyclic import by using importlib
    logger_tasks = importlib.import_module("onadata.apps.logger.tasks")

    MetaData.update_or_create_export_register(instance)

    if not created:
        transaction.on_commit(
            lambda: logger_tasks.reconstruct_xform_export_register_async.delay(
                instance.pk
            )
        )


post_save.connect(
    create_or_update_export_register,
    sender=DataDictionary,
    dispatch_uid="create_or_update_export_register",
)
