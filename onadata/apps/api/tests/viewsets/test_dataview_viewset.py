# -*- coding: utf-8 -*-
"""Test DataViewViewSet"""

import csv
import json
import os
from datetime import datetime, timedelta, timezone
from unittest.mock import patch

from django.conf import settings
from django.core.cache import cache
from django.core.files.storage import default_storage
from django.test.utils import override_settings

from flaky import flaky
from openpyxl import load_workbook

from onadata.apps.api.tests.viewsets.test_abstract_viewset import TestAbstractViewSet
from onadata.apps.api.viewsets.attachment_viewset import AttachmentViewSet
from onadata.apps.api.viewsets.dataview_viewset import (
    DataViewViewSet,
    apply_filters,
    filter_to_field_lookup,
    get_field_lookup,
    get_filter_kwargs,
)
from onadata.apps.api.viewsets.note_viewset import NoteViewSet
from onadata.apps.api.viewsets.project_viewset import ProjectViewSet
from onadata.apps.api.viewsets.xform_viewset import XFormViewSet
from onadata.apps.logger.models import Attachment, Instance
from onadata.apps.logger.models.data_view import DataView
from onadata.apps.viewer.models.export import Export
from onadata.libs.permissions import ReadOnlyRole
from onadata.libs.serializers.attachment_serializer import AttachmentSerializer
from onadata.libs.serializers.xform_serializer import XFormSerializer
from onadata.libs.utils.cache_tools import (
    DATAVIEW_COUNT,
    DATAVIEW_LAST_SUBMISSION_TIME,
    PROJECT_LINKED_DATAVIEWS,
)
from onadata.libs.utils.common_tags import EDITED, MONGO_STRFTIME
from onadata.libs.utils.common_tools import (
    filename_from_disposition,
    get_response_content,
)


class TestDataViewViewSet(TestAbstractViewSet):
    def setUp(self):
        super().setUp()
        xlsform_path = os.path.join(
            settings.PROJECT_ROOT, "libs", "tests", "utils", "fixtures", "tutorial.xlsx"
        )

        self._publish_xls_form_to_project(xlsform_path=xlsform_path)
        for fixture in range(1, 9):
            path = os.path.join(
                settings.PROJECT_ROOT,
                "libs",
                "tests",
                "utils",
                "fixtures",
                "tutorial",
                "instances",
                f"uuid{fixture}",
                "submission.xml",
            )
            self._make_submission(path)

        self.view = DataViewViewSet.as_view(
            {
                "post": "create",
                "put": "update",
                "patch": "partial_update",
                "delete": "destroy",
                "get": "retrieve",
            }
        )

    def test_create_dataview(self):
        self._create_dataview()

    def test_filter_to_field_lookup(self):
        self.assertEqual(filter_to_field_lookup("="), "__iexact")
        self.assertEqual(filter_to_field_lookup("<"), "__lt")
        self.assertEqual(filter_to_field_lookup(">"), "__gt")

    def test_get_field_lookup(self):
        self.assertEqual(get_field_lookup("q1", "="), "json__q1__iexact")
        self.assertEqual(get_field_lookup("q1", "<"), "json__q1__lt")
        self.assertEqual(get_field_lookup("q1", ">"), "json__q1__gt")

    def test_get_filter_kwargs(self):
        self.assertEqual(
            get_filter_kwargs([{"value": 2, "column": "first_column", "filter": "<"}]),
            {"json__first_column__lt": "2"},
        )
        self.assertEqual(
            get_filter_kwargs([{"value": 2, "column": "first_column", "filter": ">"}]),
            {"json__first_column__gt": "2"},
        )
        self.assertEqual(
            get_filter_kwargs([{"value": 2, "column": "first_column", "filter": "="}]),
            {"json__first_column__iexact": "2"},
        )

    def test_apply_filters(self):
        # update these filters
        filters = [{"value": "orange", "column": "fruit", "filter": "="}]
        xml = '<data id="a"><fruit>orange</fruit></data>'
        instance = Instance(xform=self.xform, xml=xml)
        instance.save()
        self.assertEqual(apply_filters(self.xform.instances, filters).first().xml, xml)
        # delete instance
        instance.delete()

    # pylint: disable=invalid-name
    def test_dataview_with_attachment_field(self):
        view = DataViewViewSet.as_view({"get": "data"})
        media_file = "test-image.png"
        attachment_file_path = os.path.join(
            settings.PROJECT_ROOT, "libs", "tests", "utils", "fixtures", media_file
        )
        submission_file_path = os.path.join(
            settings.PROJECT_ROOT,
            "libs",
            "tests",
            "utils",
            "fixtures",
            "tutorial",
            "instances",
            "uuid10",
            "submission.xml",
        )

        # make a submission with an attachment
        with open(attachment_file_path, "rb") as f:
            self._make_submission(submission_file_path, media_file=f)

        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            # ensure there's an attachment column(photo) in you dataview
            "columns": '["name", "age", "gender", "photo"]',
            "query": '[{"column":"pizza_fan","filter":"=","value":"no"}]',
        }

        self._create_dataview(data=data)
        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        for item in response.data:
            # retrieve the instance with attachment
            if item.get("photo") == media_file:
                instance_with_attachment = item

        self.assertTrue(instance_with_attachment)
        attachment_info = instance_with_attachment.get("_attachments")[0]
        self.assertEqual("image/png", attachment_info.get("mimetype"))
        self.assertEqual(
            f"{self.user.username}/attachments/{self.xform.id}_{self.xform.id_string}/{media_file}",
            attachment_info.get("filename"),
        )
        self.assertEqual(response.status_code, 200)

        # Attachment viewset works ok for filtered datasets
        attachment_list_view = AttachmentViewSet.as_view({"get": "list"})
        request = self.factory.get("/?dataview=" + str(self.data_view.pk), **self.extra)
        response = attachment_list_view(request)
        attachments = Attachment.objects.filter(instance__xform=self.data_view.xform)
        self.assertEqual(1, len(response.data))
        self.assertEqual(
            self.data_view.query,
            [{"value": "no", "column": "pizza_fan", "filter": "="}],
        )
        serialized_attachments = AttachmentSerializer(
            attachments, many=True, context={"request": request}
        ).data
        self.assertEqual(serialized_attachments, response.data)

        # create profile for alice
        alice_data = {
            "username": "alice",
            "email": "alice@localhost.com",
            "password1": "alice",
            "password2": "alice",
            "first_name": "Alice",
            "last_name": "A",
            "city": "Nairobi",
            "country": "KE",
        }
        alice_profile = self._create_user_profile(extra_post_data=alice_data)
        self.extra = {"HTTP_AUTHORIZATION": f"Token {alice_profile.user.auth_token}"}

        # check that user with no permisisons can not list attachment objects
        request = self.factory.get("/?dataview=" + str(self.data_view.pk), **self.extra)
        response = attachment_list_view(request)
        attachments = Attachment.objects.filter(instance__xform=self.data_view.xform)
        self.assertEqual(0, len(response.data))
        self.assertEqual(
            self.data_view.query,
            [{"value": "no", "column": "pizza_fan", "filter": "="}],
        )
        self.assertEqual([], response.data)

        # check that user with no permisisons can not view a specific attachment object
        attachment_list_view = AttachmentViewSet.as_view({"get": "retrieve"})
        request = self.factory.get("/?dataview=" + str(self.data_view.pk), **self.extra)
        response = attachment_list_view(request, pk=attachments.first().pk)
        self.assertEqual(
            self.data_view.query,
            [{"value": "no", "column": "pizza_fan", "filter": "="}],
        )
        self.assertEqual(response.status_code, 404)
        response_data = json.loads(json.dumps(response.data))
        self.assertEqual(
            response_data, {"detail": "No Attachment matches the given query."}
        )

        # a user with permissions can view a specific attachment object
        attachment_list_view = AttachmentViewSet.as_view({"get": "retrieve"})
        self.extra = {"HTTP_AUTHORIZATION": f"Token {self.user.auth_token}"}
        request = self.factory.get("/?dataview=" + str(self.data_view.pk), **self.extra)
        response = attachment_list_view(request, pk=attachments.first().pk)
        self.assertEqual(
            self.data_view.query,
            [{"value": "no", "column": "pizza_fan", "filter": "="}],
        )
        self.assertEqual(response.status_code, 200)
        serialized_attachment = AttachmentSerializer(
            attachments.first(), context={"request": request}
        ).data
        self.assertEqual(response.data, serialized_attachment)

    # pylint: disable=invalid-name
    def test_get_dataview_form_definition(self):
        self._create_dataview()

        data = {
            "name": "data",
            "title": "tutorial",
            "default_language": "default",
            "id_string": "tutorial",
            "type": "survey",
        }
        self.view = DataViewViewSet.as_view(
            {
                "get": "form",
            }
        )
        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        # JSON format
        response = self.view(request, pk=self.data_view.pk, format="json")
        self.assertEqual(response.status_code, 200)
        self.assertDictContainsSubset(data, response.data)

    def test_get_dataview_form_details(self):
        self._create_dataview()

        self.view = DataViewViewSet.as_view(
            {
                "get": "form_details",
            }
        )
        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        response = self.view(request, pk=self.data_view.pk, format="json")
        self.assertEqual(response.status_code, 200)

        self.assertIn("title", response.data)
        self.assertIn("created_by", response.data)
        self.assertIn("id_string", response.data)
        self.assertIn("metadata", response.data)

    def test_get_dataview(self):
        self._create_dataview()

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["dataviewid"], self.data_view.pk)
        self.assertEqual(response.data["name"], "My DataView")
        self.assertEqual(
            response.data["xform"], f"http://testserver/api/v1/forms/{self.xform.pk}"
        )
        self.assertEqual(
            response.data["project"],
            f"http://testserver/api/v1/projects/{self.project.pk}",
        )
        self.assertEqual(response.data["columns"], ["name", "age", "gender"])
        self.assertEqual(
            response.data["query"],
            [
                {"column": "age", "filter": ">", "value": "20"},
                {"column": "age", "filter": "<", "value": "50"},
            ],
        )
        self.assertEqual(
            response.data["url"],
            f"http://testserver/api/v1/dataviews/{self.data_view.pk}",
        )
        self.assertEqual(
            response.data["last_submission_time"], "2015-03-09T13:34:05.537766+00:00"
        )

        # Public
        self.project.shared = True
        self.project.save()

        anon_request = self.factory.get("/")
        anon_response = self.view(anon_request, pk=self.data_view.pk)
        self.assertEqual(anon_response.status_code, 200)

        # Private
        self.project.shared = False
        self.project.save()

        anon_request = self.factory.get("/")
        anon_response = self.view(anon_request, pk=self.data_view.pk)
        self.assertEqual(anon_response.status_code, 404)

    def test_update_dataview(self):
        self._create_dataview()

        data = {
            "name": "My DataView updated",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"age","filter":">","value":"20"}]',
        }

        request = self.factory.put("/", data=data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["name"], "My DataView updated")

        self.assertEqual(response.data["columns"], ["name", "age", "gender"])

        self.assertEqual(
            response.data["query"], [{"column": "age", "filter": ">", "value": "20"}]
        )

    def test_patch_dataview(self):
        self._create_dataview()

        data = {
            "name": "My DataView updated",
        }

        request = self.factory.patch("/", data=data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["name"], "My DataView updated")

    def test_soft_delete_dataview(self):
        """
        Tests that a dataview is soft deleted
        """
        self._create_dataview()
        dataview_id = self.data_view.pk
        self.assertIsNone(self.data_view.deleted_at)
        self.assertNotIn("-deleted-at-", self.data_view.name)

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=dataview_id)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["name"], "My DataView")

        # delete
        request = self.factory.delete("/", **self.extra)
        response = self.view(request, pk=dataview_id)
        self.assertEqual(response.status_code, 204)

        # check that it is soft deleted
        data_view = DataView.objects.get(pk=dataview_id)
        self.assertIsNotNone(data_view.deleted_at)
        self.assertIn("-deleted-at-", data_view.name)
        self.assertEqual(data_view.deleted_by.username, "bob")

    # pylint: disable=invalid-name
    def test_soft_deleted_dataview_not_in_forms_list(self):
        self._create_dataview()
        get_form_request = self.factory.get("/", **self.extra)

        xform_serializer = XFormSerializer(
            self.xform, context={"request": get_form_request}
        )

        self.assertIsNotNone(xform_serializer.data["data_views"])

        request = self.factory.delete("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 204)

        xform_serializer = XFormSerializer(
            self.xform, context={"request": get_form_request}
        )
        self.assertEqual(xform_serializer.data["data_views"], [])

    # pylint: disable=invalid-name
    def test_soft_deleted_dataview_not_in_project(self):
        """
        Test that once a filtered dataset is soft deleted
        it does not appear in the list of forms for a project
        """
        self._create_dataview()
        view = ProjectViewSet.as_view({"get": "retrieve"})
        # assert that dataview is in the returned list
        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.project.pk)
        self.assertIsNotNone(response.data["data_views"])
        # delete dataview
        request = self.factory.delete("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 204)
        # assert that deleted dataview is not in the returned list
        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.project.pk)
        self.assertEqual(response.data["data_views"], [])

    def test_list_dataview(self):
        self._create_dataview()

        data = {
            "name": "My DataView2",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"age","filter":">","value":"20"}]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "list",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)

        # delete DataView and check that we don't get it in response
        dataview = DataView.objects.get(name="My DataView2")
        deleted_dataview_id = dataview.id
        dataview.soft_delete(user=self.user)
        response = view(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertNotEqual(response.data[0]["dataviewid"], deleted_dataview_id)

        anon_request = request = self.factory.get("/")
        anon_response = view(anon_request)
        self.assertEqual(anon_response.status_code, 401)

    def test_get_dataview_no_perms(self):
        self._create_dataview()

        alice_data = {"username": "alice", "email": "alice@localhost.com"}
        self._login_user_and_profile(alice_data)

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 404)

        # assign alice the perms
        ReadOnlyRole.add(self.user, self.data_view.project)

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)

    def test_can_not_get_deleted_dataview(self):
        data = {
            "name": "Agriculture Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"age","filter":">","value":"20"},'
            '{"column":"age","filter":"<","value":"50"}]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "retrieve",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["dataviewid"], self.data_view.pk)

        dataview = DataView.objects.get(id=response.data["dataviewid"])
        dataview.soft_delete(user=self.user)

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 404)

    # pylint: disable=invalid-name
    def test_dataview_data_filter_integer(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"age","filter":">","value":"20"},'
            '{"column":"age","filter":"<","value":"50"}]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 3)
        self.assertIn("_id", response.data[0])

    def test_dataview_data_filter_decimal(self):
        """
        Test that data filter works correctly for decimal fields
        """
        # publish form with decimal field and make submissions
        path = os.path.join(
            settings.PROJECT_ROOT,
            "libs",
            "tests",
            "utils",
            "fixtures",
            "age_decimal",
            "age_decimal.xlsx",
        )
        self._publish_xls_form_to_project(xlsform_path=path)
        for fixture in range(1, 3):
            path = os.path.join(
                settings.PROJECT_ROOT,
                "libs",
                "tests",
                "utils",
                "fixtures",
                "age_decimal",
                "instances",
                f"submission{fixture}.xml",
            )
            self._make_submission(path)

        # create a dataview using filter age > 30
        data = {
            "name": "My Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age"]',
            "query": '[{"column":"age","filter":">","value":"30"}]',
        }
        self._create_dataview(data=data)
        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )
        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertEqual(response.data[0]["age"], 31)

    def test_dataview_data_filter_date(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "gender", "_submission_time"]',
            "query": '[{"column":"_submission_time",'
            '"filter":">=","value":"2015-01-01T00:00:00"}]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 7)
        self.assertIn("_id", response.data[0])

    # pylint: disable=invalid-name
    def test_dataview_data_filter_string(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "gender", "_submission_time"]',
            "query": '[{"column":"gender","filter":"<>","value":"male"}]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)

    # pylint: disable=invalid-name
    def test_dataview_data_filter_condition(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "gender", "age"]',
            "query": '[{"column":"name","filter":"=","value":"Fred",'
            ' "condition":"or"},'
            '{"column":"name","filter":"=","value":"Kameli",'
            ' "condition":"or"},'
            '{"column":"gender","filter":"=","value":"male"}]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn("_id", response.data[0])

    def test_dataview_invalid_filter(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "gender", "age"]',
            "query": '[{"column":"name","filter":"<=>","value":"Fred",'
            ' "condition":"or"}]',
        }

        request = self.factory.post("/", data=data, **self.extra)
        response = self.view(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data, {"query": ["Filter not supported"]})

    def test_dataview_sql_injection(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "gender", "age"]',
            "query": '[{"column":"age","filter":"=",'
            '"value":"1;UNION ALL SELECT NULL,version()'
            ',NULL LIMIT 1 OFFSET 1--;"}]',
        }

        request = self.factory.post("/", data=data, **self.extra)
        response = self.view(request)
        self.assertEqual(response.status_code, 400)
        self.assertIn("detail", response.data)

        self.assertTrue(
            str(response.data.get("detail")).startswith(
                "invalid input syntax for type integer"
            ),
            response.data,
        )

    def test_dataview_invalid_columns(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": "age",
        }

        request = self.factory.post("/", data=data, **self.extra)
        response = self.view(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn(
            response.data["columns"][0],
            [
                "Expecting value: line 1 column 1 (char 0)",
                "No JSON object could be decoded",
            ],
        )

    def test_dataview_invalid_query(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["age"]',
            "query": "age=10",
        }

        request = self.factory.post("/", data=data, **self.extra)
        response = self.view(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn(
            response.data["query"][0],
            [
                "Expecting value: line 1 column 1 (char 0)",
                "No JSON object could be decoded",
            ],
        )

    # pylint: disable=invalid-name
    def test_dataview_query_not_required(self):
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["age"]',
        }

        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 8)

        self.assertIn("_id", response.data[0])
        self.assertIn(EDITED, response.data[0])

    def test_csv_export_dataview(self):
        self._create_dataview()
        count = Export.objects.all().count()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk, format="csv")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(count + 1, Export.objects.all().count())

        headers = dict(response.items())
        self.assertEqual(headers["Content-Type"], "application/csv")
        content_disposition = headers["Content-Disposition"]
        filename = filename_from_disposition(content_disposition)
        _basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.csv"')

        content = get_response_content(response)
        test_file_path = os.path.join(
            settings.PROJECT_ROOT, "apps", "viewer", "tests", "fixtures", "dataview.csv"
        )
        with open(test_file_path, encoding="utf-8") as test_file:
            self.assertEqual(content, test_file.read())

    def test_csvzip_export_dataview(self):
        self._create_dataview()
        count = Export.objects.all().count()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk, format="csvzip")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(count + 1, Export.objects.all().count())

        request = self.factory.get("/", **self.extra)
        response = view(request, pk="[invalid pk]", format="csvzip")
        self.assertEqual(response.status_code, 404)

    def test_zip_export_dataview(self):
        media_file = "test-image.png"
        attachment_file_path = os.path.join(
            settings.PROJECT_ROOT, "libs", "tests", "utils", "fixtures", media_file
        )
        submission_file_path = os.path.join(
            settings.PROJECT_ROOT,
            "libs",
            "tests",
            "utils",
            "fixtures",
            "tutorial",
            "instances",
            "uuid10",
            "submission.xml",
        )

        # make a submission with an attachment
        with open(attachment_file_path, "rb") as f:
            self._make_submission(submission_file_path, media_file=f)

        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "photo"]',
            "query": '[{"column":"age","filter":"=","value":"90"}]',
        }
        self._create_dataview(data)
        count = Export.objects.all().count()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(1, len(response.data))

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk, format="zip")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(count + 1, Export.objects.all().count())

        headers = dict(response.items())
        self.assertEqual(headers["Content-Type"], "application/zip")
        content_disposition = headers["Content-Disposition"]
        filename = filename_from_disposition(content_disposition)
        _basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.zip"')

    # pylint: disable=invalid-name
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    def test_export_csv_dataview_data_async(self, async_result):
        self._create_dataview()
        self._publish_xls_form_to_project()

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        request = self.factory.get("/", data={"format": "csv"}, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertIsNotNone(response.data)

        self.assertEqual(response.status_code, 202)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)

    # pylint: disable=invalid-name
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    def test_export_csv_dataview_with_labels_async(self, async_result):
        self._create_dataview()
        self._publish_xls_form_to_project()

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        request = self.factory.get(
            "/", data={"format": "csv", "include_labels": "true"}, **self.extra
        )
        response = view(request, pk=self.data_view.pk)
        self.assertIsNotNone(response.data)

        self.assertEqual(response.status_code, 202)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)
        with default_storage.open(export.filepath, "r") as f:
            csv_reader = csv.reader(f)
            next(csv_reader)
            labels = next(csv_reader)
            self.assertIn("Gender", labels)

    # pylint: disable=invalid-name
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    def test_export_xls_dataview_with_labels_async(self, async_result):
        self._create_dataview()
        self._publish_xls_form_to_project()

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        request = self.factory.get(
            "/",
            data={"format": "xlsx", "force_xlsx": "true", "include_labels": "true"},
            **self.extra,
        )
        response = view(request, pk=self.data_view.pk)
        self.assertIsNotNone(response.data)

        self.assertEqual(response.status_code, 202)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)
        workbook = load_workbook(export.full_filepath)
        sheet_name = workbook.get_sheet_names()[0]
        main_sheet = workbook.get_sheet_by_name(sheet_name)
        self.assertIn("Gender", tuple(main_sheet.values)[1])
        self.assertEqual(len(tuple(main_sheet.values)), 5)

    def _test_csv_export_with_hxl_support(self, name, columns, expected_output):  # noqa
        data = {
            "name": name,
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": columns,
            "query": "[]",
        }

        self._create_dataview(data=data)

        dataview_pk = DataView.objects.last().pk

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get(
            "/", data={"format": "csv", "include_hxl": True}, **self.extra
        )
        response = view(request, pk=dataview_pk)

        self.assertIsNotNone(next(response.streaming_content), expected_output)

    # pylint: disable=invalid-name
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    def test_xlsx_export_with_choice_labels(self, async_result):
        """
        Test that choice labels are present in xlsx export when enabled
        """
        xform = self.xform
        project = self.project
        # add pizza_type column which is has choice labels
        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{xform.pk}",
            "project": f"http://testserver/api/v1/projects/{project.pk}",
            "columns": (
                '["name", "age", "gender", "pizza_type", "_id", "_uuid", '
                '"_submission_time", "_index", "_parent_table_name",  "_parent_index", '
                '"_tags", "_notes", "_version", "_duration","_submitted_by"]'
            ),
            "query": ('[{"column":"age","filter":"=","value":"28"}]'),
        }
        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        data = {"format": "xlsx", "show_choice_labels": "true"}

        request = self.factory.get("/", data=data, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertIsNotNone(response.data)

        self.assertEqual(response.status_code, 202)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)
        workbook = load_workbook(export.full_filepath)
        workbook.iso_dates = True
        sheet_name = workbook.sheetnames[0]
        main_sheet = workbook[sheet_name]
        sheet_headers = list(main_sheet.values)[0]
        sheet_data = list(main_sheet.values)[1]
        inst = self.xform.instances.get(id=sheet_data[4])
        expected_headers = (
            "name",
            "age",
            "gender",
            "pizza_type",
            "_id",
            "_uuid",
            "_submission_time",
            "_index",
            "_parent_table_name",
            "_parent_index",
            "_tags",
            "_notes",
            "_version",
            "_duration",
            "_submitted_by",
        )
        expected_data = (
            "Dennis Wambua",
            28,
            "Male",
            "New York think crust!",
            inst.id,
            inst.uuid,
            inst.date_created.replace(microsecond=0, tzinfo=None),
            1,
            None,
            -1,
            None,
            None,
            "4444",
            50,
            inst.user.username,
        )
        self.assertEqual(expected_headers, sheet_headers)
        self.assertEqual(expected_data, sheet_data)

    # pylint: disable=invalid-name
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    def test_csv_export_with_choice_labels(self, async_result):
        """
        Test that choice labels are present in csv export when enabled
        """
        xform = self.xform
        project = self.project
        # add pizza_type column which is has choice labels
        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{xform.pk}",
            "project": f"http://testserver/api/v1/projects/{project.pk}",
            "columns": '["name", "age", "gender", "pizza_type"]',
            "query": ('[{"column":"age","filter":"=","value":"28"}]'),
        }
        self._create_dataview(data=data)

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        data = {"format": "csv", "show_choice_labels": "true"}

        request = self.factory.get("/", data=data, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertIsNotNone(response.data)

        self.assertEqual(response.status_code, 202)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)
        with default_storage.open(export.filepath, "r") as f:
            expected_data = ["Dennis Wambua", "28", "Male", "New York think crust!"]
            expected_headers = ["name", "age", "gender", "pizza_type"]
            csv_reader = csv.reader(f)
            headers = next(csv_reader)
            self.assertEqual(expected_headers, headers)
            data = next(csv_reader)
            self.assertEqual(expected_data, data)

    # pylint: disable=invalid-name
    def test_csv_export_with_hxl_support(self):
        self._publish_form_with_hxl_support()
        self._test_csv_export_with_hxl_support(
            "test name 1", '["name"]', "name\nCristiano Ronaldo 1\nLionel Messi\n"
        )
        self._test_csv_export_with_hxl_support(
            "test name 2", '["age"]', "age\n#age,\n31\n29\n"
        )
        self._test_csv_export_with_hxl_support(
            "test name 3",
            '["age", "name"]',
            "age,name\n#age,\n31,Cristiano Ronaldo\n29,Lionel Messi\n",
        )

    def test_get_charts_data(self):
        self._create_dataview()
        self.view = DataViewViewSet.as_view(
            {
                "get": "charts",
            }
        )
        data_view_data = DataView.query_data(self.data_view)

        request = self.factory.get("/charts", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        data = {"field_name": "age"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "integer")
        self.assertEqual(response.data["field_name"], "age")
        self.assertEqual(response.data["data_type"], "numeric")
        self.assertEqual(len(response.data["data"]), len(data_view_data))

        data = {"field_xpath": "age"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "integer")
        self.assertEqual(response.data["field_name"], "age")
        self.assertEqual(response.data["data_type"], "numeric")
        self.assertEqual(len(response.data["data"]), len(data_view_data))

    # pylint: disable=invalid-name
    def test_get_charts_data_for_submission_time_field(self):
        self._create_dataview()
        self.view = DataViewViewSet.as_view(
            {
                "get": "charts",
            }
        )

        data = {"field_name": "_submission_time"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "datetime")
        self.assertEqual(response.data["field_name"], "_submission_time")
        self.assertEqual(response.data["data_type"], "time_based")

        data = {"field_name": "_submitted_by"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "text")
        self.assertEqual(response.data["field_name"], "_submitted_by")
        self.assertEqual(response.data["data_type"], "categorized")

        data = {"field_name": "_duration"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "integer")
        self.assertEqual(response.data["field_name"], "_duration")
        self.assertEqual(response.data["data_type"], "numeric")

    def test_get_charts_data_for_grouped_field(self):
        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender", "a_group/grouped"]',
            "query": '[{"column":"age","filter":">","value":"20"}]',
        }
        self._create_dataview(data)
        self.view = DataViewViewSet.as_view(
            {
                "get": "charts",
            }
        )

        request = self.factory.get("/charts", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        data = {"field_name": "grouped"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "select one")
        self.assertEqual(response.data["field_name"], "grouped")
        self.assertEqual(response.data["data_type"], "categorized")
        self.assertEqual(len(response.data["data"]), 2)

    # pylint: disable=invalid-name
    def test_get_charts_data_field_not_in_dataview_columns(self):
        self._create_dataview()
        self.view = DataViewViewSet.as_view(
            {
                "get": "charts",
            }
        )

        data = {"field_name": "grouped"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 404)

    def test_get_charts_data_with_empty_query(self):
        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": "[]",
        }
        self._create_dataview(data)
        self.view = DataViewViewSet.as_view(
            {
                "get": "charts",
            }
        )
        data_view_data = DataView.query_data(self.data_view)

        request = self.factory.get("/charts", **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        data = {"field_name": "age"}
        request = self.factory.get("/charts", data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotEqual(response.get("Cache-Control"), None)
        self.assertEqual(response.data["field_type"], "integer")
        self.assertEqual(response.data["field_name"], "age")
        self.assertEqual(response.data["data_type"], "numeric")
        self.assertEqual(len(response.data["data"]), len(data_view_data))

    def test_geopoint_dataview(self):
        # Dataview with geolocation column selected.
        # -> instances_with_geopoints= True
        data = {
            "name": "My DataView1",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender", "location"]',
            "query": '[{"column":"age","filter":">","value":"20"}]',
        }
        self._create_dataview(data)

        self.assertTrue(self.data_view.instances_with_geopoints)

        # Dataview with geolocation column NOT selected
        # -> instances_with_geopoints= False
        data = {
            "name": "My DataView2",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"age","filter":">","value":"20"}]',
        }
        self._create_dataview(data)

        self.assertFalse(self.data_view.instances_with_geopoints)

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["dataviewid"], self.data_view.pk)
        self.assertEqual(response.data["name"], "My DataView2")
        self.assertEqual(response.data["instances_with_geopoints"], False)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertNotIn("location", response.data[0])
        self.assertNotIn("_geolocation", response.data[0])

    # pylint: disable=invalid-name
    def test_geopoint_submission_dataview(self):
        data = {
            "name": "My DataView3",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender", "location"]',
            "query": '[{"column":"age","filter":">=","value":"87"}]',
        }
        self._create_dataview(data)

        self.assertTrue(self.data_view.instances_with_geopoints)

        # make submission with geopoint
        path = os.path.join(
            settings.PROJECT_ROOT,
            "libs",
            "tests",
            "utils",
            "fixtures",
            "tutorial",
            "instances",
            "uuid9",
            "submission.xml",
        )
        self._make_submission(path)

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["dataviewid"], self.data_view.pk)
        self.assertEqual(response.data["name"], "My DataView3")
        self.assertEqual(response.data["instances_with_geopoints"], True)

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertIn("location", response.data[0])
        self.assertIn("_geolocation", response.data[0])

        # geojson pagination, fields and geofield params works ok
        request = self.factory.get(
            "/?geofield=_geolocation&page=1&page_size=1&fields=name", **self.extra
        )
        response = view(request, pk=self.data_view.pk, format="geojson")
        # we get correct content type
        headers = dict(response.items())
        self.assertEqual(headers["Content-Type"], "application/geo+json")
        self.assertEqual(response.status_code, 200)
        del response.data["features"][0]["properties"]["xform"]
        del response.data["features"][0]["properties"]["id"]
        self.assertEqual(
            {
                "type": "FeatureCollection",
                "features": [
                    {
                        "type": "Feature",
                        "geometry": None,
                        "properties": {"name": "Kameli"},
                    }
                ],
            },
            response.data,
        )
        request = self.factory.get(
            "/?geofield=_geolocation&page=9&page_size=1&fields=name", **self.extra
        )
        response = view(request, pk=self.data_view.pk, format="geojson")
        self.assertEqual(response.status_code, 200)
        del response.data["features"][0]["properties"]["xform"]
        del response.data["features"][0]["properties"]["id"]
        self.assertEqual(
            {
                "type": "FeatureCollection",
                "features": [
                    {
                        "type": "Feature",
                        "geometry": {
                            "type": "GeometryCollection",
                            "geometries": [
                                {"type": "Point", "coordinates": [36.8304, -1.2655]}
                            ],
                        },
                        "properties": {"name": "Kameli"},
                    }
                ],
            },
            response.data,
        )
        request = self.factory.get(
            "/?geofield=_geolocation&page=10&page_size=1&fields=name", **self.extra
        )
        response = view(request, pk=self.data_view.pk, format="geojson")
        self.assertEqual(response.status_code, 404)
        self.assertEqual({"detail": "Invalid page."}, response.data)

    # pylint: disable=invalid-name
    def test_dataview_project_cache_cleared(self):
        self._create_dataview()

        view = ProjectViewSet.as_view(
            {
                "get": "retrieve",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.project.pk)

        self.assertEqual(response.status_code, 200)

        cached_dataviews = cache.get(f"{PROJECT_LINKED_DATAVIEWS}{self.project.pk}")

        self.assertIsNotNone(cached_dataviews)

        # update the dataview
        self.data_view.name = "updated name"
        self.data_view.save()

        updated_cache = cache.get(f"{PROJECT_LINKED_DATAVIEWS}{self.project.pk}")

        self.assertIsNone(updated_cache)

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.project.pk)

        self.assertEqual(response.status_code, 200)

        cached_dataviews = cache.get(f"{PROJECT_LINKED_DATAVIEWS}{self.project.pk}")

        self.assertIsNotNone(cached_dataviews)

        self.data_view.delete()

        updated_cache = cache.get(f"{PROJECT_LINKED_DATAVIEWS}{self.project.pk}")
        self.assertIsNone(updated_cache)

    # pylint: disable=invalid-name
    def test_dataview_update_refreshes_cached_data(self):
        self._create_dataview()
        cache.set(f"{DATAVIEW_COUNT}{self.data_view.xform.pk}", 5)
        cache.set(
            f"{DATAVIEW_LAST_SUBMISSION_TIME}{self.data_view.xform.pk}",
            "2015-03-09T13:34:05",
        )
        self.data_view.name = "Updated Dataview"
        self.data_view.save()

        self.assertIsNone(cache.get(f"{DATAVIEW_COUNT}{self.data_view.xform.pk}"))
        self.assertIsNone(
            cache.get(f"{DATAVIEW_LAST_SUBMISSION_TIME}{self.data_view.xform.pk}")
        )

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        expected_count = 3
        expected_last_submission_time = "2015-03-09T13:34:05.537766+00:00"

        self.assertEqual(response.data["count"], expected_count)
        self.assertEqual(
            response.data["last_submission_time"], "2015-03-09T13:34:05.537766+00:00"
        )

        cache_dict = cache.get(f"{DATAVIEW_COUNT}{self.data_view.xform.pk}")
        self.assertEqual(cache_dict.get(self.data_view.pk), expected_count)
        self.assertEqual(
            cache.get(f"{DATAVIEW_LAST_SUBMISSION_TIME}{self.data_view.xform.pk}"),
            expected_last_submission_time,
        )

    # pylint: disable=invalid-name
    def test_export_dataview_not_affected_by_normal_exports(self):
        count = Export.objects.all().count()

        view = XFormViewSet.as_view(
            {
                "get": "retrieve",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.xform.pk, format="csv")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(count + 1, Export.objects.all().count())

        self._create_dataview()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk, format="csv")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(count + 2, Export.objects.all().count())

        headers = dict(response.items())
        self.assertEqual(headers["Content-Type"], "application/csv")
        content_disposition = headers["Content-Disposition"]
        filename = filename_from_disposition(content_disposition)
        _basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.csv"')

        content = get_response_content(response)

        # count csv headers and ensure they are three
        self.assertEqual(len(content.split("\n")[0].split(",")), 3)

    def test_matches_parent(self):
        self._create_dataview()
        self.assertFalse(self.data_view.matches_parent)
        columns = [
            "name",
            "age",
            "gender",
            "photo",
            "date",
            "location",
            "pizza_fan",
            "pizza_hater",
            "pizza_type",
            "favorite_toppings",
            "test_location2.latitude",
            "test_location2.longitude",
            "test_location.precision",
            "test_location2.precision",
            "test_location.altitude",
            "test_location.latitude",
            "test_location2.altitude",
            "test_location.longitude",
            "thanks",
            "a_group",
            "a_group/grouped",
            "a_group/a_text",
            "start_time",
            "end_time",
            "today",
            "imei",
            "phonenumber",
            "meta",
            "meta/instanceID",
        ]
        data = {
            "name": "My DataView2",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": json.dumps(columns),
            "query": '[{"column":"age","filter":">","value":"20"}]',
        }

        self._create_dataview(data)
        self.assertTrue(self.data_view.matches_parent)

    # pylint: disable=invalid-name
    def test_dataview_create_data_filter_invalid_date(self):
        invalid_query = (
            '[{"column":"_submission_time","filter":">","value":"30/06/2015"}]'
        )
        data = {
            "name": "Transportation Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "gender", "_submission_time"]',
            "query": invalid_query,
        }

        view = DataViewViewSet.as_view(
            {"get": "data", "post": "create", "patch": "partial_update"}
        )

        request = self.factory.post("/", data=data, **self.extra)
        response = view(request)

        # Confirm you cannot create an invalid dataview
        self.assertEqual(response.status_code, 400)

    # pylint: disable=invalid-name
    def test_dataview_update_data_filter_invalid_date(self):
        invalid_query = (
            '[{"column":"_submission_time","filter":">","value":"30/06/2015"}]'
        )
        self._create_dataview()

        data = {"query": invalid_query}
        request = self.factory.patch("/", data=data, **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        # Confirm you cannot update an invalid dataview
        self.assertEqual(response.status_code, 400)

    # pylint: disable=invalid-name
    def test_dataview_serializer_exception(self):
        invalid_query = [
            {"column": "_submission_time", "filter": ">", "value": "30/06/2015"}
        ]
        self._create_dataview()

        self.data_view.query = invalid_query
        self.data_view.save()

        request = self.factory.get("/", **self.extra)
        response = self.view(request, pk=self.data_view.pk)

        self.assertEqual(response.status_code, 400)

    # pylint: disable=invalid-name
    def test_dataview_notes_added_to_data(self):
        # Create note
        view = NoteViewSet.as_view({"post": "create"})
        comment = "Dataview note"
        note = {"note": comment}
        data_id = self.xform.instances.all().order_by("pk")[0].pk
        note["instance"] = data_id
        request = self.factory.post("/", data=note, **self.extra)
        self.assertTrue(self.xform.instances.count())
        response = view(request)
        self.assertEqual(response.status_code, 201)

        # Get dataview with added notes
        data = {
            "name": "My Dataview",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["age"]',
        }
        self._create_dataview(data=data)
        view = DataViewViewSet.as_view({"get": "data"})
        request = self.factory.get("/", **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 8)
        data_with_notes = next((d for d in response.data if d["_id"] == data_id))
        self.assertIn("_notes", data_with_notes)
        self.assertEqual(
            [
                {
                    "created_by": self.user.id,
                    "id": 1,
                    "instance_field": None,
                    "note": comment,
                    "owner": self.user.username,
                }
            ],
            data_with_notes["_notes"],
        )

    def test_sort_dataview_data(self):
        self._create_dataview()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        data = {"sort": '{"age": -1}'}
        request = self.factory.get("/", data=data, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.is_sorted_desc([r.get("age") for r in response.data]))

    def test_invalid_date_filter(self):
        view = DataViewViewSet.as_view(
            {
                "get": "retrieve",
                "post": "create",
            }
        )
        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"_submission_time","filter":">",'
            '"value":"26-01-2016"}]',
        }

        request = self.factory.post("/", data=data, **self.extra)

        response = view(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data,
            {
                "non_field_errors": [
                    "Date value in _submission_time should be"
                    " yyyy-mm-ddThh:m:s or yyyy-mm-dd"
                ]
            },
        )

        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"_submission_time","filter":">",'
            '"value":"26/01/2016"}]',
        }

        request = self.factory.post("/", data=data, **self.extra)

        response = view(request)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.data,
            {
                "non_field_errors": [
                    "Date value in _submission_time should be"
                    " yyyy-mm-ddThh:m:s or yyyy-mm-dd"
                ]
            },
        )

        data = {
            "name": "My DataView",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"_submission_time","filter":">",'
            '"value":"2016-01-16T00:00:00"}]',
        }

        request = self.factory.post("/", data=data, **self.extra)

        response = view(request)

        self.assertEqual(response.status_code, 201)

        data = {
            "name": "My DataView2",
            "xform": f"http://testserver/api/v1/forms/{self.xform.pk}",
            "project": f"http://testserver/api/v1/projects/{self.project.pk}",
            "columns": '["name", "age", "gender"]',
            "query": '[{"column":"_submission_time","filter":">",'
            '"value":"2016-01-16"}]',
        }

        request = self.factory.post("/", data=data, **self.extra)

        response = view(request)

        self.assertEqual(response.status_code, 201)

    def test_search_dataview_data(self):
        self._create_dataview()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        data = {"query": "Fred"}
        request = self.factory.get("/", data=data, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(1, len(response.data))
        self.assertEqual("Fred", response.data[0].get("name"))

        data = {"query": '{"age": 22}'}
        request = self.factory.get("/", data=data, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(1, len(response.data))
        self.assertEqual(22, response.data[0].get("age"))

        data = {"query": '{"age": {"$gte": 30}}'}
        request = self.factory.get("/", data=data, **self.extra)
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 200)

        self.assertEqual(1, len(response.data))
        self.assertEqual(45, response.data[0].get("age"))

    def test_invalid_url_parameters(self):
        response = self.client.get("/api/v1/dataviews/css/ona.css/")
        self.assertEqual(response.status_code, 404)

    # pylint: disable=invalid-name,too-many-locals
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    def test_export_xls_dataview_with_date_filter(self, async_result):
        """
        Test dataview export with a date filter.
        """
        self._create_dataview()
        self._publish_xls_form_to_project()
        start_date = datetime(2014, 9, 12, tzinfo=timezone.utc)
        first_datetime = start_date.strftime(MONGO_STRFTIME)
        second_datetime = start_date + timedelta(days=1, hours=20)
        query_str = (
            '{"_submission_time": {"$gte": "'
            + first_datetime
            + '", "$lte": "'
            + second_datetime.strftime(MONGO_STRFTIME)
            + '"}}'
        )

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        request = self.factory.get(
            "/",
            data={
                "format": "xlsx",
                "force_xlsx": "true",
                "include_labels": "true",
                "query": query_str,
            },
            **self.extra,
        )
        response = view(request, pk=self.data_view.pk)
        self.assertIsNotNone(response.data)

        self.assertEqual(response.status_code, 202)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)
        workbook = load_workbook(export.full_filepath)
        workbook.iso_dates = True
        sheet_name = workbook.get_sheet_names()[0]
        main_sheet = workbook.get_sheet_by_name(sheet_name)
        self.assertIn("Gender", tuple(main_sheet.values)[1])
        self.assertEqual(len(tuple(main_sheet.values)), 3)

    def test_csv_export_dataview_date_filter(self):
        """
        Test dataview csv export with a date filter.
        """
        self._create_dataview()
        self._publish_xls_form_to_project()
        start_date = datetime(2014, 9, 12, tzinfo=timezone.utc)
        first_datetime = start_date.strftime(MONGO_STRFTIME)
        second_datetime = start_date + timedelta(days=1, hours=20)
        query_str = (
            '{"_submission_time": {"$gte": "'
            + first_datetime
            + '", "$lte": "'
            + second_datetime.strftime(MONGO_STRFTIME)
            + '"}}'
        )
        count = Export.objects.all().count()

        view = DataViewViewSet.as_view(
            {
                "get": "data",
            }
        )

        request = self.factory.get("/", data={"query": query_str}, **self.extra)
        response = view(request, pk=self.data_view.pk, format="csv")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(count + 1, Export.objects.all().count())

        headers = dict(response.items())
        self.assertEqual(headers["Content-Type"], "application/csv")
        content_disposition = headers["Content-Disposition"]
        filename = filename_from_disposition(content_disposition)
        _basename, ext = os.path.splitext(filename)
        self.assertEqual(ext, '.csv"')

        content = get_response_content(response)
        self.assertEqual(content, "name,age,gender\nDennis Wambua,28,male\n")

    # pylint: disable=too-many-locals
    @override_settings(CELERY_TASK_ALWAYS_EAGER=True)
    @patch("onadata.apps.api.viewsets.dataview_viewset.AsyncResult")
    @flaky(max_runs=10)
    def test_csv_export_async_dataview_date_filter(self, async_result):
        """
        Test dataview csv export async with a date filter.
        """
        self._create_dataview()
        self._publish_xls_form_to_project()
        start_date = datetime(2014, 9, 12, tzinfo=timezone.utc)
        first_datetime = start_date.strftime(MONGO_STRFTIME)
        second_datetime = start_date + timedelta(days=1, hours=20)
        query_str = (
            '{"_submission_time": {"$gte": "'
            + first_datetime
            + '", "$lte": "'
            + second_datetime.strftime(MONGO_STRFTIME)
            + '"}}'
        )
        count = Export.objects.all().count()

        view = DataViewViewSet.as_view(
            {
                "get": "export_async",
            }
        )

        request = self.factory.get(
            "/", data={"format": "csv", "query": query_str}, **self.extra
        )
        response = view(request, pk=self.data_view.pk)
        self.assertEqual(response.status_code, 202)
        self.assertIsNotNone(response.data)
        self.assertTrue("job_uuid" in response.data)
        task_id = response.data.get("job_uuid")
        self.assertEqual(count + 1, Export.objects.all().count())

        export_pk = Export.objects.all().order_by("pk").reverse()[0].pk

        # metaclass for mocking results
        job = type("AsyncResultMock", (), {"state": "SUCCESS", "result": export_pk})
        async_result.return_value = job

        get_data = {"job_uuid": task_id}
        request = self.factory.get("/", data=get_data, **self.extra)
        response = view(request, pk=self.data_view.pk)

        self.assertIn("export_url", response.data)

        self.assertTrue(async_result.called)
        self.assertEqual(response.status_code, 202)
        export = Export.objects.get(task_id=task_id)
        self.assertTrue(export.is_successful)
        with open(export.full_filepath, encoding="utf-8") as csv_file:
            self.assertEqual(
                csv_file.read(), "name,age,gender\nDennis Wambua,28,male\n"
            )
